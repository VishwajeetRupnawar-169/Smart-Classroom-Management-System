"""
Complete Attendance System with PRN tracking, auto-detection, and web preview
Save this as: attendance_system_complete.py
"""

from flask import Flask, jsonify, request, render_template_string, Response, send_file
from flask_cors import CORS
from pymongo import MongoClient
from datetime import datetime, timedelta
import os, threading, cv2, numpy as np, face_recognition, sys, io, traceback
from time import sleep
from threading import Lock, Event
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import time
from openpyxl.styles import Font, PatternFill

sys.path.append(os.path.abspath('../'))
try:
    from Excel_Format import get_current_session
except:
    def get_current_session():
        h = datetime.now().hour
        return f"Session {min(max((h - 9) + 1, 1), 8)}"

app = Flask(__name__)
CORS(app)

MONGODB_CONFIG = {'host': 'localhost', 'port': 27017, 'database': 'Attendance_system'}
TEMPLATE_FILE = 'Book2.xlsx'
ALL_SESSIONS = [f"Session {i}" for i in range(1, 9)]
YEAR_MAPPING = {'2022': 'B.Tech', '2023': 'TY', '2024': 'SY', '2025': 'FY'}

attendance_system = None
camera_running = False

class FilePathResolver:
    @staticmethod
    def find_file(filename):
        dirs = ['.', '..', '../..', os.path.dirname(__file__), 
                os.path.join(os.path.dirname(__file__), '..')]
        for d in dirs:
            p = os.path.join(d, filename)
            if os.path.exists(p):
                return os.path.abspath(p)
        raise FileNotFoundError(f"File '{filename}' not found")
    
    @staticmethod
    def find_training_folder(dept_year, mode_name):
        bases = ['Training_images', '../Training_images', 
                 os.path.join(os.path.dirname(__file__), 'Training_images')]
        for b in bases:
            for p in [os.path.join(b, dept_year, mode_name), os.path.join(b, mode_name)]:
                if os.path.exists(p):
                    return os.path.abspath(p)
        raise FileNotFoundError(f"Training folder not found: {dept_year}/{mode_name}")

class DatabaseManager:
    def __init__(self, config):
        self.config = config
        self.client = None
        self.db = None
        self.lock = Lock()
        self._init()
    
    def _init(self):
        self.client = MongoClient(self.config['host'], self.config['port'], serverSelectionTimeoutMS=5000)
        self.db = self.client[self.config['database']]
        try:
            self.db.lecture_metadata.create_index([('collection_name', 1)], unique=True)
        except:
            pass
    
    def _get_year_code(self, year):
        return YEAR_MAPPING.get(str(year), 'B.Tech')
    
    def load_students_from_excel(self, file, sheet):
        path = FilePathResolver.find_file(file)
        wb = load_workbook(path, data_only=True)
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found")
        s = wb[sheet]
        data, headers = [], None
        for row in s.iter_rows(values_only=True):
            if all(c is None or str(c).strip() == '' for c in row):
                continue
            r = [c if c is not None else '' for c in row]
            if headers is None:
                headers = [str(h).strip() for h in r]
            else:
                data.append(r)
        wb.close()
        return headers, data
    
    # Update the create_or_get_daily_collection method to handle PRN better
    def create_or_get_daily_collection(self, dept, year, date, room, teacher, headers, data, cams=None):
        yc = self._get_year_code(year)
        cn = f"{dept}_{yc}_{date}"
        with self.lock:
            if cn in self.db.list_collection_names():
                return cn
            
            col = self.db[cn]
            ct = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            sessions = {}
            
            prn_count = 0
            roll_fallback_count = 0
            skipped_count = 0
            
            for sn in ALL_SESSIONS:
                students = {}
                for row in data:
                    doc = {headers[i]: str(row[i]).strip() if i < len(row) and row[i] not in (None, '') else ''
                        for i in range(len(headers))}
                    
                    # Extract PRN with multiple variations
                    prn = ''
                    for k, v in doc.items():
                        k_lower = str(k).strip().lower()
                        if k_lower in ['prn no.', 'prn no', 'prn_no', 'prn', 'prnno', 'prn number']:
                            prn = str(v).strip()
                            if prn:
                                break
                    
                    # Extract Roll Number
                    roll_no = ''
                    for k, v in doc.items():
                        k_lower = str(k).strip().lower()
                        if k_lower in ['roll no.', 'roll no', 'roll_no', 'rollno', 'roll number']:
                            roll_no = str(v).strip()
                            if roll_no:
                                break
                    
                    # Determine identifier to use
                    if prn:
                        identifier = prn
                        using_roll = False
                        prn_count += 1
                    elif roll_no:
                        identifier = roll_no
                        using_roll = True
                        roll_fallback_count += 1
                    else:
                        skipped_count += 1
                        print(f"⚠️  Skipping student - no PRN or Roll: {doc.get('Name', 'Unknown')}")
                        continue
                    
                    students[identifier] = {
                        'prn_no': prn if prn else roll_no,
                        'roll_no': roll_no,
                        'name': doc.get('Name', ''),
                        'status': 'Absent',
                        'timestamps': {
                            'first_seen': None,
                            'last_seen': None,
                            'present_timer_start': None,
                            'absence_timer_start': None,
                            'last_updated': ct
                        },
                        'durations': {
                            'total_present_seconds': 0,
                            'total_absent_seconds': 0,
                            'total_present_human': '0 sec',
                            'total_absent_human': '0 sec'
                        },
                        'flags': {
                            'manual_override': False,
                            'is_temp_absent': False,
                            'is_perm_absent': False,
                            'using_roll_as_prn': using_roll
                        }
                    }
                
                sessions[sn] = {
                    'start_time': None,
                    'end_time': None,
                    'students': students
                }
            
            print(f"\n📊 Collection Created: {cn}")
            print(f"   ✅ Students with PRN: {prn_count}")
            if roll_fallback_count > 0:
                print(f"   ⚠️  Students using Roll as ID: {roll_fallback_count}")
            if skipped_count > 0:
                print(f"   ❌ Students skipped (no ID): {skipped_count}")
            print(f"   📝 Total loaded: {len(sessions['Session 1']['students'])}\n")
            
            col.insert_one({
                'date': date, 'department': dept, 'year': year, 'year_code': yc,
                'classroom': room, 'teacher_name': teacher, 'camera_ids': cams or [],
                'created_at': datetime.now(), 'sessions': sessions,
                'metadata': {
                    'prn_count': prn_count,
                    'roll_fallback_count': roll_fallback_count,
                    'skipped_count': skipped_count
                }
            })
            
            self.db.lecture_metadata.insert_one({
                'collection_name': cn, 'date': date, 'department': dept, 'year': year,
                'year_code': yc, 'classroom': room, 'teacher_name': teacher,
                'created_at': datetime.now()
            })
            
            return cn
    
    def find_prn_by_identifier(self, cn, sn, ident):
        with self.lock:
            doc = self.db[cn].find_one({})
            if not doc or sn not in doc.get('sessions', {}):
                return None
            for prn, stu in doc['sessions'][sn]['students'].items():
                if stu.get('roll_no', '').upper() == ident.upper() or stu.get('name', '').upper() == ident.upper():
                    return prn
            return None
    
    def update_student_attendance(self, cn, sn, prn, status, manual=False):
        with self.lock:
            doc = self.db[cn].find_one({})
            if not doc or sn not in doc.get('sessions', {}) or prn not in doc['sessions'][sn]['students']:
                return False
            sp = f'sessions.{sn}.students.{prn}'
            stu = doc['sessions'][sn]['students'][prn]
            ct = datetime.now()
            cts = ct.strftime('%Y-%m-%d %H:%M:%S')
            ps = stu.get('status', 'Absent')
            pstart = stu['timestamps'].get('present_timer_start')
            astart = stu['timestamps'].get('absence_timer_start')
            tp = stu['durations'].get('total_present_seconds', 0)
            ta = stu['durations'].get('total_absent_seconds', 0)
            upd = {}
            if ps == 'Present' and status != 'Present':
                if pstart:
                    try:
                        tp += (ct - datetime.strptime(pstart, '%Y-%m-%d %H:%M:%S')).total_seconds()
                    except:
                        pass
                upd[f'{sp}.timestamps.present_timer_start'] = None
                upd[f'{sp}.timestamps.absence_timer_start'] = cts
            elif ps != 'Present' and status == 'Present':
                if astart:
                    try:
                        ta += (ct - datetime.strptime(astart, '%Y-%m-%d %H:%M:%S')).total_seconds()
                    except:
                        pass
                upd[f'{sp}.timestamps.absence_timer_start'] = None
                upd[f'{sp}.timestamps.present_timer_start'] = cts
            elif status == 'Present' and not pstart:
                upd[f'{sp}.timestamps.present_timer_start'] = cts
            elif status != 'Present' and not astart:
                upd[f'{sp}.timestamps.absence_timer_start'] = cts
            upd[f'{sp}.status'] = status
            upd[f'{sp}.timestamps.last_updated'] = cts
            upd[f'{sp}.flags.manual_override'] = manual
            if stu['timestamps']['first_seen'] is None and status == 'Present':
                upd[f'{sp}.timestamps.first_seen'] = cts
            if status == 'Present':
                upd[f'{sp}.timestamps.last_seen'] = ct.strftime('%H:%M:%S')
            upd[f'{sp}.durations.total_present_seconds'] = int(tp)
            upd[f'{sp}.durations.total_absent_seconds'] = int(ta)
            upd[f'{sp}.durations.total_present_human'] = self._fmt(tp)
            upd[f'{sp}.durations.total_absent_human'] = self._fmt(ta)
            if doc['sessions'][sn]['start_time'] is None:
                upd[f'sessions.{sn}.start_time'] = cts
            self.db[cn].update_one({}, {'$set': upd})
            return True
    def batch_update_attendance(self, cn, sn, updates_dict):
        """
        Batch update multiple students' attendance
        updates_dict: {prn: {'status': 'Present', 'timestamp': datetime}, ...}
        """
        with self.lock:
            doc = self.db[cn].find_one({})
            if not doc or sn not in doc.get('sessions', {}):
                return 0
            
            ct = datetime.now()
            bulk_updates = {}
            success_count = 0
            
            for prn, update_info in updates_dict.items():
                if prn not in doc['sessions'][sn]['students']:
                    continue
                
                sp = f'sessions.{sn}.students.{prn}'
                stu = doc['sessions'][sn]['students'][prn]
                status = update_info['status']
                cts = ct.strftime('%Y-%m-%d %H:%M:%S')
                
                ps = stu.get('status', 'Absent')
                pstart = stu['timestamps'].get('present_timer_start')
                astart = stu['timestamps'].get('absence_timer_start')
                tp = stu['durations'].get('total_present_seconds', 0)
                ta = stu['durations'].get('total_absent_seconds', 0)
                
                # Timer logic
                if ps == 'Present' and status != 'Present':
                    if pstart:
                        try:
                            tp += (ct - datetime.strptime(pstart, '%Y-%m-%d %H:%M:%S')).total_seconds()
                        except:
                            pass
                    bulk_updates[f'{sp}.timestamps.present_timer_start'] = None
                    bulk_updates[f'{sp}.timestamps.absence_timer_start'] = cts
                elif ps != 'Present' and status == 'Present':
                    if astart:
                        try:
                            ta += (ct - datetime.strptime(astart, '%Y-%m-%d %H:%M:%S')).total_seconds()
                        except:
                            pass
                    bulk_updates[f'{sp}.timestamps.absence_timer_start'] = None
                    bulk_updates[f'{sp}.timestamps.present_timer_start'] = cts
                elif status == 'Present' and not pstart:
                    bulk_updates[f'{sp}.timestamps.present_timer_start'] = cts
                elif status != 'Present' and not astart:
                    bulk_updates[f'{sp}.timestamps.absence_timer_start'] = cts
                
                # Update status and timestamps
                bulk_updates[f'{sp}.status'] = status
                bulk_updates[f'{sp}.timestamps.last_updated'] = cts
                bulk_updates[f'{sp}.flags.manual_override'] = False
                
                if stu['timestamps']['first_seen'] is None and status == 'Present':
                    bulk_updates[f'{sp}.timestamps.first_seen'] = cts
                
                if status == 'Present':
                    bulk_updates[f'{sp}.timestamps.last_seen'] = ct.strftime('%H:%M:%S')
                
                bulk_updates[f'{sp}.durations.total_present_seconds'] = int(tp)
                bulk_updates[f'{sp}.durations.total_absent_seconds'] = int(ta)
                bulk_updates[f'{sp}.durations.total_present_human'] = self._fmt(tp)
                bulk_updates[f'{sp}.durations.total_absent_human'] = self._fmt(ta)
                
                success_count += 1
            
            # Single database update for all students
            if bulk_updates:
                if doc['sessions'][sn]['start_time'] is None:
                    bulk_updates[f'sessions.{sn}.start_time'] = ct.strftime('%Y-%m-%d %H:%M:%S')
                
                self.db[cn].update_one({}, {'$set': bulk_updates})
            
            return success_count


    def _fmt(self, sec):
        s = int(sec)
        if s < 60:
            return f"{s} sec"
        elif s < 3600:
            return f"{s//60} min {s%60} sec"
        else:
            return f"{s//3600} hr {(s%3600)//60} min"
    
    def get_session_attendance(self, cn, sn):
        with self.lock:
            doc = self.db[cn].find_one({})
            if not doc or sn not in doc.get('sessions', {}):
                return []
            return list(doc['sessions'][sn]['students'].values())
    
    def get_session_summary(self, cn, sn):
        with self.lock:
            doc = self.db[cn].find_one({})
            if not doc or sn not in doc.get('sessions', {}):
                return {}
            stu = doc['sessions'][sn]['students']
            t = len(stu)
            p = sum(1 for s in stu.values() if s['status'] == 'Present')
            return {
                'total': t, 'present': p,
                'temporary_absent': sum(1 for s in stu.values() if s['status'] == 'Temporary Absent'),
                'permanently_absent': sum(1 for s in stu.values() if s['status'] == 'Permanently Absent'),
                'absent': sum(1 for s in stu.values() if s['status'] == 'Absent'),
                'attendance_percentage': round((p / t * 100), 2) if t > 0 else 0
            }
    
    def get_all_daily_collections(self):
        with self.lock:
            cols = list(self.db.lecture_metadata.find().sort('created_at', -1))
            for c in cols:
                c['_id'] = str(c['_id'])
                if 'created_at' in c:
                    c['created_at'] = c['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            return cols
    
    def get_student_history(self, ident, field='prn_no'):
        with self.lock:
            metas = list(self.db.lecture_metadata.find().sort('created_at', -1))
            hist = []
            for m in metas:
                doc = self.db[m['collection_name']].find_one({})
                if not doc:
                    continue
                for sn, sd in doc.get('sessions', {}).items():
                    stu = None
                    stus = sd.get('students', {})
                    if field == 'prn_no':
                        stu = stus.get(ident)
                    else:
                        for s in stus.values():
                            if (field == 'roll_no' and s.get('roll_no', '').upper() == ident.upper()) or \
                               (field == 'name' and s.get('name', '').upper() == ident.upper()):
                                stu = s
                                break
                    if stu:
                        hist.append({
                            'date': m['date'], 'session': sn, 'status': stu.get('status', 'N/A'),
                            'first_seen': stu.get('timestamps', {}).get('first_seen', 'N/A'),
                            'last_seen': stu.get('timestamps', {}).get('last_seen', 'N/A'),
                            'present_duration': stu.get('durations', {}).get('total_present_human', '0 sec'),
                            'department': m.get('department', ''), 'classroom': m.get('classroom', ''),
                            'prn_no': stu.get('prn_no', ''), 'roll_no': stu.get('roll_no', ''),
                            'name': stu.get('name', '')
                        })
            return hist
    
    def clear_session_data(self, cn, sn):
        with self.lock:
            doc = self.db[cn].find_one({})
            if not doc or sn not in doc.get('sessions', {}):
                return 0
            cts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            upd = {}
            for prn in doc['sessions'][sn]['students'].keys():
                p = f'sessions.{sn}.students.{prn}'
                upd.update({f'{p}.status': 'Absent', f'{p}.timestamps.first_seen': None,
                           f'{p}.timestamps.last_seen': None, f'{p}.timestamps.present_timer_start': None,
                           f'{p}.timestamps.absence_timer_start': None, f'{p}.timestamps.last_updated': cts,
                           f'{p}.durations.total_present_seconds': 0, f'{p}.durations.total_absent_seconds': 0,
                           f'{p}.durations.total_present_human': '0 sec', f'{p}.durations.total_absent_human': '0 sec',
                           f'{p}.flags.manual_override': False, f'{p}.flags.is_temp_absent': False,
                           f'{p}.flags.is_perm_absent': False})
            self.db[cn].update_one({}, {'$set': upd})
            return len(doc['sessions'][sn]['students'])
    
    def generate_excel_report(self, cn, sn=None):
        try:
            wb = Workbook()
            wb.remove(wb.active)
            doc = self.db[cn].find_one({})
            if not doc:
                return None
            summ = wb.create_sheet("Summary", 0)
            summ['A1'] = f"Attendance Report - {doc['date']}"
            summ['A1'].font = Font(bold=True, size=14)
            summ['A2'] = f"Department: {doc['department']}"
            summ['A3'] = f"Classroom: {doc['classroom']}"
            summ['A4'] = f"Teacher: {doc['teacher_name']}"
            for sess in ([sn] if sn else ALL_SESSIONS):
                if sess not in doc['sessions']:
                    continue
                sh = wb.create_sheet(sess[:31])
                stus = list(doc['sessions'][sess]['students'].values())
                hdrs = ['PRN No', 'Roll No', 'Name', 'Status', 'First Seen', 'Last Seen', 
                        'Present Duration', 'Absent Duration']
                for c, h in enumerate(hdrs, 1):
                    cell = sh.cell(1, c, h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
                for r, s in enumerate(stus, 2):
                    sh.cell(r, 1, s['prn_no'])
                    sh.cell(r, 2, s['roll_no'])
                    sh.cell(r, 3, s['name'])
                    sh.cell(r, 4, s['status'])
                    sh.cell(r, 5, s['timestamps'].get('first_seen', 'N/A'))
                    sh.cell(r, 6, s['timestamps'].get('last_seen', 'N/A'))
                    sh.cell(r, 7, s['durations'].get('total_present_human', '0 sec'))
                    sh.cell(r, 8, s['durations'].get('total_absent_human', '0 sec'))
            ef = io.BytesIO()
            wb.save(ef)
            ef.seek(0)
            return ef
        except:
            return None
    
    def get_session_data_for_preview(self, cn, sn):
        with self.lock:
            doc = self.db[cn].find_one({})
            if not doc or sn not in doc.get('sessions', {}):
                return None
            stus = list(doc['sessions'][sn]['students'].values())
            t = len(stus)
            p = sum(1 for s in stus if s['status'] == 'Present')
            return {
                'date': doc['date'], 'department': doc['department'], 'classroom': doc['classroom'],
                'teacher_name': doc['teacher_name'], 'session_name': sn,
                'summary': {
                    'total': t, 'present': p,
                    'temporary_absent': sum(1 for s in stus if s['status'] == 'Temporary Absent'),
                    'permanently_absent': sum(1 for s in stus if s['status'] == 'Permanently Absent'),
                    'absent': sum(1 for s in stus if s['status'] == 'Absent'),
                    'attendance_percentage': round((p / t * 100), 2) if t > 0 else 0
                },
                'students': stus
            }
    
    def close(self):
        if self.client:
            self.client.close()

class AttendanceSystem:
    def __init__(self, mode, year, dept, room, teacher, cams=None):
        self.mode = mode
        self.student_status = {}
        self.stop_event = Event()
        self.attendance_count = 0
        self.current_faces_count = 0
        self.year = year
        self.dept = dept
        self.room = room
        self.teacher = teacher
        self.cams = cams or ['CAM-01']
        self.db = DatabaseManager(MONGODB_CONFIG)
        yc = self.db._get_year_code(year)
        sheet = f"{dept}_{yc}"
        self.headers, self.data = self.db.load_students_from_excel(TEMPLATE_FILE, sheet)
        self.total_students = len(self.data)
        self.current_session = None
        self.current_collection = None
        self.current_date = None
        self.class_names, self.encodings = self._load_training(f"{dept}_{yc}")
        self.attendance_queue = defaultdict(dict)  # NEW: Queue for batched updates
        self.queue_lock = Lock()  # NEW: Lock for queue access
    
    def _load_training(self, dy):
        path = FilePathResolver.find_training_folder(dy, 'Name' if self.mode == 1 else 'Roll No.')
        imgs, names = [], []
        for f in os.listdir(path):
            if f.lower().endswith(('.jpg', '.jpeg', '.png')):
                img = cv2.imread(os.path.join(path, f))
                if img is not None:
                    imgs.append(img)
                    names.append(os.path.splitext(f)[0])
        encs = []
        for img in imgs:
            try:
                rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                e = face_recognition.face_encodings(rgb)
                if e:
                    encs.append(e[0])
            except:
                pass
        return names, encs
    
    def mark_attendance(self, ident):
            """Queue attendance marking for batch processing"""
            ident = ident.upper()
            ct = datetime.now()
            
            if not self.current_session or not self.current_collection:
                return False
            
            prn = self.db.find_prn_by_identifier(self.current_collection, self.current_session, ident)
            if not prn:
                return False
            
            # Queue the update instead of immediate DB write
            with self.queue_lock:
                if prn not in self.student_status:
                    self.student_status[prn] = {
                        'last_seen': ct,
                        'status': 'Present',
                        'timer_start': None,
                        'first_detection': ct
                    }
                    self.attendance_queue[prn] = {
                        'status': 'Present',
                        'timestamp': ct
                    }
                else:
                    self.student_status[prn]['last_seen'] = ct
                    self.student_status[prn]['status'] = 'Present'
                    self.student_status[prn]['timer_start'] = None
            
            return True
    def process_attendance_queue(self):
        """Process queued attendance updates in batch"""
        while not self.stop_event.is_set():
            try:
                if self.attendance_queue and self.current_collection and self.current_session:
                    with self.queue_lock:
                        # Get all queued updates
                        updates = dict(self.attendance_queue)
                        self.attendance_queue.clear()
                    
                    if updates:
                        # Batch update to database
                        success_count = self.db.batch_update_attendance(
                            self.current_collection,
                            self.current_session,
                            updates
                        )
                        
                        if success_count > 0:
                            self.attendance_count = len([s for s in self.student_status.values() 
                                                        if s['status'] == 'Present'])
                            print(f"✅ Batch updated {success_count} students")
                
                sleep(0.5)  # Process queue every 0.5 seconds
            except Exception as e:
                print(f"Queue processing error: {e}")
                sleep(0.5)
        
    def check_absence_continuously(self):
        while not self.stop_event.is_set():
            try:
                if self.current_session:
                    ct = datetime.now()
                    for prn, info in list(self.student_status.items()):
                        diff = ct - info['last_seen']
                        if info['status'] == 'Present':
                            if diff >= timedelta(seconds=5):
                                if info['timer_start'] is None:
                                    self.student_status[prn]['timer_start'] = ct
                                else:
                                    ta = ct - info['timer_start']
                                    if ta >= timedelta(seconds=15):
                                        self.student_status[prn]['status'] = 'Permanently Absent'
                                        self.db.update_student_attendance(self.current_collection, 
                                                                         self.current_session, prn, 
                                                                         'Permanently Absent')
                                    elif ta >= timedelta(seconds=10):
                                        if info['status'] != 'Temporary Absent':
                                            self.student_status[prn]['status'] = 'Temporary Absent'
                                            self.db.update_student_attendance(self.current_collection, 
                                                                             self.current_session, prn, 
                                                                             'Temporary Absent')
                            else:
                                self.db.update_student_attendance(self.current_collection, 
                                                                 self.current_session, prn, 
                                                                 'Present', manual=False)
                sleep(2)
            except:
                sleep(2)
    
    def process_frame(self, frame):
        try:
            small = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            rgb = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)
            locs = face_recognition.face_locations(rgb)
            encs = face_recognition.face_encodings(rgb, locs)
            
            self.current_faces_count = len(locs)
            
            sess = get_current_session()
            date = datetime.now().strftime('%Y-%m-%d')
            
            if sess != self.current_session or date != self.current_date:
                self.current_session = sess
                self.current_date = date
                self.student_status = {}
                self.attendance_count = 0
                with self.queue_lock:
                    self.attendance_queue.clear()
                
                self.current_collection = self.db.create_or_get_daily_collection(
                    self.dept, self.year, date, self.room, self.teacher,
                    self.headers, self.data, self.cams
                )
            
            # Process all detected faces
            detected_this_frame = []
            for enc, loc in zip(encs, locs):
                matches = face_recognition.compare_faces(self.encodings, enc, tolerance=0.6)
                dists = face_recognition.face_distance(self.encodings, enc)
                
                if len(dists) > 0:
                    idx = np.argmin(dists)
                    if matches[idx] and dists[idx] < 0.6:  # Additional distance check
                        name = self.class_names[idx].upper()
                        color = (0, 255, 0)
                        self.mark_attendance(name)  # Queue the update
                        detected_this_frame.append(name)
                    else:
                        name = "UNKNOWN"
                        color = (0, 0, 255)
                else:
                    name = "UNKNOWN"
                    color = (0, 0, 255)
                
                # Draw bounding box
                y1, x2, y2, x1 = [c * 4 for c in loc]
                cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                cv2.rectangle(frame, (x1, y2 - 35), (x2, y2), color, cv2.FILLED)
                cv2.putText(frame, name, (x1 + 6, y2 - 6), 
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
            
            # Draw overlay with stats
            overlay = frame.copy()
            cv2.rectangle(overlay, (0, 0), (frame.shape[1], 100), (0, 0, 0), -1)
            cv2.addWeighted(overlay, 0.7, frame, 0.3, 0, frame)
            
            cv2.putText(frame, f"Session: {sess}", (20, 25), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
            cv2.putText(frame, f"Attendance: {self.attendance_count}/{self.total_students}",
                       (20, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.6,
                       (0, 255, 0) if self.attendance_count > 0 else (255, 255, 255), 2)
            cv2.putText(frame, f"Faces Detected: {self.current_faces_count}",
                       (20, 75), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)
            
            # Show currently detected students
            if detected_this_frame:
                cv2.putText(frame, f"Detecting: {', '.join(detected_this_frame[:3])}",
                           (20, 95), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
            
            return frame
        except Exception as e:
            print(f"Frame processing error: {e}")
            return frame

    def stop(self):
        self.stop_event.set()

@app.route('/api/health')
def health():
    try:
        c = MongoClient(MONGODB_CONFIG['host'], MONGODB_CONFIG['port'], serverSelectionTimeoutMS=2000)
        c.server_info()
        c.close()
        dbc = True
    except:
        dbc = False
    return jsonify({'status': 'healthy' if dbc else 'degraded', 'database': 'connected' if dbc else 'disconnected',
                    'camera_status': 'running' if camera_running else 'stopped', 'system_initialized': attendance_system is not None})

@app.route('/api/test-excel')
def test_excel():
    """Test endpoint to verify Excel file and list all sheets"""
    try:
        result = {
            'excel_file': TEMPLATE_FILE,
            'file_found': False,
            'file_path': None,
            'sheets': [],
            'error': None
        }
        
        try:
            excel_path = FilePathResolver.find_file(TEMPLATE_FILE)
            result['file_found'] = True
            result['file_path'] = excel_path
            
            wb = load_workbook(excel_path, data_only=True)
            result['sheets'] = wb.sheetnames
            
            # Get sample data from each sheet
            sheets_info = {}
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheets_info[sheet_name] = {
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column,
                    'has_data': sheet.max_row > 1
                }
            
            result['sheets_info'] = sheets_info
            wb.close()
            
        except FileNotFoundError as e:
            result['error'] = f"File not found: {str(e)}"
        except Exception as e:
            result['error'] = f"Error: {str(e)}"
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/preview-config', methods=['POST'])
def preview_config():
    try:
        d = request.get_json()
        print(f"\n{'#'*70}")
        print(f"API ENDPOINT: /api/preview-config")
        print(f"{'#'*70}")
        print(f"Request data: {d}")
        
        y, dept, room, teach = d.get('year'), d.get('department'), d.get('classroom'), d.get('teacher_name')
        
        if not all([y, dept, room, teach]):
            print("❌ Missing required fields")
            return jsonify({'success': False, 'message': 'All fields required'}), 400
        
        print(f"\nConfiguration:")
        print(f"  Year: {y}")
        print(f"  Department: {dept}")
        print(f"  Classroom: {room}")
        print(f"  Teacher: {teach}")
        
        db = DatabaseManager(MONGODB_CONFIG)
        yc = db._get_year_code(y)
        sheet = f"{dept}_{yc}"
        
        print(f"\nTarget sheet: {sheet}")
        print(f"Year code: {yc}")
        
        try:
            print(f"\nAttempting to load Excel...")
            h, data = db.load_students_from_excel(TEMPLATE_FILE, sheet)
            print(f"✅ Excel loaded successfully!")
            print(f"   Headers: {len(h)} columns")
            print(f"   Data: {len(data)} rows")
            
            if len(data) == 0:
                print("⚠️  WARNING: No data rows found!")
                db.close()
                return jsonify({
                    'success': False, 
                    'message': f'Sheet "{sheet}" is empty or contains no valid data rows',
                    'error': 'No student data found'
                }), 400
                
        except Exception as e:
            print(f"❌ Failed to load Excel!")
            print(f"Error: {e}")
            traceback.print_exc()
            db.close()
            return jsonify({
                'success': False, 
                'message': f'Failed to load sheet: {sheet}', 
                'error': str(e),
                'details': 'Check if Book2.xlsx exists and has the correct sheet name'
            }), 400
        
        sess = get_current_session()
        date = datetime.now().strftime('%Y-%m-%d')
        
        print(f"\nCreating/getting collection...")
        print(f"  Session: {sess}")
        print(f"  Date: {date}")
        
        cn = db.create_or_get_daily_collection(dept, y, date, room, teach, h, data, d.get('camera_ids', ['CAM-01']))
        
        print(f"\nFetching attendance data...")
        att = db.get_session_attendance(cn, sess)
        summ = db.get_session_summary(cn, sess)
        
        print(f"\n{'='*70}")
        print(f"RESULT:")
        print(f"{'='*70}")
        print(f"Attendance records: {len(att)}")
        print(f"Summary: {summ}")
        print(f"{'='*70}\n")
        
        db.close()
        
        return jsonify({
            'success': True, 
            'year_code': yc, 
            'sheet_loaded': sheet,
            'collection_name': cn, 
            'session_name': sess, 
            'date': date, 
            'summary': summ, 
            'attendance': att
        })
        
    except Exception as e:
        print(f"\n❌ UNEXPECTED ERROR in preview_config:")
        print(f"Error: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False, 
            'error': str(e),
            'message': 'Internal server error. Check server logs for details.'
        }), 500

# Update the start_camera endpoint to start the queue processor
@app.route('/api/camera/start', methods=['POST'])
def start_camera():
    global attendance_system, camera_running
    try:
        d = request.get_json()
        if camera_running:
            return jsonify({'success': False, 'message': 'Camera already running'}), 400
        
        attendance_system = AttendanceSystem(
            d.get('mode', 1), d['year'], d['department'],
            d['classroom'], d['teacher_name'], d.get('camera_ids', ['CAM-01'])
        )
        
        # Start background threads
        threading.Thread(target=attendance_system.check_absence_continuously, daemon=True).start()
        threading.Thread(target=attendance_system.process_attendance_queue, daemon=True).start()  # NEW
        
        camera_running = True
        sess = get_current_session()
        date = datetime.now().strftime('%Y-%m-%d')
        cn = attendance_system.db.create_or_get_daily_collection(
            d['department'], d['year'], date,
            d['classroom'], d['teacher_name'],
            attendance_system.headers, attendance_system.data,
            d.get('camera_ids', ['CAM-01'])
        )
        
        attendance_system.current_collection = cn
        attendance_system.current_session = sess
        attendance_system.current_date = date
        
        print(f"🎥 Camera started - Multi-face detection enabled")
        print(f"📊 Ready to track {attendance_system.total_students} students")
        
        return jsonify({
            'success': True,
            'message': 'Camera started with multi-face detection',
            'year_code': YEAR_MAPPING.get(d['year'], 'B.Tech'),
            'sheet_loaded': f"{d['department']}_{YEAR_MAPPING.get(d['year'], 'B.Tech')}",
            'initial_data': {
                'collection_name': cn,
                'session_name': sess,
                'date': date,
                'summary': attendance_system.db.get_session_summary(cn, sess),
                'attendance': attendance_system.db.get_session_attendance(cn, sess)
            }
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/camera/stop', methods=['POST'])
def stop_camera():
    global attendance_system, camera_running
    try:
        if not camera_running:
            return jsonify({'success': False, 'message': 'Camera not running'}), 400
        camera_running = False
        if attendance_system:
            attendance_system.stop()
        return jsonify({'success': True, 'message': 'Camera stopped'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/camera/status')
def camera_status():
    global camera_running, attendance_system
    status = {'running': camera_running, 'attendance_count': 0, 'current_faces': 0,
              'current_session': None, 'current_collection': None, 'camera_ids': []}
    if attendance_system:
        status.update({'attendance_count': attendance_system.attendance_count,
                      'current_faces': attendance_system.current_faces_count,
                      'current_session': attendance_system.current_session,
                      'current_collection': attendance_system.current_collection,
                      'camera_ids': attendance_system.cams})
    return jsonify(status)

def generate_frames():
    global attendance_system, camera_running
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        return
    try:
        while camera_running:
            ret, frame = cap.read()
            if not ret:
                break
            if attendance_system:
                frame = attendance_system.process_frame(frame)
            ret, buffer = cv2.imencode('.jpg', frame)
            if not ret:
                continue
            yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
    except Exception as e:
        print(f"Video error: {e}")
    finally:
        cap.release()

@app.route('/api/video_feed')
def video_feed():
    if not camera_running:
        return jsonify({'error': 'Camera not running'}), 400
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/api/current-session')
def get_current_session_data():
    try:
        print(f"Current session request - attendance_system exists: {attendance_system is not None}")
        if attendance_system and attendance_system.current_collection and attendance_system.current_session:
            print(f"Getting data for: {attendance_system.current_collection}, {attendance_system.current_session}")
            att = attendance_system.db.get_session_attendance(attendance_system.current_collection, 
                                                              attendance_system.current_session)
            summ = attendance_system.db.get_session_summary(attendance_system.current_collection, 
                                                           attendance_system.current_session)
            print(f"Returning {len(att)} records")
            return jsonify({'success': True, 'active': True,
                          'collection_name': attendance_system.current_collection,
                          'session_name': attendance_system.current_session,
                          'date': attendance_system.current_date,
                          'summary': summ,
                          'attendance': att})
        print("No active attendance system")
        return jsonify({'success': True, 'active': False})
    except Exception as e:
        print(f"Current session error: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/attendance/update', methods=['POST'])
def update_attendance_manual():
    try:
        d = request.get_json()
        print(f"\n{'='*60}")
        print(f"MANUAL ATTENDANCE UPDATE")
        print(f"{'='*60}")
        print(f"Request data: {d}")
        
        collection_name = d.get('collection_name')
        session_name = d.get('session_name')
        prn_no = d.get('prn_no', '').strip()
        status = d.get('status')
        
        print(f"Collection: {collection_name}")
        print(f"Session: {session_name}")
        print(f"PRN: '{prn_no}' (length: {len(prn_no)})")
        print(f"New Status: {status}")
        
        # Detailed validation
        if not collection_name:
            print("❌ Missing collection_name")
            return jsonify({'success': False, 'error': 'Missing collection name'}), 400
        if not session_name:
            print("❌ Missing session_name")
            return jsonify({'success': False, 'error': 'Missing session name'}), 400
        if not prn_no:
            print("❌ Missing or empty prn_no")
            return jsonify({'success': False, 'error': 'Missing PRN number'}), 400
        if not status:
            print("❌ Missing status")
            return jsonify({'success': False, 'error': 'Missing status'}), 400
        
        db = DatabaseManager(MONGODB_CONFIG)
        
        print(f"Updating attendance...")
        success = db.update_student_attendance(collection_name, session_name, prn_no, status, manual=True)
        
        db.close()
        
        if success:
            print(f"✅ Successfully updated PRN {prn_no} to {status}")
            return jsonify({'success': True, 'message': f'Marked as {status}'})
        else:
            print(f"❌ Failed to update - student not found")
            return jsonify({'success': False, 'error': 'Student not found or update failed'}), 404
            
    except Exception as e:
        print(f"\n❌ ERROR in update_attendance_manual:")
        print(f"Error: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/attendance/clear', methods=['POST'])
def clear_attendance_data():
    try:
        d = request.get_json()
        if attendance_system:
            attendance_system.student_status = {}
            attendance_system.attendance_count = 0
            count = attendance_system.db.clear_session_data(d['collection_name'], d['session_name'])
        else:
            db = DatabaseManager(MONGODB_CONFIG)
            count = db.clear_session_data(d['collection_name'], d['session_name'])
            db.close()
        return jsonify({'success': True, 'message': f'Cleared {count} students', 'count': count})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/collections')
def get_collections():
    try:
        db = DatabaseManager(MONGODB_CONFIG)
        cols = db.get_all_daily_collections()
        db.close()
        return jsonify({'success': True, 'data': cols})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/student/history/<identifier>')
def get_student_history(identifier):
    try:
        field = request.args.get('search_by', 'prn_no')
        db = DatabaseManager(MONGODB_CONFIG)
        hist = db.get_student_history(identifier, field)
        db.close()
        t = len(hist)
        p = sum(1 for h in hist if h['status'] == 'Present')
        det = hist[0] if hist else {}
        return jsonify({'success': True, 'identifier': identifier,
                       'student_details': {'prn_no': det.get('prn_no', ''), 'roll_no': det.get('roll_no', ''),
                                          'name': det.get('name', '')},
                       'statistics': {'total_sessions': t, 'present': p, 'absent': t - p,
                                     'attendance_percentage': round((p / t * 100), 2) if t > 0 else 0},
                       'history': hist})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reports/preview/<collection_name>')
def preview_report(collection_name):
    try:
        sn = request.args.get('session')
        db = DatabaseManager(MONGODB_CONFIG)
        if sn:
            data = db.get_session_data_for_preview(collection_name, sn)
            db.close()
            return jsonify({'success': True, 'data': data}) if data else \
                   jsonify({'success': False, 'error': 'No data found'}), 404
        doc = db.db[collection_name].find_one({})
        if not doc:
            db.close()
            return jsonify({'success': False, 'error': 'Collection not found'}), 404
        all_sess = [db.get_session_data_for_preview(collection_name, s) 
                    for s in ALL_SESSIONS if s in doc['sessions']]
        all_sess = [s for s in all_sess if s]
        db.close()
        return jsonify({'success': True, 'collection_name': collection_name, 'date': doc['date'],
                       'department': doc['department'], 'classroom': doc['classroom'],
                       'teacher_name': doc['teacher_name'], 'sessions': all_sess})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reports/export/<collection_name>')
def export_report(collection_name):
    try:
        sn = request.args.get('session')
        db = DatabaseManager(MONGODB_CONFIG)
        ef = db.generate_excel_report(collection_name, sn)
        db.close()
        if ef:
            return send_file(ef, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name=f"{collection_name}_{sn if sn else 'all'}.xlsx")
        return jsonify({'success': False, 'error': 'Failed to generate'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/reports')
def reports_page():
    return render_template_string(REPORTS_HTML)

@app.route('/student')
def student_page():
    return render_template_string(STUDENT_HTML)

HTML_TEMPLATE = '''<!DOCTYPE html>
<html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Attendance System</title><style>
*{margin:0;padding:0;box-sizing:border-box}body{font-family:'Segoe UI',sans-serif;background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);min-height:100vh;padding:15px}
.container{max-width:1800px;margin:0 auto}.header{background:white;padding:20px 25px;border-radius:12px;box-shadow:0 5px 15px rgba(0,0,0,0.2);margin-bottom:15px;display:flex;justify-content:space-between;align-items:center}
.header h1{color:#667eea;font-size:1.8em}.nav-links{display:flex;gap:10px}.nav-links a{padding:8px 16px;background:#667eea;color:white;text-decoration:none;border-radius:6px;font-weight:600;transition:background 0.3s}
.nav-links a:hover{background:#5568d3}.camera-section{background:white;padding:15px;border-radius:12px;box-shadow:0 5px 15px rgba(0,0,0,0.1);margin-bottom:15px}
.config-row{display:grid;grid-template-columns:repeat(5,1fr) auto;gap:10px;margin-bottom:10px;align-items:end}.form-group{display:flex;flex-direction:column;gap:5px}
.form-group label{font-weight:600;color:#555;font-size:0.85em}.form-group input,.form-group select{padding:8px 10px;border:2px solid #ddd;border-radius:6px;font-size:0.9em}
.btn{padding:8px 16px;border:none;border-radius:6px;font-size:0.9em;font-weight:bold;cursor:pointer;transition:all 0.3s}.btn-confirm{background:#3b82f6;color:white}
.btn:disabled{background:#9ca3af;cursor:not-allowed;opacity:0.6}.camera-controls{display:flex;gap:10px;align-items:center}.btn-start{background:#10b981;color:white;flex:1}
.btn-stop{background:#ef4444;color:white;flex:1}.status-badge{display:inline-flex;align-items:center;gap:6px;padding:6px 12px;background:#f3f4f6;border-radius:6px;font-weight:600;font-size:0.85em}
.status-dot{width:10px;height:10px;border-radius:50%;background:#ef4444}.status-dot.active{background:#10b981;animation:pulse 2s infinite}@keyframes pulse{0%,100%{opacity:1}50%{opacity:0.5}}
.video-row{display:grid;grid-template-columns:1fr auto;gap:15px}.video-container{background:#000;border-radius:8px;overflow:hidden;height:400px;display:flex;align-items:center;justify-content:center}
.video-container img{max-width:100%;max-height:100%;object-fit:contain}.video-placeholder{color:#9ca3af;font-size:1.1em;text-align:center}.info-stats{display:flex;flex-direction:column;gap:8px;min-width:200px}
.stat-box{background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);padding:12px;border-radius:8px;color:white;text-align:center}.stat-box h3{font-size:0.7em;opacity:0.9;margin-bottom:4px}
.stat-box .value{font-size:1.6em;font-weight:bold}.attendance-section{background:white;padding:20px;border-radius:12px;box-shadow:0 5px 15px rgba(0,0,0,0.1)}
.section-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:15px}.section-header h2{color:#333;font-size:1.4em}.action-buttons{display:flex;gap:8px}
.btn-refresh{background:#667eea;color:white;border:none;padding:8px 16px;border-radius:6px;cursor:pointer;font-weight:600}.btn-clear{background:#f59e0b;color:white;border:none;padding:8px 16px;border-radius:6px;cursor:pointer;font-weight:600}
.stats-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin-bottom:15px}.stat-card{background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);padding:15px;border-radius:8px;color:white;text-align:center}
.stat-card h3{font-size:0.7em;text-transform:uppercase;margin-bottom:6px;opacity:0.9}.stat-card .value{font-size:1.8em;font-weight:bold}.table-container{overflow-x:auto;max-height:450px;overflow-y:auto}
table{width:100%;border-collapse:collapse;font-size:0.85em}table th{background:#f8f9fa;padding:10px 8px;text-align:left;font-weight:bold;color:#333;position:sticky;top:0;z-index:10}
table td{padding:10px 8px;border-bottom:1px solid #eee}table tr:hover{background:#f8f9fa}.status-badge-table{padding:3px 10px;border-radius:20px;font-size:0.8em;font-weight:bold;display:inline-block}
.status-badge-table.present{background:#d1fae5;color:#065f46}.status-badge-table.absent{background:#fee2e2;color:#991b1b}.status-badge-table.temporary-absent{background:#fef3c7;color:#92400e}
.status-toggle{padding:4px 8px;border:none;border-radius:4px;cursor:pointer;font-size:0.75em;font-weight:bold}.status-toggle.to-present{background:#10b981;color:white}
.status-toggle.to-absent{background:#ef4444;color:white}.manual-badge{background:#fbbf24;color:#78350f;padding:2px 6px;border-radius:4px;font-size:0.7em;margin-left:5px}
.auto-badge{background:#10b981;color:white;padding:2px 6px;border-radius:4px;font-size:0.7em;margin-left:5px}.config-confirmed{background:#d1fae5;border:2px solid #10b981;padding:10px;border-radius:6px;margin-bottom:10px;display:none}
.config-confirmed.show{display:block}.error-msg{background:#fee2e2;border:2px solid #ef4444;color:#991b1b;padding:10px;border-radius:6px;margin-bottom:10px;display:none}.error-msg.show{display:block}
.info-msg{background:#dbeafe;border:2px solid #3b82f6;color:#1e40af;padding:10px;border-radius:6px;margin-bottom:10px;display:none}.info-msg.show{display:block}
.row-auto-detected{background:#f0fdf4!important}
</style></head><body><div class="container"><div class="header"><h1>📸 Attendance System</h1><div class="nav-links">
<a href="/">Dashboard</a><a href="/reports">Reports</a><a href="/student">Student View</a></div></div>
<div class="camera-section"><div class="error-msg" id="errorBanner"></div><div class="info-msg" id="infoBanner"></div><div class="config-row">
<div class="form-group"><label>Year (2022-2025)</label><input type="text" id="year" placeholder="e.g., 2022"></div>
<div class="form-group"><label>Department</label><input type="text" id="department" placeholder="e.g., CSBS"></div>
<div class="form-group"><label>Classroom</label><input type="text" id="classroom" placeholder="e.g., 301"></div>
<div class="form-group"><label>Teacher</label><input type="text" id="teacherInput" placeholder="Prof. Name"></div>
<div class="form-group"><label>Camera ID</label><input type="text" id="cameraId" placeholder="CAM-01" value="CAM-01"></div>
<button class="btn btn-confirm" onclick="confirmConfig()" id="confirmBtn">Confirm</button></div>
<div class="config-confirmed" id="confirmedBanner"><strong>Config:</strong> <span id="confirmedText"></span> | <strong>Sheet:</strong> <span id="sheetName"></span> | <strong>Cam:</strong> <span id="confirmedCamera"></span></div>
<div class="camera-controls" style="margin-bottom:10px"><div class="form-group" style="flex:0 0 150px"><label>Mode</label><select id="modeSelect">
<option value="1">By Name</option><option value="2">By Roll Number</option></select></div>
<button class="btn btn-start" id="startBtn" onclick="startCamera()" disabled>Start</button>
<button class="btn btn-stop" id="stopBtn" onclick="stopCamera()" disabled>Stop</button>
<div class="status-badge"><div class="status-dot" id="statusDot"></div><span id="statusText">Stopped</span></div></div>
<div class="video-row"><div class="video-container" id="videoContainer"><div class="video-placeholder">Configure and confirm to load data</div></div>
<div class="info-stats"><div class="stat-box"><h3>Faces</h3><div class="value" id="currentFaces">0</div></div>
<div class="stat-box"><h3>Session</h3><div class="value" style="font-size:1.2em" id="currentSession">-</div></div>
<div class="stat-box"><h3>Total</h3><div class="value" id="quickTotal">-</div></div>
<div class="stat-box"><h3>Present</h3><div class="value" id="quickPresent">-</div></div>
<div class="stat-box"><h3>Absent</h3><div class="value" id="quickAbsent">-</div></div></div></div></div>
<div class="attendance-section"><div class="section-header"><h2>Current Session</h2><div class="action-buttons">
<button class="btn-refresh" onclick="refreshData()">Refresh</button><button class="btn-clear" onclick="clearSessionData()" id="clearBtn" disabled>Clear</button></div></div>
<div class="stats-grid"><div class="stat-card"><h3>Total</h3><div class="value" id="totalStudents">-</div></div>
<div class="stat-card"><h3>Present</h3><div class="value" id="presentCount">-</div></div>
<div class="stat-card"><h3>Absent</h3><div class="value" id="absentCount">-</div></div>
<div class="stat-card"><h3>Temp Absent</h3><div class="value" id="tempAbsentCount">-</div></div>
<div class="stat-card"><h3>Attendance %</h3><div class="value" id="attendancePercentage">-</div></div></div>
<div class="table-container"><table><thead><tr><th>PRN</th><th>Roll</th><th>Name</th><th>Status</th><th>First</th><th>Last</th><th>Present</th><th>Absent</th><th>Action</th></tr></thead>
<tbody id="attendanceBody"><tr><td colspan="9" style="text-align:center">Configure and confirm to view data</td></tr></tbody></table></div></div></div>
<script>const API=window.location.origin+'/api';const YM={'2022':'B.Tech','2023':'TY','2024':'SY','2025':'FY'};
let running=false,interval=null,col=null,sess=null,conf=false,cfg={};function showError(m){const e=document.getElementById('errorBanner');
e.textContent='❌ '+m;e.classList.add('show');setTimeout(()=>e.classList.remove('show'),5000)}
function showInfo(m){const i=document.getElementById('infoBanner');i.textContent='ℹ️ '+m;i.classList.add('show');setTimeout(()=>i.classList.remove('show'),8000)}
async function confirmConfig(){const y=document.getElementById('year').value.trim(),d=document.getElementById('department').value.trim(),
c=document.getElementById('classroom').value.trim(),t=document.getElementById('teacherInput').value.trim(),
cam=document.getElementById('cameraId').value.trim();if(!y||!d||!c||!t){showError('Please fill all fields');return}
const btn=document.getElementById('confirmBtn');btn.disabled=true;btn.textContent='Loading...';try{cfg={year:y,department:d,classroom:c,teacher_name:t,camera_ids:[cam]};
const r=await fetch(`${API}/preview-config`,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(cfg)});
const data=await r.json();if(!data.success){showError(data.message||data.error||'Failed to load config');btn.disabled=false;btn.textContent='Confirm';return}
conf=true;col=data.collection_name;sess=data.session_name;document.getElementById('confirmedText').textContent=`${y}|${d}|${c}|${t}`;
document.getElementById('sheetName').textContent=data.sheet_loaded;document.getElementById('confirmedCamera').textContent=cam;
document.getElementById('confirmedBanner').classList.add('show');document.getElementById('startBtn').disabled=false;
document.getElementById('clearBtn').disabled=false;document.getElementById('currentSession').textContent=sess;
console.log('Config data received:', data);console.log('Summary:', data.summary);console.log('Attendance records:', data.attendance);
// Check if any students are using Roll as PRN
if(data.attendance&&data.attendance.length>0){
const usingRoll=data.attendance.some(s=>s.flags&&s.flags.using_roll_as_prn);
if(usingRoll)showInfo('Note: Some students have empty PRN numbers. Using Roll Numbers as identifiers.');}
if(data.summary)updateStats(data.summary);if(data.attendance&&data.attendance.length>0)displayData(data.attendance);
else{document.getElementById('attendanceBody').innerHTML='<tr><td colspan="9" style="text-align:center">No students found in Excel sheet</td></tr>'}
document.getElementById('videoContainer').innerHTML='<div class="video-placeholder">Ready to start camera</div>'}
catch(e){console.error('Config error:',e);showError('Network error: '+e.message)}finally{btn.disabled=false;btn.textContent='Confirm'}}
async function startCamera(){if(!conf){showError('Please confirm configuration first');return}const m=document.getElementById('modeSelect').value,
sb=document.getElementById('startBtn'),stb=document.getElementById('stopBtn');sb.disabled=true;sb.textContent='Starting...';try{
const r=await fetch(`${API}/camera/start`,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({mode:parseInt(m),...cfg})});
const d=await r.json();if(d.success){running=true;updateStatus(true);sb.disabled=true;stb.disabled=false;sb.textContent='Start';
document.getElementById('videoContainer').innerHTML='<img src="'+API+'/video_feed?t='+Date.now()+'">';if(d.initial_data){col=d.initial_data.collection_name;
sess=d.initial_data.session_name;updateStats(d.initial_data.summary);displayData(d.initial_data.attendance)}startRefresh()}
else throw new Error(d.message||d.error)}catch(e){showError(e.message);sb.disabled=false;sb.textContent='Start'}}
async function stopCamera(){const sb=document.getElementById('startBtn'),stb=document.getElementById('stopBtn');stb.disabled=true;stb.textContent='Stopping...';
try{const r=await fetch(`${API}/camera/stop`,{method:'POST'});const d=await r.json();if(d.success){running=false;updateStatus(false);sb.disabled=false;
stb.disabled=true;stb.textContent='Stop';document.getElementById('videoContainer').innerHTML='<div class="video-placeholder">Camera stopped</div>';
stopRefresh()}}catch(e){showError('Failed to stop: '+e.message);stb.disabled=false;stb.textContent='Stop'}}
function updateStatus(r){const dot=document.getElementById('statusDot'),txt=document.getElementById('statusText');
if(r){dot.classList.add('active');txt.textContent='Running'}else{dot.classList.remove('active');txt.textContent='Stopped'}}
async function updateInfo(){if(!running)return;try{const r=await fetch(`${API}/camera/status`);const d=await r.json();
document.getElementById('currentFaces').textContent=d.current_faces||0;document.getElementById('currentSession').textContent=d.current_session||'-';
if(d.current_session&&d.current_session!==sess){sess=d.current_session;await loadSession()}}catch(e){}}
async function loadSession(){if(!col||!sess)return;try{const r=await fetch(`${API}/current-session`);const d=await r.json();
if(d.success&&d.active){col=d.collection_name;sess=d.session_name;updateStats(d.summary);displayData(d.attendance)}}catch(e){}}
function updateStats(s){document.getElementById('totalStudents').textContent=s.total||0;document.getElementById('presentCount').textContent=s.present||0;
document.getElementById('absentCount').textContent=s.absent||0;document.getElementById('tempAbsentCount').textContent=s.temporary_absent||0;
document.getElementById('attendancePercentage').textContent=(s.attendance_percentage||0).toFixed(2)+'%';document.getElementById('quickTotal').textContent=s.total||0;
document.getElementById('quickPresent').textContent=s.present||0;document.getElementById('quickAbsent').textContent=s.absent||0}
function displayData(recs){const tb=document.getElementById('attendanceBody');if(!recs||recs.length===0){
tb.innerHTML='<tr><td colspan="9" style="text-align:center">No records found</td></tr>';return}tb.innerHTML='';
recs.forEach(r=>{const row=document.createElement('tr');const st=r.status||'Absent';let bc='absent';
if(st==='Present')bc='present';else if(st==='Temporary Absent')bc='temporary-absent';else if(st==='Permanently Absent')bc='permanently-absent';
const ts=r.timestamps||{},du=r.durations||{},fl=r.flags||{};const autoDetected=st==='Present'&&!fl.manual_override;
if(autoDetected)row.classList.add('row-auto-detected');


let prn = r.prn_no || '';
if (!prn) {
  console.warn('Using Roll No as PRN for student:', r);
  prn = r.roll_no || '(No ID)';
}


const usingRollAsPrn=fl.using_roll_as_prn||false;
const prnDisplay=usingRollAsPrn?`${prn} <span style="color:#f59e0b;font-size:0.7em">(Roll)</span>`:prn;
row.innerHTML=`<td>${prnDisplay}</td><td>${r.roll_no||'-'}</td><td>${r.name||'-'}</td>
<td><span class="status-badge-table ${bc}">${st}</span>${fl.manual_override?'<span class="manual-badge">Manual</span>':''}${autoDetected?'<span class="auto-badge">Auto</span>':''}</td>
<td>${ts.first_seen||'N/A'}</td><td>${ts.last_seen||'N/A'}</td><td>${du.total_present_human||'0 sec'}</td><td>${du.total_absent_human||'0 sec'}</td>
<td><button class="status-toggle ${st==='Present'?'to-absent':'to-present'}" data-prn="${prn}" data-status="${st}" onclick="toggleAtt(this)">${st==='Present'?'Mark Absent':'Mark Present'}</button></td>`;
tb.appendChild(row)})}
async function toggleAtt(btn){const prn=btn.getAttribute('data-prn');const cur=btn.getAttribute('data-status');
console.log('Toggle attendance - PRN:',prn,'Current:',cur,'Collection:',col,'Session:',sess);
if(!prn||prn==='-'){showError('Invalid PRN number');return}if(!col||!sess){showError('No active session');return}
const ns=cur==='Present'?'Absent':'Present',origText=btn.textContent;btn.disabled=true;btn.textContent='Updating...';
try{console.log('Sending update request...');const r=await fetch(`${API}/attendance/update`,
{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({collection_name:col,session_name:sess,prn_no:prn,status:ns})});
const d=await r.json();console.log('Update response:',d);if(d.success){await refreshData();console.log('✓ Successfully updated to',ns)}
else showError(d.error||'Failed to update')}catch(e){console.error('Update error:',e);showError('Network error: '+e.message)}
finally{btn.disabled=false;btn.textContent=origText}}
async function refreshData(){await loadSession()}
async function clearSessionData(){if(!col||!sess){showError('No active session');return}if(!confirm('Clear all attendance data for this session?'))return;
try{const r=await fetch(`${API}/attendance/clear`,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({collection_name:col,session_name:sess})});
const d=await r.json();if(d.success){alert('✓ '+d.message);await refreshData()}else showError(d.error||'Failed to clear')}catch(e){showError('Network error: '+e.message)}}
function startRefresh(){if(interval)clearInterval(interval);updateInfo();interval=setInterval(()=>{updateInfo();refreshData()},2000)}
function stopRefresh(){if(interval){clearInterval(interval);interval=null}}window.onload=()=>updateStatus(false);</script></body></html>'''

REPORTS_HTML = '''<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Reports</title><style>
*{margin:0;padding:0;box-sizing:border-box}body{font-family:'Segoe UI',sans-serif;background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);min-height:100vh;padding:20px}
.container{max-width:1400px;margin:0 auto}.header{background:white;padding:25px 30px;border-radius:15px;box-shadow:0 10px 30px rgba(0,0,0,0.2);margin-bottom:20px;display:flex;justify-content:space-between;align-items:center}
.header h1{color:#667eea;font-size:2em}.nav-links{display:flex;gap:15px}.nav-links a{padding:10px 20px;background:#667eea;color:white;text-decoration:none;border-radius:8px;font-weight:bold}
.reports-section{background:white;padding:25px;border-radius:15px;box-shadow:0 5px 15px rgba(0,0,0,0.1)}.collections-grid{display:grid;gap:15px}
.collection-card{background:linear-gradient(135deg,#f8f9fa 0%,#e9ecef 100%);padding:20px;border-radius:10px;border-left:5px solid #667eea}
.collection-header{display:flex;justify-content:space-between;margin-bottom:15px}.collection-title{font-size:1.2em;font-weight:bold;color:#333}
.collection-meta{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:15px}.meta-item{background:white;padding:10px;border-radius:6px;text-align:center}
.btn{padding:8px 16px;border:none;border-radius:6px;font-weight:bold;cursor:pointer;font-size:0.9em}.btn-preview{background:#3b82f6;color:white}
.btn-export{background:#10b981;color:white}.btn-view{background:#8b5cf6;color:white}.collection-actions{display:flex;gap:10px;flex-wrap:wrap}
.session-details{display:none;margin-top:15px;padding-top:15px;border-top:2px solid #e5e7eb}.session-details.show{display:block}
.sessions-list{display:grid;grid-template-columns:repeat(4,1fr);gap:10px}.session-btn{padding:10px;background:white;border:2px solid #667eea;border-radius:6px;cursor:pointer;font-weight:600;color:#667eea;text-align:center}
.preview-modal{display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.5);z-index:1000;overflow-y:auto}.preview-modal.show{display:block}
.preview-content{background:white;margin:50px auto;max-width:1200px;border-radius:15px;padding:30px;position:relative}.preview-close{position:absolute;top:15px;right:15px;background:#ef4444;color:white;border:none;padding:8px 16px;border-radius:6px;cursor:pointer;font-weight:bold}
.preview-header h2{color:#667eea;font-size:1.8em;margin-bottom:10px}.preview-info{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:20px}
.preview-info-item{background:#f8f9fa;padding:10px;border-radius:6px}.preview-stats{display:grid;grid-template-columns:repeat(5,1fr);gap:15px;margin-bottom:20px}
.preview-stat{background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);padding:15px;border-radius:8px;color:white;text-align:center}.preview-stat h3{font-size:0.75em;opacity:0.9}
.preview-stat .value{font-size:2em;font-weight:bold}.preview-table{overflow-x:auto;max-height:500px;overflow-y:auto;margin-top:20px}
.preview-table table{width:100%;border-collapse:collapse}.preview-table th{background:#667eea;color:white;padding:12px;text-align:left;position:sticky;top:0}
.preview-table td{padding:10px 12px;border-bottom:1px solid #e5e7eb}.preview-table tr:hover{background:#f8f9fa}.status-badge{padding:4px 12px;border-radius:20px;font-size:0.85em;font-weight:bold}
.status-badge.present{background:#d1fae5;color:#065f46}.status-badge.absent{background:#fee2e2;color:#991b1b}.status-badge.temporary-absent{background:#fef3c7;color:#92400e}
.session-tabs{display:flex;gap:10px;margin-bottom:20px;flex-wrap:wrap}.session-tab{padding:10px 20px;background:#f8f9fa;border:2px solid #e5e7eb;border-radius:6px;cursor:pointer;font-weight:600;color:#666}
.session-tab.active{background:#667eea;color:white;border-color:#667eea}.download-btn{background:#10b981;color:white;padding:10px 20px;border:none;border-radius:6px;font-weight:bold;cursor:pointer;margin-top:20px}
</style></head><body><div class="container"><div class="header"><h1>📊 Reports</h1><div class="nav-links">
<a href="/">Dashboard</a><a href="/reports">Reports</a><a href="/student">Student View</a></div></div>
<div class="reports-section"><h2>Daily Collections</h2><div id="collectionsContainer" class="collections-grid"><div class="loading">Loading reports...</div></div></div></div>
<div class="preview-modal" id="previewModal"><div class="preview-content"><button class="preview-close" onclick="closePreview()">✕ Close</button>
<div class="preview-header"><h2 id="previewTitle">Attendance Report</h2></div><div class="preview-info" id="previewInfo"></div>
<div class="session-tabs" id="sessionTabs"></div><div class="preview-stats" id="previewStats"></div><div class="preview-table">
<table><thead><tr><th>PRN</th><th>Roll No</th><th>Name</th><th>Status</th><th>First Seen</th><th>Last Seen</th><th>Present Duration</th><th>Absent Duration</th></tr></thead>
<tbody id="previewTableBody"></tbody></table></div><button class="download-btn" id="downloadBtn">📥 Download Excel</button></div></div>
<script>const API=window.location.origin+'/api';let currentCollection=null,allSessionsData=[];
async function loadCollections(){try{const r=await fetch(`${API}/collections`);const d=await r.json();
if(d.success&&d.data.length>0)displayCollections(d.data);else document.getElementById('collectionsContainer').innerHTML='<div>No collections found</div>'}
catch(e){document.getElementById('collectionsContainer').innerHTML='<div>Error loading reports</div>'}}
function displayCollections(cols){const c=document.getElementById('collectionsContainer');c.innerHTML='';cols.forEach(col=>{const card=document.createElement('div');
card.className='collection-card';card.innerHTML=`<div class="collection-header"><div class="collection-title">${col.collection_name}</div></div>
<div class="collection-meta"><div class="meta-item"><label>Date</label><value>${col.date}</value></div>
<div class="meta-item"><label>Department</label><value>${col.department}</value></div>
<div class="meta-item"><label>Classroom</label><value>${col.classroom}</value></div></div>
<div class="collection-meta"><div class="meta-item"><label>Teacher</label><value>${col.teacher_name}</value></div>
<div class="meta-item"><label>Year</label><value>${col.year} (${col.year_code})</value></div>
<div class="meta-item"><label>Created</label><value>${col.created_at}</value></div></div>
<div class="collection-actions"><button class="btn btn-preview" onclick="previewReport('${col.collection_name}')">👁️ Preview</button>
<button class="btn btn-view" onclick="toggleSessions('${col.collection_name}')">View Sessions</button>
<button class="btn btn-export" onclick="exportReport('${col.collection_name}')">📥 Export All</button></div>
<div class="session-details" id="sessions-${col.collection_name}"><h3 style="margin-bottom:10px;color:#555">Preview/Export by Session:</h3>
<div class="sessions-list">${[1,2,3,4,5,6,7,8].map(i=>`<button class="session-btn" onclick="previewSession('${col.collection_name}','Session ${i}')">Session ${i}</button>`).join('')}</div></div>`;
c.appendChild(card)})}function toggleSessions(cn){document.getElementById(`sessions-${cn}`).classList.toggle('show')}
async function previewReport(cn){try{currentCollection=cn;const r=await fetch(`${API}/reports/preview/${cn}`);const d=await r.json();
if(d.success){allSessionsData=d.sessions||[];showPreviewModal(d,null)}else alert('Failed to load preview')}catch(e){alert('Error loading preview')}}
async function previewSession(cn,sn){try{currentCollection=cn;const r=await fetch(`${API}/reports/preview/${cn}?session=${encodeURIComponent(sn)}`);
const d=await r.json();if(d.success){allSessionsData=[d.data];showPreviewModal(d.data,sn)}else alert('Failed to load session')}catch(e){alert('Error loading session')}}
function showPreviewModal(data,singleSession){const modal=document.getElementById('previewModal'),title=document.getElementById('previewTitle'),
info=document.getElementById('previewInfo'),tabs=document.getElementById('sessionTabs'),downloadBtn=document.getElementById('downloadBtn');
if(singleSession){title.textContent=`${data.session_name} - ${data.date}`;info.innerHTML=`<div class="preview-info-item"><label>Department</label><value>${data.department}</value></div>
<div class="preview-info-item"><label>Classroom</label><value>${data.classroom}</value></div>
<div class="preview-info-item"><label>Teacher</label><value>${data.teacher_name}</value></div>
<div class="preview-info-item"><label>Session</label><value>${data.session_name}</value></div>`;tabs.innerHTML='';displaySessionData(data);
downloadBtn.onclick=()=>exportSession(currentCollection,singleSession)}else{title.textContent=`All Sessions - ${data.date}`;
info.innerHTML=`<div class="preview-info-item"><label>Date</label><value>${data.date}</value></div>
<div class="preview-info-item"><label>Department</label><value>${data.department}</value></div>
<div class="preview-info-item"><label>Classroom</label><value>${data.classroom}</value></div>
<div class="preview-info-item"><label>Teacher</label><value>${data.teacher_name}</value></div>`;tabs.innerHTML='';
allSessionsData.forEach((sess,idx)=>{const tab=document.createElement('div');tab.className='session-tab'+(idx===0?' active':'');
tab.textContent=sess.session_name;tab.onclick=()=>switchSession(idx);tabs.appendChild(tab)});displaySessionData(allSessionsData[0]);
downloadBtn.onclick=()=>exportReport(currentCollection)}modal.classList.add('show')}
function switchSession(idx){document.querySelectorAll('.session-tab').forEach((t,i)=>{if(i===idx)t.classList.add('active');else t.classList.remove('active')});
displaySessionData(allSessionsData[idx])}function displaySessionData(data){const stats=document.getElementById('previewStats'),tbody=document.getElementById('previewTableBody'),s=data.summary;
stats.innerHTML=`<div class="preview-stat"><h3>Total</h3><div class="value">${s.total}</div></div>
<div class="preview-stat"><h3>Present</h3><div class="value">${s.present}</div></div>
<div class="preview-stat"><h3>Absent</h3><div class="value">${s.absent}</div></div>
<div class="preview-stat"><h3>Temp Absent</h3><div class="value">${s.temporary_absent}</div></div>
<div class="preview-stat"><h3>Attendance %</h3><div class="value">${s.attendance_percentage.toFixed(2)}%</div></div>`;tbody.innerHTML='';
if(!data.students||data.students.length===0){tbody.innerHTML='<tr><td colspan="8" style="text-align:center">No students found</td></tr>';return}
data.students.forEach(stu=>{const st=stu.status||'Absent';let bc='absent';if(st==='Present')bc='present';
else if(st==='Temporary Absent')bc='temporary-absent';else if(st==='Permanently Absent')bc='permanently-absent';
const row=document.createElement('tr');row.innerHTML=`<td>${stu.prn_no||'-'}</td><td>${stu.roll_no||'-'}</td><td>${stu.name||'-'}</td>
<td><span class="status-badge ${bc}">${st}</span></td><td>${stu.timestamps?.first_seen||'N/A'}</td>
<td>${stu.timestamps?.last_seen||'N/A'}</td><td>${stu.durations?.total_present_human||'0 sec'}</td>
<td>${stu.durations?.total_absent_human||'0 sec'}</td>`;tbody.appendChild(row)})}
function closePreview(){document.getElementById('previewModal').classList.remove('show');currentCollection=null;allSessionsData=[]}
async function exportReport(cn){window.location.href=`${API}/reports/export/${cn}`}
async function exportSession(cn,sn){window.location.href=`${API}/reports/export/${cn}?session=${encodeURIComponent(sn)}`}
window.onclick=e=>{if(e.target===document.getElementById('previewModal'))closePreview()};window.onload=loadCollections;</script></body></html>'''

STUDENT_HTML = '''<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Student View</title><style>
*{margin:0;padding:0;box-sizing:border-box}body{font-family:'Segoe UI',sans-serif;background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);min-height:100vh;padding:20px}
.container{max-width:1200px;margin:0 auto}.header{background:white;padding:25px 30px;border-radius:15px;box-shadow:0 10px 30px rgba(0,0,0,0.2);margin-bottom:20px;display:flex;justify-content:space-between;align-items:center}
.header h1{color:#667eea;font-size:2em}.nav-links{display:flex;gap:15px}.nav-links a{padding:10px 20px;background:#667eea;color:white;text-decoration:none;border-radius:8px;font-weight:bold}
.search-section{background:white;padding:25px;border-radius:15px;box-shadow:0 5px 15px rgba(0,0,0,0.1);margin-bottom:20px}.search-form{display:grid;grid-template-columns:150px 1fr auto;gap:15px;align-items:end}
.form-group{display:flex;flex-direction:column;gap:8px}.form-group label{font-weight:bold;color:#555}.form-group select,.form-group input{padding:10px;border:2px solid #ddd;border-radius:8px}
.btn-search{padding:10px 30px;background:#667eea;color:white;border:none;border-radius:8px;font-weight:bold;cursor:pointer}.results-section{background:white;padding:25px;border-radius:15px;box-shadow:0 5px 15px rgba(0,0,0,0.1);display:none}
.results-section.show{display:block}.student-info{background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);padding:20px;border-radius:10px;color:white;margin-bottom:20px}
.student-info h2{font-size:1.5em;margin-bottom:10px}.student-details{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-top:15px;background:rgba(255,255,255,0.1);padding:15px;border-radius:8px}
.detail-item{text-align:center}.detail-item label{font-size:0.8em;opacity:0.9;display:block;margin-bottom:5px}.detail-item value{font-size:1.1em;font-weight:bold}
.stats-row{display:grid;grid-template-columns:repeat(4,1fr);gap:15px;margin-bottom:20px}.stat-box{background:white;padding:20px;border-radius:10px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,0.1)}
.stat-box h3{font-size:0.8em;color:#666;margin-bottom:8px;text-transform:uppercase}.stat-box .value{font-size:2em;font-weight:bold;color:#667eea}
.history-table{overflow-x:auto;max-height:500px;overflow-y:auto}table{width:100%;border-collapse:collapse}table th{background:#f8f9fa;padding:12px;text-align:left;font-weight:bold;color:#333;position:sticky;top:0}
table td{padding:12px;border-bottom:1px solid #eee}table tr:hover{background:#f8f9fa}.status-badge{padding:4px 12px;border-radius:20px;font-size:0.85em;font-weight:bold}
.status-badge.present{background:#d1fae5;color:#065f46}.status-badge.absent{background:#fee2e2;color:#991b1b}
</style></head><body><div class="container"><div class="header"><h1>👤 Student View</h1><div class="nav-links">
<a href="/">Dashboard</a><a href="/reports">Reports</a><a href="/student">Student View</a></div></div>
<div class="search-section"><h2 style="margin-bottom:20px;color:#333">Search Student</h2><div class="search-form">
<div class="form-group"><label>Search By</label><select id="searchBy">
<option value="prn_no">PRN Number</option><option value="roll_no">Roll Number</option><option value="name">Name</option></select></div>
<div class="form-group"><label>Enter PRN/Roll/Name</label><input type="text" id="searchInput" placeholder="e.g., PRN123456, 101, or John Doe"></div>
<button class="btn-search" onclick="searchStudent()">Search</button></div></div>
<div class="results-section" id="resultsSection"><div class="student-info"><h2>Student Information</h2>
<div class="student-details"><div class="detail-item"><label>PRN Number</label><value id="studentPRN">-</value></div>
<div class="detail-item"><label>Roll Number</label><value id="studentRoll">-</value></div>
<div class="detail-item"><label>Student Name</label><value id="studentFullName">-</value></div></div></div>
<div class="stats-row"><div class="stat-box"><h3>Total Sessions</h3><div class="value" id="totalSessions">0</div></div>
<div class="stat-box"><h3>Present</h3><div class="value" id="presentSessions">0</div></div>
<div class="stat-box"><h3>Absent</h3><div class="value" id="absentSessions">0</div></div>
<div class="stat-box"><h3>Attendance %</h3><div class="value" id="attendancePercent">0%</div></div></div>
<h3 style="margin-bottom:15px;color:#333">Attendance History</h3><div class="history-table"><table><thead>
<tr><th>Date</th><th>Session</th><th>Department</th><th>Classroom</th><th>Status</th><th>First</th><th>Last</th><th>Duration</th></tr></thead>
<tbody id="historyBody"><tr><td colspan="8" style="text-align:center">Search for a student</td></tr></tbody></table></div></div></div>
<script>const API=window.location.origin+'/api';async function searchStudent(){const sb=document.getElementById('searchBy').value,
id=document.getElementById('searchInput').value.trim();if(!id){alert('Please enter a PRN, roll number, or name');return}
const hb=document.getElementById('historyBody');hb.innerHTML='<tr><td colspan="8" style="text-align:center">Loading...</td></tr>';
document.getElementById('resultsSection').classList.add('show');try{const r=await fetch(`${API}/student/history/${encodeURIComponent(id)}?search_by=${sb}`);
const d=await r.json();if(d.success)displayData(d);else hb.innerHTML='<tr><td colspan="8" style="text-align:center">Error: '+d.error+'</td></tr>'}
catch(e){hb.innerHTML='<tr><td colspan="8" style="text-align:center">Network error</td></tr>'}}
function displayData(d){const det=d.student_details||{};document.getElementById('studentPRN').textContent=det.prn_no||'-';
document.getElementById('studentRoll').textContent=det.roll_no||'-';document.getElementById('studentFullName').textContent=det.name||'-';
const s=d.statistics;document.getElementById('totalSessions').textContent=s.total_sessions;
document.getElementById('presentSessions').textContent=s.present;document.getElementById('absentSessions').textContent=s.absent;
document.getElementById('attendancePercent').textContent=s.attendance_percentage.toFixed(2)+'%';const hb=document.getElementById('historyBody');
if(d.history.length===0){hb.innerHTML='<tr><td colspan="8" style="text-align:center">No attendance records found</td></tr>';return}
hb.innerHTML='';d.history.forEach(r=>{const row=document.createElement('tr');const sc=r.status==='Present'?'present':'absent';
row.innerHTML=`<td>${r.date}</td><td>${r.session}</td><td>${r.department||'-'}</td><td>${r.classroom||'-'}</td>
<td><span class="status-badge ${sc}">${r.status}</span></td><td>${r.first_seen||'N/A'}</td><td>${r.last_seen||'N/A'}</td>
<td>${r.present_duration||'0 sec'}</td>`;hb.appendChild(row)})}
document.getElementById('searchInput').addEventListener('keypress',e=>{if(e.key==='Enter')searchStudent()});</script></body></html>'''

if __name__=='__main__':
    print("="*80)
    print("ATTENDANCE SYSTEM - COMPLETE VERSION WITH FIXES")
    print("="*80)
    print("\n✅ ALL FEATURES IMPLEMENTED:")
    print("   1. PRN-based attendance tracking (with Roll Number fallback)")
    print("   2. Automatic face recognition with auto-detection badges")
    print("   3. Manual override with manual badges (FIXED)")
    print("   4. Student search by PRN/Roll/Name")
    print("   5. Web-based reports preview with modal")
    print("   6. Excel export from preview")
    print("   7. Session-wise preview and download")
    print("   8. Real-time duration tracking")
    print("\n🔧 FIXES APPLIED:")
    print("   - Fixed empty PRN issue in manual updates")
    print("   - Added data attributes for proper PRN extraction")
    print("   - Enhanced logging for debugging")
    print("   - Better validation and error messages")
    print("   - Auto-fallback to Roll Number when PRN is empty")
    print("\n💡 IMPORTANT:")
    print("   - If PRN column is empty, system will use Roll Number as identifier")
    print("   - Students with neither PRN nor Roll Number will be skipped")
    print("   - Consider adding PRN numbers to Excel for better tracking")
    print("\n🚀 Starting server on http://localhost:5000")
    print("="*80+"\n")
    try:
        app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)
    except KeyboardInterrupt:
        print("\n\n👋 Server stopped")
    except Exception as e:
        print(f"\n\n❌ Error: {e}")
        traceback.print_exc()