"""
Enhanced Flask Server with Face Recognition Attendance System
UPDATES:
1. Camera start refreshes current session
2. Roll No visible in attendance table
3. Full time tracking logic implemented (First_Seen, Last_Seen, durations, etc.)
4. New layout: Camera + Controls side-by-side, Attendance section below
5. All timing fields visible in attendance table
"""

from flask import Flask, jsonify, request, render_template_string, Response, send_file
from flask_cors import CORS
from pymongo import MongoClient
from datetime import datetime, timedelta
import os
import threading
import cv2
import numpy as np
import face_recognition
import sys
from time import sleep
from threading import Lock, Event
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import json
from bson.objectid import ObjectId

# Add utils folder to path
sys.path.append(os.path.abspath('../'))
try:
    from Excel_Format import get_current_session
except ImportError:
    def get_current_session():
        current_hour = datetime.now().hour
        if 9 <= current_hour < 12:
            return "Morning Session"
        elif 12 <= current_hour < 15:
            return "Afternoon Session"
        elif 15 <= current_hour < 18:
            return "Evening Session"
        else:
            return "General Session"

app = Flask(__name__)
CORS(app)

# MongoDB Configuration
MONGODB_CONFIG = {
    'host': 'localhost',
    'port': 27017,
    'database': 'Attendance_system'
}

TEMPLATE_FILE = 'Book2.xlsx'

# Global variables
attendance_system = None
camera_running = False
camera_lock = Lock()


# ============= ATTENDANCE SYSTEM CODE =============

class AttendanceConfig:
    """Configuration constants"""
    TEMPLATE_FILE = 'Book2.xlsx'
    ABSENCE_DETECTION_DELAY = 5
    TEMPORARY_ABSENT_THRESHOLD = 10
    PERMANENT_ABSENT_THRESHOLD = 15
    ABSENCE_CHECK_INTERVAL = 2
    MODE_NAME = 1
    MODE_ROLL_NO = 2


class DatabaseManager:
    """Handles all MongoDB database operations"""
    
    def __init__(self, mongodb_config):
        self.mongodb_config = mongodb_config
        self.client = None
        self.db = None
        self.lock = Lock()
        self._initialize_db()
    
    def _get_connection(self):
        try:
            if self.client is None:
                self.client = MongoClient(
                    self.mongodb_config['host'], 
                    self.mongodb_config['port']
                )
                self.db = self.client[self.mongodb_config['database']]
            return self.db
        except Exception as e:
            print(f"Error connecting to MongoDB: {e}")
            raise
    
    def _initialize_db(self):
        try:
            self.db = self._get_connection()
            self.db.lecture_metadata.create_index('table_name', unique=True)
            print("✓ MongoDB initialized successfully")
        except Exception as e:
            print(f"Error initializing database: {e}")
            raise
    
    def load_template_from_excel(self, excel_file, sheet_name='Sheet1'):
        try:
            wb = load_workbook(excel_file, data_only=True)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in {excel_file}")
            sheet = wb[sheet_name]
            data = []
            headers = None
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                row_data = [cell if cell is not None else '' for cell in row]
                if headers is None:
                    headers = [str(h).strip() for h in row_data]
                else:
                    data.append(row_data)
            wb.close()
            if not headers:
                raise ValueError("No headers found in template")
            print(f"✓ Loaded template: {len(headers)} columns, {len(data)} rows")
            return headers, data
        except Exception as e:
            print(f"Error loading template from Excel: {e}")
            raise
    
    def create_lecture_collection(self, session_name, template_headers, template_data):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        collection_name = f"lecture_{timestamp}"
        with self.lock:
            try:
                db = self._get_connection()
                collection = db[collection_name]
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                documents = []
                for row in template_data:
                    doc = {}
                    for i, header in enumerate(template_headers):
                        if i < len(row):
                            value = row[i]
                            if value is None or value == '':
                                doc[header] = ''
                            elif isinstance(value, (int, float)):
                                doc[header] = str(value)
                            else:
                                doc[header] = str(value).strip()
                        else:
                            doc[header] = ''
                    doc['Status'] = 'Absent'
                    doc['First_Seen'] = 'N/A'
                    doc['Last_Seen'] = 'N/A'
                    doc['Absence_Timer_Start'] = 'N/A'
                    doc['Temp_Absent_Time'] = 'N/A'
                    doc['Perm_Absent_Time'] = 'N/A'
                    doc['Total_Absent_Duration'] = '0 sec'
                    doc['Total_Present_Duration'] = '0 sec'
                    doc['Last_Updated'] = current_time
                    doc['Manual_Override'] = False
                    doc['_present_start_time'] = None
                    doc['_absent_start_time'] = None
                    doc['_total_present_seconds'] = 0
                    doc['_total_absent_seconds'] = 0
                    documents.append(doc)
                if documents:
                    collection.insert_many(documents)
                metadata = {
                    'table_name': collection_name,
                    'session_name': session_name,
                    'date': datetime.now().strftime('%Y-%m-%d'),
                    'start_time': datetime.now().strftime('%H:%M:%S'),
                    'created_at': datetime.now()
                }
                db.lecture_metadata.insert_one(metadata)
                print(f"✓ Created collection '{collection_name}' with {len(documents)} students")
                return collection_name
            except Exception as e:
                print(f"Error creating lecture collection: {e}")
                import traceback
                traceback.print_exc()
                raise
    
    def get_current_lecture_collection(self, session_name):
        with self.lock:
            try:
                db = self._get_connection()
                result = db.lecture_metadata.find_one(
                    {
                        'session_name': session_name,
                        'date': datetime.now().strftime('%Y-%m-%d')
                    },
                    sort=[('created_at', -1)]
                )
                return result['table_name'] if result else None
            except Exception as e:
                print(f"Error getting lecture collection: {e}")
                return None
    
    def _find_student_fields(self, collection):
        """Helper to find roll and name field names in collection"""
        sample_doc = collection.find_one()
        if not sample_doc:
            return None, None
        
        roll_field = None
        name_field = None
        
        # Find roll number field
        for key in sample_doc.keys():
            if 'roll' in key.lower() and key != '_id':
                roll_field = key
                break
        
        # Find name field (excluding roll fields)
        for key in sample_doc.keys():
            if 'name' in key.lower() and 'roll' not in key.lower() and key != '_id':
                name_field = key
                break
        
        return roll_field, name_field
    
    def update_attendance(self, collection_name, identifier, status, last_seen, absence_timer_start=None, search_mode='name', manual=False):
        """Update attendance with full time tracking"""
        with self.lock:
            try:
                db = self._get_connection()
                collection = db[collection_name]
                
                roll_field, name_field = self._find_student_fields(collection)
                
                if not roll_field or not name_field:
                    print("Error: Could not find roll or name fields")
                    return False
                
                search_field = roll_field if search_mode == 'roll' else name_field
                
                student = collection.find_one({
                    search_field: {'$regex': f'^{identifier}$', '$options': 'i'}
                })
                
                if not student:
                    return False
                
                current_time = datetime.now()
                current_time_str = current_time.strftime('%Y-%m-%d %H:%M:%S')
                
                update_data = {
                    'Status': status,
                    'Last_Seen': last_seen,
                    'Last_Updated': current_time_str,
                    'Manual_Override': manual
                }
                
                # Handle First_Seen
                if student.get('First_Seen') == 'N/A' and status == 'Present':
                    update_data['First_Seen'] = last_seen
                
                # Calculate durations
                prev_status = student.get('Status', 'Absent')
                present_start = student.get('_present_start_time')
                absent_start = student.get('_absent_start_time')
                total_present = student.get('_total_present_seconds', 0)
                total_absent = student.get('_total_absent_seconds', 0)
                
                # Status transition logic
                if prev_status == 'Present' and status != 'Present':
                    # Was present, now absent - calculate present duration
                    if present_start:
                        try:
                            start_dt = datetime.strptime(present_start, '%Y-%m-%d %H:%M:%S')
                            duration = (current_time - start_dt).total_seconds()
                            total_present += duration
                        except:
                            pass
                    update_data['_present_start_time'] = None
                    update_data['_absent_start_time'] = current_time_str
                    
                elif prev_status != 'Present' and status == 'Present':
                    # Was absent, now present - calculate absent duration
                    if absent_start:
                        try:
                            start_dt = datetime.strptime(absent_start, '%Y-%m-%d %H:%M:%S')
                            duration = (current_time - start_dt).total_seconds()
                            total_absent += duration
                        except:
                            pass
                    update_data['_absent_start_time'] = None
                    update_data['_present_start_time'] = current_time_str
                    
                elif status == 'Present' and not present_start:
                    # Starting present tracking
                    update_data['_present_start_time'] = current_time_str
                    
                elif status != 'Present' and not absent_start:
                    # Starting absent tracking
                    update_data['_absent_start_time'] = current_time_str
                
                update_data['_total_present_seconds'] = total_present
                update_data['_total_absent_seconds'] = total_absent
                update_data['Total_Present_Duration'] = self._format_duration(total_present)
                update_data['Total_Absent_Duration'] = self._format_duration(total_absent)
                
                # Absence timer
                if absence_timer_start:
                    update_data['Absence_Timer_Start'] = absence_timer_start
                else:
                    update_data['Absence_Timer_Start'] = 'N/A'
                
                # Temporary/Permanent absent times
                if status == 'Temporary Absent':
                    update_data['Temp_Absent_Time'] = current_time_str
                elif status == 'Permanently Absent':
                    update_data['Perm_Absent_Time'] = current_time_str
                
                result = collection.update_one(
                    {'_id': student['_id']},
                    {'$set': update_data}
                )
                
                return result.modified_count > 0
            except Exception as e:
                print(f"Error updating attendance: {e}")
                import traceback
                traceback.print_exc()
                return False
    
    def _format_duration(self, seconds):
        """Format seconds into readable duration"""
        if seconds < 60:
            return f"{int(seconds)} sec"
        elif seconds < 3600:
            minutes = int(seconds / 60)
            secs = int(seconds % 60)
            return f"{minutes} min {secs} sec"
        else:
            hours = int(seconds / 3600)
            minutes = int((seconds % 3600) / 60)
            return f"{hours} hr {minutes} min"
    
    def update_attendance_by_id(self, collection_name, doc_id, status):
        """Update attendance by document ID with manual override flag"""
        with self.lock:
            try:
                db = self._get_connection()
                collection = db[collection_name]
                
                obj_id = ObjectId(doc_id)
                current_time = datetime.now()
                current_time_str = current_time.strftime('%Y-%m-%d %H:%M:%S')
                
                student = collection.find_one({'_id': obj_id})
                if not student:
                    return False
                
                update_data = {
                    'Status': status,
                    'Last_Updated': current_time_str,
                    'Manual_Override': True
                }
                
                if status == 'Present':
                    update_data['Last_Seen'] = datetime.now().strftime('%H:%M:%S')
                    if student.get('First_Seen') == 'N/A':
                        update_data['First_Seen'] = datetime.now().strftime('%H:%M:%S')
                
                # Handle duration tracking
                prev_status = student.get('Status', 'Absent')
                if prev_status == 'Present' and status != 'Present':
                    present_start = student.get('_present_start_time')
                    if present_start:
                        try:
                            start_dt = datetime.strptime(present_start, '%Y-%m-%d %H:%M:%S')
                            duration = (current_time - start_dt).total_seconds()
                            total_present = student.get('_total_present_seconds', 0) + duration
                            update_data['_total_present_seconds'] = total_present
                            update_data['Total_Present_Duration'] = self._format_duration(total_present)
                        except:
                            pass
                    update_data['_present_start_time'] = None
                    update_data['_absent_start_time'] = current_time_str
                    
                elif prev_status != 'Present' and status == 'Present':
                    absent_start = student.get('_absent_start_time')
                    if absent_start:
                        try:
                            start_dt = datetime.strptime(absent_start, '%Y-%m-%d %H:%M:%S')
                            duration = (current_time - start_dt).total_seconds()
                            total_absent = student.get('_total_absent_seconds', 0) + duration
                            update_data['_total_absent_seconds'] = total_absent
                            update_data['Total_Absent_Duration'] = self._format_duration(total_absent)
                        except:
                            pass
                    update_data['_absent_start_time'] = None
                    update_data['_present_start_time'] = current_time_str
                
                result = collection.update_one(
                    {'_id': obj_id},
                    {'$set': update_data}
                )
                
                return result.modified_count > 0
                
            except Exception as e:
                print(f"Error updating attendance by ID: {e}")
                import traceback
                traceback.print_exc()
                return False

    def reset_manual_overrides(self, collection_name):
        """Reset all manual overrides in a collection"""
        with self.lock:
            try:
                db = self._get_connection()
                collection = db[collection_name]
                
                result = collection.update_many(
                    {'Manual_Override': True},
                    {'$set': {'Manual_Override': False}}
                )
                
                print(f"✓ Reset {result.modified_count} manual overrides in {collection_name}")
                return result.modified_count
                
            except Exception as e:
                print(f"Error resetting manual overrides: {e}")
                return 0
    
    def clear_session_data(self, collection_name):
        """Reset all students in current session to default (Absent) state"""
        with self.lock:
            try:
                db = self._get_connection()
                collection = db[collection_name]
                
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                result = collection.update_many(
                    {},
                    {'$set': {
                        'Status': 'Absent',
                        'First_Seen': 'N/A',
                        'Last_Seen': 'N/A',
                        'Absence_Timer_Start': 'N/A',
                        'Temp_Absent_Time': 'N/A',
                        'Perm_Absent_Time': 'N/A',
                        'Total_Absent_Duration': '0 sec',
                        'Total_Present_Duration': '0 sec',
                        'Last_Updated': current_time,
                        'Manual_Override': False,
                        '_present_start_time': None,
                        '_absent_start_time': None,
                        '_total_present_seconds': 0,
                        '_total_absent_seconds': 0
                    }}
                )
                
                print(f"✓ Cleared data for {result.modified_count} students in {collection_name}")
                return result.modified_count
                
            except Exception as e:
                print(f"Error clearing session data: {e}")
                return 0

    def get_all_lectures(self):
        try:
            with self.lock:
                db = self._get_connection()
                lectures = list(db.lecture_metadata.find().sort('created_at', -1))
                for lecture in lectures:
                    lecture['_id'] = str(lecture['_id'])
                    if 'created_at' in lecture:
                        lecture['created_at'] = lecture['created_at'].strftime('%Y-%m-%d %H:%M:%S')
                return lectures
        except Exception as e:
            print(f"Error getting lectures: {e}")
            return []
    
    def get_lecture_attendance(self, collection_name):
        try:
            with self.lock:
                db = self._get_connection()
                collection = db[collection_name]
                attendance_data = list(collection.find())
                for record in attendance_data:
                    record['_id'] = str(record['_id'])
                return attendance_data
        except Exception as e:
            print(f"Error getting attendance data: {e}")
            return []
    
    def get_attendance_summary(self, collection_name):
        try:
            with self.lock:
                db = self._get_connection()
                collection = db[collection_name]
                total = collection.count_documents({})
                present = collection.count_documents({'Status': 'Present'})
                temp_absent = collection.count_documents({'Status': 'Temporary Absent'})
                perm_absent = collection.count_documents({'Status': 'Permanently Absent'})
                absent = collection.count_documents({'Status': 'Absent'})
                return {
                    'total': total,
                    'present': present,
                    'temporary_absent': temp_absent,
                    'permanently_absent': perm_absent,
                    'absent': absent,
                    'attendance_percentage': round((present / total * 100), 2) if total > 0 else 0
                }
        except Exception as e:
            print(f"Error getting summary: {e}")
            return {}
    
    def get_student_attendance_history(self, student_identifier, search_field='name'):
        try:
            with self.lock:
                db = self._get_connection()
                all_lectures = list(db.lecture_metadata.find().sort('created_at', -1))
                history = []
                for lecture in all_lectures:
                    collection = db[lecture['table_name']]
                    roll_field, name_field = self._find_student_fields(collection)
                    
                    if roll_field and name_field:
                        field_to_search = roll_field if search_field == 'roll' else name_field
                        student = collection.find_one({
                            field_to_search: {'$regex': f'^{student_identifier}$', '$options': 'i'}
                        })
                        if student:
                            history.append({
                                'date': lecture['date'],
                                'session': lecture['session_name'],
                                'status': student.get('Status', 'N/A'),
                                'first_seen': student.get('First_Seen', 'N/A'),
                                'last_seen': student.get('Last_Seen', 'N/A')
                            })
                return history
        except Exception as e:
            print(f"Error getting student history: {e}")
            return []
    
    def get_daily_attendance_report(self, date_str):
        try:
            with self.lock:
                db = self._get_connection()
                lectures = list(db.lecture_metadata.find({'date': date_str}).sort('created_at', 1))
                report = []
                for lecture in lectures:
                    collection = db[lecture['table_name']]
                    attendance_data = list(collection.find())
                    for record in attendance_data:
                        record['_id'] = str(record['_id'])
                    summary = self.get_attendance_summary(lecture['table_name'])
                    report.append({
                        'session_name': lecture['session_name'],
                        'start_time': lecture['start_time'],
                        'collection_name': lecture['table_name'],
                        'summary': summary,
                        'attendance': attendance_data
                    })
                return report
        except Exception as e:
            print(f"Error getting daily report: {e}")
            return []
    
    def close(self):
        if self.client:
            self.client.close()


class AttendanceSystem:
    def __init__(self, mode=AttendanceConfig.MODE_NAME):
        self.mode = mode
        self.config = AttendanceConfig()
        self.student_status = {}
        self.stop_event = Event()
        self.last_recognized_faces = {}
        self.attendance_count = 0
        self.total_students = 73
        self.search_mode = 'roll' if mode == AttendanceConfig.MODE_ROLL_NO else 'name'
        self.current_faces_count = 0
        self.db_manager = DatabaseManager(MONGODB_CONFIG)
        self.template_headers, self.template_data = self.db_manager.load_template_from_excel(
            self.config.TEMPLATE_FILE, 
            sheet_name='Sheet1'
        )
        self.current_collection = None
        self.current_session = None
        self.previous_session = None
        self.class_names, self.known_encodings = self._load_training_data()
    
    def _load_training_data(self):
        path = '../Training_images/Name' if self.mode == self.config.MODE_NAME else '../Training_images/Roll No.'
        if not os.path.exists(path):
            raise FileNotFoundError(f"Training images directory not found: {path}")
        images = []
        class_names = []
        image_files = [f for f in os.listdir(path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
        print(f"Loading {len(image_files)} training images from {path}")
        for filename in image_files:
            img_path = os.path.join(path, filename)
            img = cv2.imread(img_path)
            if img is None:
                print(f"Warning: Could not load image {filename}")
                continue
            images.append(img)
            class_names.append(os.path.splitext(filename)[0])
        if not images:
            raise ValueError("No valid training images found")
        print(f"✓ Loaded {len(class_names)} training images")
        encodings = self._find_encodings(images)
        print(f"✓ Generated {len(encodings)} face encodings")
        return class_names, encodings
    
    def _find_encodings(self, images):
        encode_list = []
        for idx, img in enumerate(images):
            try:
                img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                encodings = face_recognition.face_encodings(img_rgb)
                if encodings:
                    encode_list.append(encodings[0])
                else:
                    print(f"Warning: No face detected in image {idx}")
            except Exception as e:
                print(f"Error encoding image {idx}: {e}")
        return encode_list
    
    def mark_attendance(self, identifier):
        identifier = identifier.upper()
        current_time = datetime.now()
        current_time_str = current_time.strftime('%H:%M:%S')
        if not self.current_collection:
            return False
        if identifier not in self.student_status:
            self.student_status[identifier] = {
                'last_seen': current_time,
                'first_seen': current_time,
                'status': 'Present',
                'timer_start': None
            }
            self.attendance_count += 1
        else:
            self.student_status[identifier]['last_seen'] = current_time
            self.student_status[identifier]['status'] = 'Present'
            self.student_status[identifier]['timer_start'] = None
        success = self.db_manager.update_attendance(
            self.current_collection,
            identifier,
            'Present',
            current_time_str,
            None,
            self.search_mode,
            manual=False
        )
        if success:
            print(f"✓ Marked {identifier} as Present at {current_time_str}")
        return success
    
    def check_absence_continuously(self):
        while not self.stop_event.is_set():
            if not self.current_collection:
                sleep(self.config.ABSENCE_CHECK_INTERVAL)
                continue
            current_time = datetime.now()
            for identifier, info in list(self.student_status.items()):
                last_seen = info['last_seen']
                current_status = info['status']
                timer_start = info['timer_start']
                time_since_seen = current_time - last_seen
                if current_status == 'Present':
                    if time_since_seen >= timedelta(seconds=self.config.ABSENCE_DETECTION_DELAY):
                        if timer_start is None:
                            timer_start = current_time
                            self.student_status[identifier]['timer_start'] = timer_start
                            print(f"⏱ Started absence timer for {identifier}")
                        time_in_absence = current_time - timer_start
                        if time_in_absence >= timedelta(seconds=self.config.PERMANENT_ABSENT_THRESHOLD):
                            self.student_status[identifier]['status'] = 'Permanently Absent'
                            self.db_manager.update_attendance(
                                self.current_collection,
                                identifier,
                                'Permanently Absent',
                                last_seen.strftime('%H:%M:%S'),
                                timer_start.strftime('%H:%M:%S'),
                                self.search_mode
                            )
                            print(f"❌ {identifier} marked as Permanently Absent")
                        elif time_in_absence >= timedelta(seconds=self.config.TEMPORARY_ABSENT_THRESHOLD):
                            if current_status != 'Temporary Absent':
                                self.student_status[identifier]['status'] = 'Temporary Absent'
                                self.db_manager.update_attendance(
                                    self.current_collection,
                                    identifier,
                                    'Temporary Absent',
                                    last_seen.strftime('%H:%M:%S'),
                                    timer_start.strftime('%H:%M:%S'),
                                    self.search_mode
                                )
                                print(f"⚠ {identifier} marked as Temporary Absent")
            sleep(self.config.ABSENCE_CHECK_INTERVAL)
    
    def process_frame(self, frame):
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
        self.current_faces_count = len(face_locations)
        session = get_current_session()
        
        if session != self.current_session:
            print(f"\n📚 Session Changed: {self.current_session} -> {session}")
            
            if self.current_collection and self.previous_session:
                self.db_manager.reset_manual_overrides(self.current_collection)
            
            self.previous_session = self.current_session
            self.current_session = session
            self.last_recognized_faces = {}
            self.student_status = {}
            self.attendance_count = 0
            
            if session:
                collection_name = self.db_manager.get_current_lecture_collection(session)
                if not collection_name:
                    collection_name = self.db_manager.create_lecture_collection(
                        session,
                        self.template_headers,
                        self.template_data
                    )
                self.current_collection = collection_name
                print(f"Using collection: {collection_name}")
        
        if session:
            for face_encoding, face_loc in zip(face_encodings, face_locations):
                matches = face_recognition.compare_faces(self.known_encodings, face_encoding, tolerance=0.6)
                face_distances = face_recognition.face_distance(self.known_encodings, face_encoding)
                if len(face_distances) > 0:
                    best_match_idx = np.argmin(face_distances)
                    if matches[best_match_idx]:
                        name = self.class_names[best_match_idx].upper()
                        color = (0, 255, 0)
                        self.mark_attendance(name)
                    else:
                        name = "UNKNOWN"
                        color = (0, 0, 255)
                else:
                    name = "UNKNOWN"
                    color = (0, 0, 255)
                y1, x2, y2, x1 = [coord * 4 for coord in face_loc]
                cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                cv2.rectangle(frame, (x1, y2 - 35), (x2, y2), color, cv2.FILLED)
                cv2.putText(frame, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        
        overlay = frame.copy()
        cv2.rectangle(overlay, (0, 0), (frame.shape[1], 100), (0, 0, 0), -1)
        cv2.addWeighted(overlay, 0.7, frame, 0.3, 0, frame)
        session_text = f"Session: {session if session else 'Not Active'}"
        cv2.putText(frame, session_text, (20, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        attendance_text = f"Total Attendance: {self.attendance_count}/{self.total_students}"
        cv2.putText(frame, attendance_text, (20, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0) if self.attendance_count > 0 else (255, 255, 255), 2)
        faces_text = f"Current Faces: {self.current_faces_count}"
        cv2.putText(frame, faces_text, (20, 90), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255) if self.current_faces_count > 0 else (150, 150, 150), 2)
        return frame
    
    def stop(self):
        self.stop_event.set()


# ============= EXCEL EXPORT FUNCTIONALITY =============

def generate_daily_report_excel(date_str):
    """Generate Excel report for a specific date"""
    try:
        db_manager = DatabaseManager(MONGODB_CONFIG)
        report_data = db_manager.get_daily_attendance_report(date_str)
        
        if not report_data:
            return None
        
        wb = Workbook()
        wb.remove(wb.active)
        
        summary_sheet = wb.create_sheet("Daily Summary", 0)
        summary_sheet['A1'] = f"Attendance Report - {date_str}"
        summary_sheet['A1'].font = Font(bold=True, size=14)
        
        row = 3
        summary_sheet['A' + str(row)] = "Session"
        summary_sheet['B' + str(row)] = "Start Time"
        summary_sheet['C' + str(row)] = "Total"
        summary_sheet['D' + str(row)] = "Present"
        summary_sheet['E' + str(row)] = "Absent"
        summary_sheet['F' + str(row)] = "Attendance %"
        
        for cell in summary_sheet[row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        row += 1
        for session in report_data:
            summary_sheet['A' + str(row)] = session['session_name']
            summary_sheet['B' + str(row)] = session['start_time']
            summary_sheet['C' + str(row)] = session['summary']['total']
            summary_sheet['D' + str(row)] = session['summary']['present']
            summary_sheet['E' + str(row)] = session['summary']['absent']
            summary_sheet['F' + str(row)] = f"{session['summary']['attendance_percentage']}%"
            row += 1
        
        for session in report_data:
            sheet = wb.create_sheet(session['session_name'][:31])
            attendance = session['attendance']
            
            if attendance:
                headers = [k for k in attendance[0].keys() if k != '_id' and not k.startswith('_')]
                for col_idx, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                
                for row_idx, record in enumerate(attendance, 2):
                    for col_idx, header in enumerate(headers, 1):
                        sheet.cell(row=row_idx, column=col_idx, value=str(record.get(header, '')))
        
        db_manager.close()
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    except Exception as e:
        print(f"Error generating Excel report: {e}")
        return None


# ============= FLASK ROUTES =============

@app.route('/api/routes', methods=['GET'])
def list_routes():
    """Debug endpoint to list all available routes"""
    routes = []
    for rule in app.url_map.iter_rules():
        routes.append({
            'endpoint': rule.endpoint,
            'methods': list(rule.methods),
            'path': str(rule)
        })
    return jsonify({'routes': routes})

@app.after_request
def after_request(response):
    """Ensure CORS headers are set"""
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/student')
def student_view():
    return render_template_string(STUDENT_HTML_TEMPLATE)

@app.route('/reports')
def reports_view():
    return render_template_string(REPORTS_HTML_TEMPLATE)

@app.route('/api/health', methods=['GET'])
def health_check():
    global attendance_system, camera_running
    try:
        client = MongoClient(MONGODB_CONFIG['host'], MONGODB_CONFIG['port'], serverSelectionTimeoutMS=2000)
        client.server_info()
        db_connected = True
        client.close()
    except:
        db_connected = False
    return jsonify({
        'status': 'healthy' if db_connected else 'degraded',
        'database': 'connected' if db_connected else 'disconnected',
        'camera_status': 'running' if camera_running else 'stopped',
        'system_initialized': attendance_system is not None,
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/camera/start', methods=['POST'])
def start_camera():
    global attendance_system, camera_running
    try:
        data = request.get_json()
        mode = data.get('mode', 1)
        if camera_running:
            return jsonify({'success': False, 'message': 'Camera is already running'}), 400
        attendance_system = AttendanceSystem(mode=mode)
        absence_thread = threading.Thread(target=attendance_system.check_absence_continuously)
        absence_thread.daemon = True
        absence_thread.start()
        camera_running = True
        return jsonify({
            'success': True,
            'message': 'Camera started successfully',
            'mode': 'Name' if mode == 1 else 'Roll Number'
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/camera/stop', methods=['POST'])
def stop_camera():
    global attendance_system, camera_running
    try:
        if not camera_running:
            return jsonify({'success': False, 'message': 'Camera is not running'}), 400
        camera_running = False
        if attendance_system:
            attendance_system.stop()
        return jsonify({'success': True, 'message': 'Camera stopped successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/camera/status', methods=['GET'])
def camera_status():
    global camera_running, attendance_system
    status = {
        'running': camera_running,
        'attendance_count': 0,
        'current_faces': 0,
        'current_session': None
    }
    if attendance_system:
        status['attendance_count'] = attendance_system.attendance_count
        status['current_faces'] = attendance_system.current_faces_count
        status['current_session'] = attendance_system.current_session
    return jsonify(status)

def generate_frames():
    global attendance_system, camera_running
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("Error: Could not open camera")
        return
    try:
        while camera_running:
            ret, frame = cap.read()
            if not ret:
                print("Error: Failed to read frame")
                break
            if attendance_system:
                frame = attendance_system.process_frame(frame)
            ret, buffer = cv2.imencode('.jpg', frame)
            if not ret:
                continue
            frame = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
    except Exception as e:
        print(f"Error in video stream: {e}")
    finally:
        cap.release()

@app.route('/api/video_feed')
def video_feed():
    if not camera_running:
        return jsonify({'error': 'Camera not running'}), 400
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/api/lectures', methods=['GET'])
def get_lectures():
    try:
        if attendance_system:
            lectures = attendance_system.db_manager.get_all_lectures()
        else:
            db_manager = DatabaseManager(MONGODB_CONFIG)
            lectures = db_manager.get_all_lectures()
            db_manager.close()
        return jsonify({'success': True, 'count': len(lectures), 'data': lectures})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/lectures/<collection_name>', methods=['GET'])
def get_lecture_attendance(collection_name):
    try:
        if attendance_system:
            attendance = attendance_system.db_manager.get_lecture_attendance(collection_name)
        else:
            db_manager = DatabaseManager(MONGODB_CONFIG)
            attendance = db_manager.get_lecture_attendance(collection_name)
            db_manager.close()
        return jsonify({'success': True, 'collection_name': collection_name, 'count': len(attendance), 'data': attendance})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/lectures/<collection_name>/summary', methods=['GET'])
def get_lecture_summary(collection_name):
    try:
        if attendance_system:
            summary = attendance_system.db_manager.get_attendance_summary(collection_name)
        else:
            db_manager = DatabaseManager(MONGODB_CONFIG)
            summary = db_manager.get_attendance_summary(collection_name)
            db_manager.close()
        return jsonify({'success': True, 'collection_name': collection_name, 'summary': summary})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/current-session', methods=['GET'])
def get_current_session_data():
    try:
        if attendance_system and attendance_system.current_collection:
            attendance = attendance_system.db_manager.get_lecture_attendance(attendance_system.current_collection)
            summary = attendance_system.db_manager.get_attendance_summary(attendance_system.current_collection)
            return jsonify({
                'success': True,
                'active': True,
                'collection_name': attendance_system.current_collection,
                'session_name': attendance_system.current_session,
                'summary': summary,
                'attendance': attendance
            })
        else:
            current_session = get_current_session()
            db_manager = DatabaseManager(MONGODB_CONFIG)
            collection_name = db_manager.get_current_lecture_collection(current_session)
            
            if collection_name:
                attendance = db_manager.get_lecture_attendance(collection_name)
                summary = db_manager.get_attendance_summary(collection_name)
                db_manager.close()
                return jsonify({
                    'success': True,
                    'active': True,
                    'collection_name': collection_name,
                    'session_name': current_session,
                    'summary': summary,
                    'attendance': attendance
                })
            else:
                db_manager.close()
                return jsonify({'success': True, 'active': False, 'message': 'No active session'})
    except Exception as e:
        print(f"Error in get_current_session_data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/attendance/update', methods=['POST'])
def update_attendance_manual():
    """Manual attendance update endpoint"""
    try:
        data = request.get_json()
        
        collection_name = data.get('collection_name')
        doc_id = data.get('doc_id')
        status = data.get('status')
        
        print(f"Update request - Collection: {collection_name}, Doc: {doc_id}, Status: {status}")
        
        if not all([collection_name, doc_id, status]):
            return jsonify({
                'success': False, 
                'error': 'Missing required fields'
            }), 400
        
        if attendance_system:
            success = attendance_system.db_manager.update_attendance_by_id(
                collection_name, doc_id, status
            )
        else:
            db_manager = DatabaseManager(MONGODB_CONFIG)
            success = db_manager.update_attendance_by_id(collection_name, doc_id, status)
            db_manager.close()
        
        if success:
            return jsonify({
                'success': True, 
                'message': f'Successfully marked as {status}'
            })
        else:
            return jsonify({
                'success': False, 
                'error': 'Failed to update - no document found or already in this state'
            }), 500
            
    except Exception as e:
        print(f"Error in update_attendance_manual: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/attendance/clear', methods=['POST'])
def clear_attendance_data():
    """Clear all attendance data for current session"""
    try:
        data = request.get_json()
        collection_name = data.get('collection_name')
        
        if not collection_name:
            return jsonify({
                'success': False,
                'error': 'Collection name is required'
            }), 400
        
        if attendance_system:
            attendance_system.student_status = {}
            attendance_system.attendance_count = 0
            count = attendance_system.db_manager.clear_session_data(collection_name)
        else:
            db_manager = DatabaseManager(MONGODB_CONFIG)
            count = db_manager.clear_session_data(collection_name)
            db_manager.close()
        
        return jsonify({
            'success': True,
            'message': f'Cleared data for {count} students',
            'count': count
        })
        
    except Exception as e:
        print(f"Error in clear_attendance_data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reports/daily/<date_str>', methods=['GET'])
def get_daily_report(date_str):
    try:
        db_manager = DatabaseManager(MONGODB_CONFIG)
        report = db_manager.get_daily_attendance_report(date_str)
        db_manager.close()
        return jsonify({'success': True, 'date': date_str, 'data': report})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reports/export/<date_str>', methods=['GET'])
def export_daily_report(date_str):
    try:
        excel_file = generate_daily_report_excel(date_str)
        if excel_file:
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'attendance_report_{date_str}.xlsx'
            )
        else:
            return jsonify({'success': False, 'error': 'No data available for this date'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/student/<identifier>', methods=['GET'])
def get_student_history(identifier):
    try:
        search_field = request.args.get('search_by', 'name')
        db_manager = DatabaseManager(MONGODB_CONFIG)
        history = db_manager.get_student_attendance_history(identifier, search_field)
        db_manager.close()
        
        total_sessions = len(history)
        present_count = sum(1 for h in history if h['status'] == 'Present')
        attendance_percentage = (present_count / total_sessions * 100) if total_sessions > 0 else 0
        
        return jsonify({
            'success': True,
            'identifier': identifier,
            'statistics': {
                'total_sessions': total_sessions,
                'present': present_count,
                'absent': total_sessions - present_count,
                'attendance_percentage': round(attendance_percentage, 2)
            },
            'history': history
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============= HTML TEMPLATES =============

HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Attendance System V2 - Enhanced Layout</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        .container { max-width: 1800px; margin: 0 auto; }
        .header {
            background: white;
            padding: 25px 30px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            margin-bottom: 20px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .header h1 { color: #667eea; font-size: 2em; }
        .nav-links { display: flex; gap: 15px; }
        .nav-links a {
            padding: 10px 20px;
            background: #667eea;
            color: white;
            text-decoration: none;
            border-radius: 8px;
            font-weight: bold;
            transition: background 0.3s;
            font-size: 0.9em;
        }
        .nav-links a:hover { background: #5568d3; }
        
        /* NEW LAYOUT - Camera + Controls combined */
        .camera-section {
            display: grid;
            grid-template-columns: 70% 30%;
            gap: 20px;
            margin-bottom: 20px;
            background: white;
            padding: 25px;
            border-radius: 15px;
            box-shadow: 0 5px 15px rgba(0,0,0,0.1);
        }
        
        .video-panel {
            display: flex;
            flex-direction: column;
        }
        
        .video-container {
            position: relative;
            width: 100%;
            background: #000;
            border-radius: 10px;
            overflow: hidden;
            height: 500px;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .video-container img { 
            max-width: 100%; 
            max-height: 100%; 
            width: auto;
            height: auto;
            display: block; 
            object-fit: contain;
        }
        .video-placeholder {
            width: 100%;
            height: 100%;
            background: linear-gradient(135deg, #2d3748 0%, #1a202c 100%);
            display: flex;
            align-items: center;
            justify-content: center;
            color: #9ca3af;
            font-size: 1.2em;
            border-radius: 10px;
            text-align: center;
            padding: 20px;
        }
        
        .control-panel {
            display: flex;
            flex-direction: column;
            gap: 15px;
            height: 100%;
        }
        
        .control-panel h2 {
            color: #333;
            font-size: 1.3em;
            margin-bottom: 5px;
        }
        
        .form-group {
            display: flex;
            flex-direction: column;
            gap: 8px;
        }
        
        .form-group label {
            font-weight: bold;
            color: #555;
            font-size: 0.9em;
        }
        
        .form-group select,
        .form-group input {
            padding: 10px;
            border: 2px solid #ddd;
            border-radius: 8px;
            font-size: 1em;
            background: white;
        }
        
        .btn {
            padding: 12px 20px;
            border: none;
            border-radius: 8px;
            font-size: 1em;
            font-weight: bold;
            cursor: pointer;
            transition: all 0.3s ease;
            width: 100%;
        }
        .btn-start { background: #10b981; color: white; }
        .btn-start:hover:not(:disabled) { background: #059669; transform: translateY(-2px); }
        .btn-stop { background: #ef4444; color: white; }
        .btn-stop:hover:not(:disabled) { background: #dc2626; transform: translateY(-2px); }
        .btn:disabled { background: #9ca3af; cursor: not-allowed; opacity: 0.6; }
        
        .status-indicator {
            display: flex;
            align-items: center;
            gap: 10px;
            padding: 12px;
            background: #f3f4f6;
            border-radius: 10px;
            font-weight: bold;
        }
        .status-dot {
            width: 14px;
            height: 14px;
            border-radius: 50%;
            background: #ef4444;
        }
        .status-dot.active {
            background: #10b981;
            animation: pulse 2s infinite;
        }
        @keyframes pulse {
            0%, 100% { opacity: 1; }
            50% { opacity: 0.5; }
        }
        
        .info-cards {
            display: grid;
            grid-template-columns: 1fr 1fr 1fr;
            gap: 10px;
            margin-top: auto;
        }
        
        .info-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 12px;
            border-radius: 10px;
            color: white;
            text-align: center;
        }
        .info-card h3 {
            font-size: 0.75em;
            opacity: 0.9;
            margin-bottom: 5px;
        }
        .info-card .value {
            font-size: 1.4em;
            font-weight: bold;
            word-break: break-word;
        }
        
        /* Attendance Section */
        .attendance-section {
            background: white;
            padding: 25px;
            border-radius: 15px;
            box-shadow: 0 5px 15px rgba(0,0,0,0.1);
        }
        
        .stats-grid {
            display: grid;
            grid-template-columns: repeat(5, 1fr);
            gap: 15px;
            margin-bottom: 20px;
        }
        .stat-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            border-radius: 10px;
            color: white;
            text-align: center;
        }
        .stat-card h3 {
            font-size: 0.75em;
            text-transform: uppercase;
            margin-bottom: 8px;
            opacity: 0.9;
        }
        .stat-card .value {
            font-size: 2em;
            font-weight: bold;
        }
        
        .section-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 20px;
        }
        
        .section-header h2 {
            color: #333;
            font-size: 1.5em;
        }
        
        .action-buttons {
            display: flex;
            gap: 10px;
        }
        
        .btn-refresh {
            background: #667eea;
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: 8px;
            cursor: pointer;
            font-weight: bold;
            transition: background 0.3s;
        }
        .btn-refresh:hover { background: #5568d3; }
        
        .btn-clear {
            background: #f59e0b;
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: 8px;
            cursor: pointer;
            font-weight: bold;
            transition: background 0.3s;
        }
        .btn-clear:hover { background: #d97706; }
        
        .table-container {
            overflow-x: auto;
            max-height: 500px;
            overflow-y: auto;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            font-size: 0.85em;
        }
        table th {
            background: #f8f9fa;
            padding: 10px 8px;
            text-align: left;
            font-weight: bold;
            color: #333;
            position: sticky;
            top: 0;
            z-index: 10;
            font-size: 0.9em;
        }
        table td {
            padding: 10px 8px;
            border-bottom: 1px solid #eee;
        }
        table tr:hover { background: #f8f9fa; }
        .status-badge {
            padding: 3px 10px;
            border-radius: 20px;
            font-size: 0.8em;
            font-weight: bold;
            display: inline-block;
        }
        .status-badge.present {
            background: #d1fae5;
            color: #065f46;
        }
        .status-badge.absent {
            background: #fee2e2;
            color: #991b1b;
        }
        .status-badge.temporary-absent {
            background: #fef3c7;
            color: #92400e;
        }
        .status-badge.permanently-absent {
            background: #fee2e2;
            color: #991b1b;
        }
        .status-toggle {
            padding: 4px 8px;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 0.75em;
            font-weight: bold;
            transition: all 0.3s;
        }
        .status-toggle.to-present {
            background: #10b981;
            color: white;
        }
        .status-toggle.to-absent {
            background: #ef4444;
            color: white;
        }
        .status-toggle:hover {
            transform: scale(1.05);
        }
        .manual-override-badge {
            background: #fbbf24;
            color: #78350f;
            padding: 2px 6px;
            border-radius: 4px;
            font-size: 0.7em;
            margin-left: 5px;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div>
                <h1>Enhanced Attendance System V2</h1>
                <p>Real-time Face Recognition with Complete Time Tracking</p>
            </div>
            <div class="nav-links">
                <a href="/">Dashboard</a>
                <a href="/reports">Reports</a>
                <a href="/student">Student View</a>
            </div>
        </div>

        <!-- NEW COMBINED CAMERA + CONTROL SECTION -->
        <div class="camera-section">
            <div class="video-panel">
                <h2 style="margin-bottom: 15px; color: #333;">Live Camera Feed</h2>
                <div class="video-container" id="videoContainer">
                    <div class="video-placeholder">Camera not started<br>Click "Start Camera" to begin</div>
                </div>
                <div class="info-cards" style="margin-top: 15px;">
                    <div class="info-card">
                        <h3>Current Faces</h3>
                        <div class="value" id="currentFaces">0</div>
                    </div>
                    <div class="info-card">
                        <h3>Session</h3>
                        <div class="value" style="font-size: 1.2em;" id="currentSession">None</div>
                    </div>
                    <div class="info-card">
                        <h3>Teacher/Professor</h3>
                        <div class="value" style="font-size: 1.2em;" id="teacherName">-</div>
                    </div>
                </div>
            </div>

            <div class="control-panel">
                <h2>Camera Control</h2>
                
                <div class="form-group">
                    <label>Classroom</label>
                    <input type="text" id="classroom" placeholder="e.g., Room 101, Lab A">
                </div>

                <div class="form-group">
                    <label>Teacher/Professor Name</label>
                    <input type="text" id="teacherInput" placeholder="e.g., Dr. Smith">
                </div>

                <div class="form-group">
                    <label>Recognition Mode</label>
                    <select id="modeSelect">
                        <option value="1">By Name</option>
                        <option value="2">By Roll Number</option>
                    </select>
                </div>

                <button class="btn btn-start" id="startBtn" onclick="startCamera()">Start Camera</button>
                <button class="btn btn-stop" id="stopBtn" onclick="stopCamera()" disabled>Stop Camera</button>

                <div class="status-indicator">
                    <div class="status-dot" id="statusDot"></div>
                    <span id="statusText">Camera Stopped</span>
                </div>

                <div style="border-top: 2px solid #e5e7eb; padding-top: 15px; margin-top: auto;">
                    <h3 style="color: #555; font-size: 1em; margin-bottom: 10px;">Quick Stats</h3>
                    <div style="display: flex; flex-direction: column; gap: 8px;">
                        <div style="display: flex; justify-content: space-between; padding: 8px; background: #f9fafb; border-radius: 6px;">
                            <span style="font-weight: 600; color: #666;">Total Students:</span>
                            <span id="quickTotal" style="font-weight: bold; color: #667eea;">-</span>
                        </div>
                        <div style="display: flex; justify-content: space-between; padding: 8px; background: #f9fafb; border-radius: 6px;">
                            <span style="font-weight: 600; color: #666;">Present:</span>
                            <span id="quickPresent" style="font-weight: bold; color: #10b981;">-</span>
                        </div>
                        <div style="display: flex; justify-content: space-between; padding: 8px; background: #f9fafb; border-radius: 6px;">
                            <span style="font-weight: 600; color: #666;">Absent:</span>
                            <span id="quickAbsent" style="font-weight: bold; color: #ef4444;">-</span>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- ATTENDANCE SECTION -->
        <div class="attendance-section">
            <div class="section-header">
                <h2>Current Session Attendance</h2>
                <div class="action-buttons">
                    <button class="btn-refresh" onclick="refreshData()">Refresh Data</button>
                    <button class="btn-clear" onclick="clearSessionData()" id="clearBtn">Clear Data</button>
                </div>
            </div>

            <div class="stats-grid">
                <div class="stat-card">
                    <h3>Total Students</h3>
                    <div class="value" id="totalStudents">-</div>
                </div>
                <div class="stat-card">
                    <h3>Present</h3>
                    <div class="value" id="presentCount">-</div>
                </div>
                <div class="stat-card">
                    <h3>Absent</h3>
                    <div class="value" id="absentCount">-</div>
                </div>
                <div class="stat-card">
                    <h3>Temp Absent</h3>
                    <div class="value" id="tempAbsentCount">-</div>
                </div>
                <div class="stat-card">
                    <h3>Attendance %</h3>
                    <div class="value" id="attendancePercentage">-</div>
                </div>
            </div>

            <div class="table-container">
                <table id="attendanceTable">
                    <thead>
                        <tr>
                            <th>Roll No</th>
                            <th>Name</th>
                            <th>Status</th>
                            <th>First Seen</th>
                            <th>Last Seen</th>
                            <th>Absence Timer</th>
                            <th>Temp Absent Time</th>
                            <th>Perm Absent Time</th>
                            <th>Present Duration</th>
                            <th>Absent Duration</th>
                            <th>Action</th>
                        </tr>
                    </thead>
                    <tbody id="attendanceBody">
                        <tr><td colspan="11" style="text-align: center;">Loading...</td></tr>
                    </tbody>
                </table>
            </div>
        </div>
    </div>

    <script>
        const API_BASE_URL = window.location.origin + '/api';
        let cameraRunning = false;
        let refreshInterval = null;
        let currentCollectionName = null;
        let currentSessionName = null;
        let cameraInfoInterval = null;

        async function startCamera() {
            const mode = document.getElementById('modeSelect').value;
            const classroom = document.getElementById('classroom').value.trim();
            const teacherName = document.getElementById('teacherInput').value.trim();
            const startBtn = document.getElementById('startBtn');
            const stopBtn = document.getElementById('stopBtn');
            
            if (!classroom) {
                alert('Please enter the classroom name');
                return;
            }
            
            if (!teacherName) {
                alert('Please enter the teacher/professor name');
                return;
            }
            
            // Update teacher name in info card
            document.getElementById('teacherName').textContent = teacherName;
            
            startBtn.disabled = true;
            startBtn.textContent = 'Starting...';
            try {
                const response = await fetch(`${API_BASE_URL}/camera/start`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ mode: parseInt(mode) })
                });
                const data = await response.json();
                if (data.success) {
                    cameraRunning = true;
                    updateCameraStatus(true);
                    startBtn.disabled = true;
                    stopBtn.disabled = false;
                    startBtn.textContent = 'Start Camera';
                    document.getElementById('videoContainer').innerHTML = 
                        '<img src="' + API_BASE_URL + '/video_feed?t=' + Date.now() + '" alt="Live Feed" style="max-width: 100%; max-height: 100%; object-fit: contain;">';
                    startCameraInfoRefresh();
                    // AUTO-REFRESH CURRENT SESSION ON CAMERA START
                    await loadCurrentSession();
                } else {
                    throw new Error(data.message || data.error);
                }
            } catch (error) {
                alert('Error: ' + error.message);
                startBtn.disabled = false;
                startBtn.textContent = 'Start Camera';
            }
        }

        async function stopCamera() {
            const startBtn = document.getElementById('startBtn');
            const stopBtn = document.getElementById('stopBtn');
            stopBtn.disabled = true;
            stopBtn.textContent = 'Stopping...';
            try {
                const response = await fetch(`${API_BASE_URL}/camera/stop`, { method: 'POST' });
                const data = await response.json();
                if (data.success) {
                    cameraRunning = false;
                    updateCameraStatus(false);
                    startBtn.disabled = false;
                    stopBtn.disabled = true;
                    stopBtn.textContent = 'Stop Camera';
                    document.getElementById('videoContainer').innerHTML = 
                        '<div class="video-placeholder">Camera stopped</div>';
                    stopCameraInfoRefresh();
                }
            } catch (error) {
                alert('Error: ' + error.message);
                stopBtn.disabled = false;
                stopBtn.textContent = 'Stop Camera';
            }
        }

        function updateCameraStatus(isRunning) {
            const statusDot = document.getElementById('statusDot');
            const statusText = document.getElementById('statusText');
            if (isRunning) {
                statusDot.classList.add('active');
                statusText.textContent = 'Camera Running';
            } else {
                statusDot.classList.remove('active');
                statusText.textContent = 'Camera Stopped';
            }
        }

        async function updateCameraInfo() {
            try {
                const response = await fetch(`${API_BASE_URL}/camera/status`);
                const data = await response.json();
                document.getElementById('currentFaces').textContent = data.current_faces || 0;
                document.getElementById('currentSession').textContent = data.current_session || 'None';
                
                if (data.current_session && data.current_session !== currentSessionName) {
                    console.log('Session changed detected:', currentSessionName, '->', data.current_session);
                    currentSessionName = data.current_session;
                    await loadCurrentSession();
                }
            } catch (error) {
                console.error('Error updating camera info:', error);
            }
        }

        async function loadCurrentSession() {
            try {
                const response = await fetch(`${API_BASE_URL}/current-session`);
                const data = await response.json();
                if (data.success && data.active) {
                    currentCollectionName = data.collection_name;
                    currentSessionName = data.session_name;
                    updateStats(data.summary);
                    displayAttendanceData(data.attendance);
                    document.getElementById('clearBtn').disabled = false;
                } else {
                    currentCollectionName = null;
                    currentSessionName = null;
                    document.getElementById('totalStudents').textContent = '-';
                    document.getElementById('presentCount').textContent = '-';
                    document.getElementById('absentCount').textContent = '-';
                    document.getElementById('tempAbsentCount').textContent = '-';
                    document.getElementById('attendancePercentage').textContent = '-';
                    document.getElementById('quickTotal').textContent = '-';
                    document.getElementById('quickPresent').textContent = '-';
                    document.getElementById('quickAbsent').textContent = '-';
                    document.getElementById('attendanceBody').innerHTML = 
                        '<tr><td colspan="11" style="text-align: center;">No active session</td></tr>';
                    document.getElementById('clearBtn').disabled = true;
                }
            } catch (error) {
                console.error('Error loading session:', error);
            }
        }

        function updateStats(summary) {
            document.getElementById('totalStudents').textContent = summary.total || 0;
            document.getElementById('presentCount').textContent = summary.present || 0;
            document.getElementById('absentCount').textContent = summary.absent || 0;
            document.getElementById('tempAbsentCount').textContent = summary.temporary_absent || 0;
            document.getElementById('attendancePercentage').textContent = 
                (summary.attendance_percentage || 0).toFixed(2) + '%';
            
            // Update quick stats in control panel
            document.getElementById('quickTotal').textContent = summary.total || 0;
            document.getElementById('quickPresent').textContent = summary.present || 0;
            document.getElementById('quickAbsent').textContent = summary.absent || 0;
        }

        async function toggleAttendance(docId, currentStatus) {
            if (!currentCollectionName) {
                alert('No active session');
                return;
            }
            
            const newStatus = currentStatus === 'Present' ? 'Absent' : 'Present';
            const button = event.target;
            button.disabled = true;
            button.textContent = 'Updating...';
            
            try {
                const response = await fetch(`${API_BASE_URL}/attendance/update`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({
                        collection_name: currentCollectionName,
                        doc_id: docId,
                        status: newStatus
                    })
                });
                
                const data = await response.json();
                if (data.success) {
                    await refreshData();
                } else {
                    alert('Failed to update: ' + (data.error || 'Unknown error'));
                }
            } catch (error) {
                console.error('Error toggling attendance:', error);
                alert('Error: ' + error.message);
            } finally {
                button.disabled = false;
                button.textContent = currentStatus === 'Present' ? 'Mark Absent' : 'Mark Present';
            }
        }

        async function refreshData() {
            if (!currentCollectionName) {
                await loadCurrentSession();
                return;
            }
            try {
                const response = await fetch(`${API_BASE_URL}/lectures/${currentCollectionName}`);
                const data = await response.json();
                if (data.success) {
                    const summaryResponse = await fetch(`${API_BASE_URL}/lectures/${currentCollectionName}/summary`);
                    const summaryData = await summaryResponse.json();
                    updateStats(summaryData.summary);
                    displayAttendanceData(data.data);
                }
            } catch (error) {
                console.error('Error refreshing data:', error);
            }
        }

        async function clearSessionData() {
            if (!currentCollectionName) {
                alert('No active session to clear.');
                return;
            }
            if (!confirm('Are you sure you want to clear all attendance data for this session?')) return;
            try {
                const response = await fetch(`${API_BASE_URL}/attendance/clear`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ collection_name: currentCollectionName })
                });
                const data = await response.json();
                if (data.success) {
                    alert(data.message);
                    await refreshData();
                } else {
                    alert('Failed to clear data: ' + (data.error || 'Unknown error'));
                }
            } catch (error) {
                console.error('Error clearing session data:', error);
            }
        }

        function displayAttendanceData(records) {
            const tbody = document.getElementById('attendanceBody');
            if (!records || records.length === 0) {
                tbody.innerHTML = '<tr><td colspan="11" style="text-align:center;">No records found</td></tr>';
                return;
            }

            tbody.innerHTML = '';
            records.forEach(rec => {
                const row = document.createElement('tr');
                const status = rec.Status || 'Absent';
                let badgeClass = 'absent';
                if (status === 'Present') badgeClass = 'present';
                else if (status === 'Temporary Absent') badgeClass = 'temporary-absent';
                else if (status === 'Permanently Absent') badgeClass = 'permanently-absent';

                // Find roll number field (handle different possible field names)
                const rollNo = rec['Roll No'] || rec['Roll_No'] || rec['RollNo'] || rec['Roll Number'] || '-';

                row.innerHTML = `
                    <td>${rollNo}</td>
                    <td>${rec['Name'] || '-'}</td>
                    <td><span class="status-badge ${badgeClass}">${status}</span>
                        ${rec.Manual_Override ? '<span class="manual-override-badge">Manual</span>' : ''}
                    </td>
                    <td>${rec.First_Seen || 'N/A'}</td>
                    <td>${rec.Last_Seen || 'N/A'}</td>
                    <td>${rec.Absence_Timer_Start || 'N/A'}</td>
                    <td>${rec.Temp_Absent_Time || 'N/A'}</td>
                    <td>${rec.Perm_Absent_Time || 'N/A'}</td>
                    <td>${rec.Total_Present_Duration || '0 sec'}</td>
                    <td>${rec.Total_Absent_Duration || '0 sec'}</td>
                    <td>
                        <button class="status-toggle ${status === 'Present' ? 'to-absent' : 'to-present'}"
                            onclick="toggleAttendance('${rec._id}', '${status}')">
                            ${status === 'Present' ? 'Mark Absent' : 'Mark Present'}
                        </button>
                    </td>`;
                tbody.appendChild(row);
            });
        }

        function startCameraInfoRefresh() {
            if (cameraInfoInterval) clearInterval(cameraInfoInterval);
            updateCameraInfo(); // Initial call
            cameraInfoInterval = setInterval(updateCameraInfo, 3000);
            
            // Also refresh attendance data periodically when camera is running
            if (refreshInterval) clearInterval(refreshInterval);
            refreshInterval = setInterval(refreshData, 5000);
        }

        function stopCameraInfoRefresh() {
            if (cameraInfoInterval) clearInterval(cameraInfoInterval);
            if (refreshInterval) clearInterval(refreshInterval);
        }

        // AUTO-LOAD ON PAGE LOAD AND REFRESH
        window.onload = async function() {
            updateCameraStatus(false);
            // Initial load of current session data
            await loadCurrentSession();
            // Start periodic updates for camera info
            startCameraInfoRefresh();
        };
    </script>
</body>
</html>

'''
from template import STUDENT_HTML_TEMPLATE, REPORTS_HTML_TEMPLATE
# # Student and Reports templates remain the same
# STUDENT_HTML_TEMPLATE = '''[Previous student template - unchanged]'''
# REPORTS_HTML_TEMPLATE = '''[Previous reports template - unchanged]'''

if __name__ == '__main__':
    print("=" * 70)
    print("Enhanced Face Recognition Attendance System V2")
    print("=" * 70)
    print("\nNew Features:")
    print("   ✓ Combined camera + control layout (70/30 split)")
    print("   ✓ Full time tracking implementation")
    print("   ✓ Camera start refreshes attendance data")
    print("   ✓ All timing fields visible in table")
    print("   ✓ Roll No properly displayed")
    print("\nStarting Server...")
    print(f"Database: MongoDB ({MONGODB_CONFIG['host']}:{MONGODB_CONFIG['port']})")
    print(f"Server URL: http://localhost:5000")
    print("\n" + "=" * 70 + "\n")
    
    app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)