"""
Enhanced Attendance System - FIXED VERSION
MongoDB schema and Excel configuration properly integrated
"""

from flask import Flask, jsonify, request, render_template_string, Response
from flask_cors import CORS
from pymongo import MongoClient
from datetime import datetime, timedelta
import os
import threading
import cv2
import numpy as np
import face_recognition
import sys
from time import sleep
from threading import Lock, Event
from openpyxl import load_workbook

sys.path.append(os.path.abspath('../'))
try:
    from Excel_Format import get_current_session
except ImportError:
    def get_current_session():
        current_hour = datetime.now().hour
        if 9 <= current_hour < 10:
            return "Session 1"
        elif 10 <= current_hour < 11:
            return "Session 2"
        elif 11 <= current_hour < 12:
            return "Session 3"
        elif 12 <= current_hour < 13:
            return "Session 4"
        elif 13 <= current_hour < 14:
            return "Session 5"
        elif 14 <= current_hour < 15:
            return "Session 6"
        elif 15 <= current_hour < 16:
            return "Session 7"
        else:
            return "Session 8"

app = Flask(__name__)
CORS(app)

MONGODB_CONFIG = {
    'host': 'localhost',
    'port': 27017,
    'database': 'Attendance_system'
}

TEMPLATE_FILE = 'Book2.xlsx'
ALL_SESSIONS = ["Session 1", "Session 2", "Session 3", "Session 4", 
                "Session 5", "Session 6", "Session 7", "Session 8"]

YEAR_MAPPING = {
    '2022': 'B.Tech',
    '2023': 'TY',
    '2024': 'SY',
    '2025': 'FY'
}

attendance_system = None
camera_running = False
camera_lock = Lock()


class AttendanceConfig:
    TEMPLATE_FILE = 'Book2.xlsx'
    ABSENCE_DETECTION_DELAY = 5
    TEMPORARY_ABSENT_THRESHOLD = 10
    PERMANENT_ABSENT_THRESHOLD = 15
    ABSENCE_CHECK_INTERVAL = 2
    MODE_NAME = 1
    MODE_ROLL_NO = 2


class DatabaseManager:
    """Manages MongoDB with proper schema and Excel integration"""
    
    def __init__(self, mongodb_config):
        self.mongodb_config = mongodb_config
        self.client = None
        self.db = None
        self.lock = Lock()
        self._initialize_db()
    
    def _get_connection(self):
        try:
            if self.client is None:
                self.client = MongoClient(
                    self.mongodb_config['host'], 
                    self.mongodb_config['port'],
                    serverSelectionTimeoutMS=5000
                )
                self.db = self.client[self.mongodb_config['database']]
            return self.db
        except Exception as e:
            print(f"Error connecting to MongoDB: {e}")
            raise
    
    def _initialize_db(self):
        try:
            self.db = self._get_connection()
            # Create index for metadata collection
            try:
                self.db.lecture_metadata.create_index([('collection_name', 1)], unique=True)
            except Exception:
                pass  # Index might already exist
            print("✓ MongoDB initialized successfully")
        except Exception as e:
            print(f"Error initializing database: {e}")
            raise
    
    def _get_year_code(self, year_input):
        """Convert year to code using YEAR_MAPPING"""
        return YEAR_MAPPING.get(str(year_input), 'B.Tech')
    
    def _get_sheet_name(self, department, year_input):
        """Get Excel sheet name: DEPT_YEARCODE"""
        year_code = self._get_year_code(year_input)
        return f"{department}_{year_code}"
    
    def load_students_from_excel(self, excel_file, sheet_name):
        """Load students from specific Excel sheet with proper header handling"""
        try:
            # Try multiple possible paths for the Excel file
            possible_paths = [
                excel_file,
                os.path.join(os.path.dirname(__file__), excel_file),
                os.path.join(os.path.dirname(__file__), '..', excel_file),
                os.path.join(os.path.dirname(__file__), '..', '..', excel_file),
            ]
            
            excel_path = None
            for path in possible_paths:
                if os.path.exists(path):
                    excel_path = path
                    break
            
            if excel_path is None:
                raise FileNotFoundError(
                    f"Excel file not found: {excel_file}\n"
                    f"Searched in:\n" + "\n".join(f"  - {p}" for p in possible_paths) +
                    f"\nPlease ensure Book2.xlsx is in one of these locations."
                )
            
            print(f"Loading Excel from: {excel_path}")
            wb = load_workbook(excel_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            
            sheet = wb[sheet_name]
            data = []
            headers = None
            
            for row in sheet.iter_rows(values_only=True):
                # Skip completely empty rows
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                
                row_data = [cell if cell is not None else '' for cell in row]
                
                if headers is None:
                    # First non-empty row is headers
                    headers = [str(h).strip() for h in row_data]
                else:
                    # Data rows
                    data.append(row_data)
            
            wb.close()
            print(f"✓ Loaded {len(data)} students from sheet '{sheet_name}'")
            print(f"  Headers: {headers}")
            return headers, data
        except Exception as e:
            print(f"Error loading Excel: {e}")
            raise
    
    def create_or_get_daily_collection(self, department, year_input, date_str, 
                                      classroom, teacher_name, template_headers, 
                                      template_data, camera_ids=None):
        """Create or get daily collection with proper nested schema"""
        year_code = self._get_year_code(year_input)
        collection_name = f"{department}_{year_code}_{date_str}"
        
        with self.lock:
            try:
                db = self._get_connection()
                
                # Check if collection already exists
                if collection_name in db.list_collection_names():
                    print(f"✓ Using existing collection: {collection_name}")
                    return collection_name
                
                collection = db[collection_name]
                current_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # Build sessions with nested students
                sessions = {}
                for session_name in ALL_SESSIONS:
                    students_dict = {}
                    
                    for row in template_data:
                        doc_data = {}
                        for i, header in enumerate(template_headers):
                            if i < len(row):
                                value = row[i]
                                doc_data[header] = str(value).strip() if value not in (None, '') else ''
                            else:
                                doc_data[header] = ''
                        
                        # Extract roll number with multiple header variations
                        roll_no = (doc_data.get('Roll No') or 
                                  doc_data.get('Roll NO') or 
                                  doc_data.get('Roll_No') or 
                                  doc_data.get('RollNo') or '')
                        
                        if not roll_no:
                            continue
                        
                        # Build student document
                        students_dict[roll_no] = {
                            'sr_no': (doc_data.get('Sr. No') or 
                                     doc_data.get('Sr no') or 
                                     doc_data.get('Sr_No') or ''),
                            'roll_no': roll_no,
                            'prn_no': (doc_data.get('PRN No.') or 
                                      doc_data.get('PRN no') or 
                                      doc_data.get('PRN_No') or ''),
                            'name': doc_data.get('Name') or '',
                            'status': 'Absent',
                            'timestamps': {
                                'first_seen': None,
                                'last_seen': None,
                                'present_timer_start': None,
                                'absence_timer_start': None,
                                'temp_absent_time': None,
                                'perm_absent_time': None,
                                'last_updated': current_time_str
                            },
                            'durations': {
                                'total_present_seconds': 0,
                                'total_absent_seconds': 0,
                                'total_present_human': '0 sec',
                                'total_absent_human': '0 sec'
                            },
                            'flags': {
                                'manual_override': False,
                                'is_temp_absent': False,
                                'is_perm_absent': False
                            }
                        }
                    
                    sessions[session_name] = {
                        'start_time': None,
                        'end_time': None,
                        'students': students_dict
                    }
                
                # Create daily document
                daily_doc = {
                    'date': date_str,
                    'department': department,
                    'year': year_input,
                    'year_code': year_code,
                    'classroom': classroom,
                    'teacher_name': teacher_name,
                    'camera_ids': camera_ids or [],
                    'created_at': datetime.now(),
                    'sessions': sessions
                }
                
                collection.insert_one(daily_doc)
                
                # Store metadata
                metadata = {
                    'collection_name': collection_name,
                    'date': date_str,
                    'department': department,
                    'year': year_input,
                    'year_code': year_code,
                    'classroom': classroom,
                    'teacher_name': teacher_name,
                    'created_at': datetime.now()
                }
                db.lecture_metadata.insert_one(metadata)
                
                print(f"✓ Created full day document: {collection_name}")
                print(f"  - {len(ALL_SESSIONS)} sessions initialized")
                print(f"  - {len(students_dict)} students per session")
                
                return collection_name
                
            except Exception as e:
                print(f"Error creating daily collection: {e}")
                import traceback
                traceback.print_exc()
                raise
    
    def update_student_attendance(self, collection_name, session_name, roll_no, status, manual=False):
        """Update student attendance with robust duration tracking"""
        with self.lock:
            try:
                db = self._get_connection()
                collection = db[collection_name]
                
                doc = collection.find_one({})
                if not doc or session_name not in doc.get('sessions', {}):
                    print(f"Error: Session {session_name} not found in {collection_name}")
                    return False
                
                student_path = f'sessions.{session_name}.students.{roll_no}'
                student = doc['sessions'][session_name]['students'].get(roll_no)
                
                if not student:
                    print(f"Error: Student {roll_no} not found in {session_name}")
                    return False
                
                current_time = datetime.now()
                current_time_str = current_time.strftime('%Y-%m-%d %H:%M:%S')
                
                prev_status = student.get('status', 'Absent')
                present_start = student['timestamps'].get('present_timer_start')
                absent_start = student['timestamps'].get('absence_timer_start')
                total_present = student['durations'].get('total_present_seconds', 0)
                total_absent = student['durations'].get('total_absent_seconds', 0)
                
                update_fields = {}
                
                # Finalize previous present period if leaving present
                if prev_status == 'Present' and status != 'Present':
                    if present_start:
                        try:
                            start_dt = datetime.strptime(present_start, '%Y-%m-%d %H:%M:%S')
                            duration = (current_time - start_dt).total_seconds()
                            total_present += duration
                            print(f"  [{roll_no}] Added {duration:.0f}s to present (total: {total_present:.0f}s)")
                        except Exception as e:
                            print(f"  Error calculating present duration: {e}")
                    update_fields[f'{student_path}.timestamps.present_timer_start'] = None
                    update_fields[f'{student_path}.timestamps.absence_timer_start'] = current_time_str
                    
                # Finalize previous absent period if arriving to present
                elif prev_status != 'Present' and status == 'Present':
                    if absent_start:
                        try:
                            start_dt = datetime.strptime(absent_start, '%Y-%m-%d %H:%M:%S')
                            duration = (current_time - start_dt).total_seconds()
                            total_absent += duration
                            print(f"  [{roll_no}] Added {duration:.0f}s to absent (total: {total_absent:.0f}s)")
                        except Exception as e:
                            print(f"  Error calculating absent duration: {e}")
                    update_fields[f'{student_path}.timestamps.absence_timer_start'] = None
                    update_fields[f'{student_path}.timestamps.present_timer_start'] = current_time_str
                    
                # Initialize timers if needed
                elif status == 'Present' and not present_start:
                    update_fields[f'{student_path}.timestamps.present_timer_start'] = current_time_str
                    
                elif status != 'Present' and not absent_start:
                    update_fields[f'{student_path}.timestamps.absence_timer_start'] = current_time_str
                
                # Update basic fields
                update_fields[f'{student_path}.status'] = status
                update_fields[f'{student_path}.timestamps.last_updated'] = current_time_str
                update_fields[f'{student_path}.flags.manual_override'] = manual
                
                # Update first_seen if this is first present
                if student['timestamps']['first_seen'] is None and status == 'Present':
                    update_fields[f'{student_path}.timestamps.first_seen'] = current_time_str
                
                # Update last_seen if present
                if status == 'Present':
                    update_fields[f'{student_path}.timestamps.last_seen'] = current_time.strftime('%H:%M:%S')
                
                # Update durations
                update_fields[f'{student_path}.durations.total_present_seconds'] = int(total_present)
                update_fields[f'{student_path}.durations.total_absent_seconds'] = int(total_absent)
                update_fields[f'{student_path}.durations.total_present_human'] = self._format_duration(total_present)
                update_fields[f'{student_path}.durations.total_absent_human'] = self._format_duration(total_absent)
                
                # Update flags based on status
                if status == 'Temporary Absent':
                    update_fields[f'{student_path}.timestamps.temp_absent_time'] = current_time_str
                    update_fields[f'{student_path}.flags.is_temp_absent'] = True
                elif status == 'Permanently Absent':
                    update_fields[f'{student_path}.timestamps.perm_absent_time'] = current_time_str
                    update_fields[f'{student_path}.flags.is_perm_absent'] = True
                elif status == 'Present':
                    update_fields[f'{student_path}.flags.is_temp_absent'] = False
                    update_fields[f'{student_path}.flags.is_perm_absent'] = False
                
                # Set session start_time if first update
                if doc['sessions'][session_name]['start_time'] is None:
                    update_fields[f'sessions.{session_name}.start_time'] = current_time_str
                
                result = collection.update_one({}, {'$set': update_fields})
                
                return result.modified_count > 0
                
            except Exception as e:
                print(f"Error updating student attendance: {e}")
                import traceback
                traceback.print_exc()
                return False
    
    def _format_duration(self, seconds):
        """Format duration in human-readable form"""
        try:
            seconds = int(seconds)
        except:
            seconds = 0
        
        if seconds < 60:
            return f"{seconds} sec"
        elif seconds < 3600:
            minutes = seconds // 60
            secs = seconds % 60
            return f"{minutes} min {secs} sec"
        else:
            hours = seconds // 3600
            minutes = (seconds % 3600) // 60
            return f"{hours} hr {minutes} min"
    
    def get_session_attendance(self, collection_name, session_name):
        """Get attendance for a specific session"""
        with self.lock:
            try:
                db = self._get_connection()
                collection = db[collection_name]
                doc = collection.find_one({})
                
                if not doc or session_name not in doc.get('sessions', {}):
                    return []
                
                students = doc['sessions'][session_name]['students']
                return list(students.values())
                
            except Exception as e:
                print(f"Error getting session attendance: {e}")
                return []
    
    def get_session_summary(self, collection_name, session_name):
        """Get summary statistics for a session"""
        with self.lock:
            try:
                db = self._get_connection()
                collection = db[collection_name]
                doc = collection.find_one({})
                
                if not doc or session_name not in doc.get('sessions', {}):
                    return {}
                
                students = doc['sessions'][session_name]['students']
                total = len(students)
                present = sum(1 for s in students.values() if s['status'] == 'Present')
                temp_absent = sum(1 for s in students.values() if s['status'] == 'Temporary Absent')
                perm_absent = sum(1 for s in students.values() if s['status'] == 'Permanently Absent')
                absent = sum(1 for s in students.values() if s['status'] == 'Absent')
                
                return {
                    'total': total,
                    'present': present,
                    'temporary_absent': temp_absent,
                    'permanently_absent': perm_absent,
                    'absent': absent,
                    'attendance_percentage': round((present / total * 100), 2) if total > 0 else 0
                }
                
            except Exception as e:
                print(f"Error getting summary: {e}")
                return {}
    
    def clear_session_data(self, collection_name, session_name):
        """Clear all attendance data for a session"""
        with self.lock:
            try:
                db = self._get_connection()
                collection = db[collection_name]
                doc = collection.find_one({})
                
                if not doc or session_name not in doc.get('sessions', {}):
                    return 0
                
                students = doc['sessions'][session_name]['students']
                current_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                update_fields = {}
                for roll_no in students.keys():
                    prefix = f'sessions.{session_name}.students.{roll_no}'
                    update_fields[f'{prefix}.status'] = 'Absent'
                    update_fields[f'{prefix}.timestamps.first_seen'] = None
                    update_fields[f'{prefix}.timestamps.last_seen'] = None
                    update_fields[f'{prefix}.timestamps.present_timer_start'] = None
                    update_fields[f'{prefix}.timestamps.absence_timer_start'] = None
                    update_fields[f'{prefix}.timestamps.temp_absent_time'] = None
                    update_fields[f'{prefix}.timestamps.perm_absent_time'] = None
                    update_fields[f'{prefix}.timestamps.last_updated'] = current_time_str
                    update_fields[f'{prefix}.durations.total_present_seconds'] = 0
                    update_fields[f'{prefix}.durations.total_absent_seconds'] = 0
                    update_fields[f'{prefix}.durations.total_present_human'] = '0 sec'
                    update_fields[f'{prefix}.durations.total_absent_human'] = '0 sec'
                    update_fields[f'{prefix}.flags.manual_override'] = False
                    update_fields[f'{prefix}.flags.is_temp_absent'] = False
                    update_fields[f'{prefix}.flags.is_perm_absent'] = False
                
                collection.update_one({}, {'$set': update_fields})
                
                print(f"✓ Cleared {len(students)} students in {session_name}")
                return len(students)
                
            except Exception as e:
                print(f"Error clearing session: {e}")
                return 0
    
    def close(self):
        if self.client:
            self.client.close()


class AttendanceSystem:
    """Face recognition and attendance tracking system"""
    
    def __init__(self, mode, year_input, department, classroom, teacher_name, camera_ids=None):
        self.mode = mode
        self.config = AttendanceConfig()
        self.student_status = {}
        self.stop_event = Event()
        self.attendance_count = 0
        self.total_students = 0
        self.search_mode = 'roll' if mode == AttendanceConfig.MODE_ROLL_NO else 'name'
        self.current_faces_count = 0
        
        self.year_input = year_input
        self.department = department
        self.classroom = classroom
        self.teacher_name = teacher_name
        self.camera_ids = camera_ids or []
        
        # Initialize database manager
        self.db_manager = DatabaseManager(MONGODB_CONFIG)
        
        # Load students from Excel with proper sheet mapping
        sheet_name = self.db_manager._get_sheet_name(department, year_input)
        print(f"Loading students from sheet: {sheet_name}")
        
        self.template_headers, self.template_data = self.db_manager.load_students_from_excel(
            self.config.TEMPLATE_FILE, sheet_name
        )
        self.total_students = len(self.template_data)
        
        self.current_session = None
        self.current_collection = None
        self.current_date = None
        
        # Load training data for face recognition
        self.class_names, self.known_encodings = self._load_training_data()
    
    def _load_training_data(self):
        """Load face recognition training data"""
        path = '../Training_images/Name' if self.mode == self.config.MODE_NAME else '../Training_images/Roll No.'
        if not os.path.exists(path):
            print(f"Warning: Training images directory not found: {path}")
            return [], []
        
        images = []
        class_names = []
        image_files = [f for f in os.listdir(path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
        print(f"Loading {len(image_files)} training images from {path}")
        
        for filename in image_files:
            img_path = os.path.join(path, filename)
            img = cv2.imread(img_path)
            if img is None:
                continue
            images.append(img)
            class_names.append(os.path.splitext(filename)[0])
        
        print(f"✓ Loaded {len(class_names)} training images")
        encodings = self._find_encodings(images)
        return class_names, encodings
    
    def _find_encodings(self, images):
        """Generate face encodings for training images"""
        encode_list = []
        for img in images:
            try:
                img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                encodings = face_recognition.face_encodings(img_rgb)
                if encodings:
                    encode_list.append(encodings[0])
            except:
                pass
        return encode_list
    
    def mark_attendance(self, identifier):
        """Mark attendance for identified student"""
        identifier = identifier.upper()
        current_time = datetime.now()
        
        if not self.current_session or not self.current_collection:
            return False
        
        if identifier not in self.student_status:
            self.student_status[identifier] = {
                'last_seen': current_time,
                'status': 'Present',
                'timer_start': None
            }
            self.attendance_count += 1
        else:
            self.student_status[identifier]['last_seen'] = current_time
            self.student_status[identifier]['status'] = 'Present'
            self.student_status[identifier]['timer_start'] = None
        
        success = self.db_manager.update_student_attendance(
            self.current_collection,
            self.current_session,
            identifier,
            'Present',
            manual=False
        )
        
        return success
    
    def check_absence_continuously(self):
        """Background thread to check for absences"""
        while not self.stop_event.is_set():
            if not self.current_session:
                sleep(self.config.ABSENCE_CHECK_INTERVAL)
                continue
            
            current_time = datetime.now()
            for identifier, info in list(self.student_status.items()):
                last_seen = info['last_seen']
                current_status = info['status']
                timer_start = info['timer_start']
                
                time_since_seen = current_time - last_seen
                
                if current_status == 'Present':
                    if time_since_seen >= timedelta(seconds=self.config.ABSENCE_DETECTION_DELAY):
                        if timer_start is None:
                            timer_start = current_time
                            self.student_status[identifier]['timer_start'] = timer_start
                        
                        time_in_absence = current_time - timer_start
                        
                        if time_in_absence >= timedelta(seconds=self.config.PERMANENT_ABSENT_THRESHOLD):
                            self.student_status[identifier]['status'] = 'Permanently Absent'
                            self.db_manager.update_student_attendance(
                                self.current_collection,
                                self.current_session,
                                identifier,
                                'Permanently Absent'
                            )
                        elif time_in_absence >= timedelta(seconds=self.config.TEMPORARY_ABSENT_THRESHOLD):
                            if current_status != 'Temporary Absent':
                                self.student_status[identifier]['status'] = 'Temporary Absent'
                                self.db_manager.update_student_attendance(
                                    self.current_collection,
                                    self.current_session,
                                    identifier,
                                    'Temporary Absent'
                                )
            
            sleep(self.config.ABSENCE_CHECK_INTERVAL)
    
    def process_frame(self, frame):
        """Process video frame for face recognition"""
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
        
        self.current_faces_count = len(face_locations)
        session = get_current_session()
        date_str = datetime.now().strftime('%Y-%m-%d')
        
        # Check if session or date changed
        if session != self.current_session or date_str != self.current_date:
            print(f"\nSession/Date Changed: {self.current_session} -> {session}, Date: {date_str}")
            
            self.current_session = session
            self.current_date = date_str
            self.student_status = {}
            self.attendance_count = 0
            
            self.current_collection = self.db_manager.create_or_get_daily_collection(
                self.department,
                self.year_input,
                date_str,
                self.classroom,
                self.teacher_name,
                self.template_headers,
                self.template_data,
                camera_ids=self.camera_ids
            )
            
            print(f"Using collection: {self.current_collection}")
            print(f"Current session: {session}")
        
        # Process each detected face
        for face_encoding, face_loc in zip(face_encodings, face_locations):
            matches = face_recognition.compare_faces(self.known_encodings, face_encoding, tolerance=0.6)
            face_distances = face_recognition.face_distance(self.known_encodings, face_encoding)
            
            if len(face_distances) > 0:
                best_match_idx = np.argmin(face_distances)
                if matches[best_match_idx]:
                    name = self.class_names[best_match_idx].upper()
                    color = (0, 255, 0)
                    self.mark_attendance(name)
                else:
                    name = "UNKNOWN"
                    color = (0, 0, 255)
            else:
                name = "UNKNOWN"
                color = (0, 0, 255)
            
            # Draw bounding box and label
            y1, x2, y2, x1 = [coord * 4 for coord in face_loc]
            cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
            cv2.rectangle(frame, (x1, y2 - 35), (x2, y2), color, cv2.FILLED)
            cv2.putText(frame, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        
        # Draw info overlay
        overlay = frame.copy()
        cv2.rectangle(overlay, (0, 0), (frame.shape[1], 80), (0, 0, 0), -1)
        cv2.addWeighted(overlay, 0.7, frame, 0.3, 0, frame)
        
        cv2.putText(frame, f"Session: {session}", (20, 25), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
        cv2.putText(frame, f"Attendance: {self.attendance_count}/{self.total_students}", (20, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0) if self.attendance_count > 0 else (255, 255, 255), 2)
        cv2.putText(frame, f"Faces: {self.current_faces_count}", (20, 70), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)
        
        return frame
    
    def stop(self):
        """Stop the attendance system"""
        self.stop_event.set()


# ============= FLASK ROUTES =============

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    global attendance_system, camera_running
    try:
        client = MongoClient(MONGODB_CONFIG['host'], MONGODB_CONFIG['port'], serverSelectionTimeoutMS=2000)
        client.server_info()
        db_connected = True
        client.close()
    except:
        db_connected = False
    return jsonify({
        'status': 'healthy' if db_connected else 'degraded',
        'database': 'connected' if db_connected else 'disconnected',
        'camera_status': 'running' if camera_running else 'stopped',
        'system_initialized': attendance_system is not None,
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/camera/start', methods=['POST'])
def start_camera():
    """Start camera and attendance tracking"""
    global attendance_system, camera_running
    try:
        data = request.get_json()
        mode = data.get('mode', 1)
        year_input = data.get('year', '')
        department = data.get('department', '')
        classroom = data.get('classroom', '')
        teacher_name = data.get('teacher_name', '')
        camera_ids = data.get('camera_ids', [])
        
        if not all([year_input, department, classroom, teacher_name]):
            return jsonify({'success': False, 'message': 'All fields are required'}), 400
        
        if camera_running:
            return jsonify({'success': False, 'message': 'Camera is already running'}), 400
        
        attendance_system = AttendanceSystem(
            mode=mode,
            year_input=year_input,
            department=department,
            classroom=classroom,
            teacher_name=teacher_name,
            camera_ids=camera_ids
        )
        
        absence_thread = threading.Thread(target=attendance_system.check_absence_continuously)
        absence_thread.daemon = True
        absence_thread.start()
        
        camera_running = True
        
        year_code = YEAR_MAPPING.get(year_input, 'B.Tech')
        
        return jsonify({
            'success': True,
            'message': 'Camera started successfully',
            'year_code': year_code,
            'sheet_loaded': f"{department}_{year_code}"
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/camera/stop', methods=['POST'])
def stop_camera():
    """Stop camera and attendance tracking"""
    global attendance_system, camera_running
    try:
        if not camera_running:
            return jsonify({'success': False, 'message': 'Camera is not running'}), 400
        
        camera_running = False
        if attendance_system:
            attendance_system.stop()
        
        return jsonify({'success': True, 'message': 'Camera stopped successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/camera/status', methods=['GET'])
def camera_status():
    """Get current camera and attendance status"""
    global camera_running, attendance_system
    status = {
        'running': camera_running,
        'attendance_count': 0,
        'current_faces': 0,
        'current_session': None,
        'current_collection': None
    }
    if attendance_system:
        status['attendance_count'] = attendance_system.attendance_count
        status['current_faces'] = attendance_system.current_faces_count
        status['current_session'] = attendance_system.current_session
        status['current_collection'] = attendance_system.current_collection
    return jsonify(status)

def generate_frames():
    """Generate video frames for streaming"""
    global attendance_system, camera_running
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("Error: Could not open camera")
        return
    try:
        while camera_running:
            ret, frame = cap.read()
            if not ret:
                break
            if attendance_system:
                frame = attendance_system.process_frame(frame)
            ret, buffer = cv2.imencode('.jpg', frame)
            if not ret:
                continue
            frame = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
    except Exception as e:
        print(f"Error in video stream: {e}")
    finally:
        cap.release()

@app.route('/api/video_feed')
def video_feed():
    """Video feed endpoint"""
    if not camera_running:
        return jsonify({'error': 'Camera not running'}), 400
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/api/current-session', methods=['GET'])
def get_current_session_data():
    """Get current session attendance data"""
    try:
        if attendance_system:
            if not attendance_system.current_collection or not attendance_system.current_session:
                return jsonify({'success': True, 'active': False, 'message': 'No active session'})
            
            attendance = attendance_system.db_manager.get_session_attendance(
                attendance_system.current_collection,
                attendance_system.current_session
            )
            summary = attendance_system.db_manager.get_session_summary(
                attendance_system.current_collection,
                attendance_system.current_session
            )
            
            return jsonify({
                'success': True,
                'active': True,
                'collection_name': attendance_system.current_collection,
                'session_name': attendance_system.current_session,
                'date': attendance_system.current_date,
                'summary': summary,
                'attendance': attendance
            })
        else:
            return jsonify({'success': True, 'active': False, 'message': 'System not initialized'})
    except Exception as e:
        print(f"Error in get_current_session_data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/attendance/update', methods=['POST'])
def update_attendance_manual():
    """Manual attendance update endpoint"""
    try:
        data = request.get_json()
        
        collection_name = data.get('collection_name')
        session_name = data.get('session_name')
        roll_no = data.get('roll_no')
        status = data.get('status')
        
        if not all([collection_name, session_name, roll_no, status]):
            return jsonify({'success': False, 'error': 'Missing required fields'}), 400
        
        db_manager = DatabaseManager(MONGODB_CONFIG)
        success = db_manager.update_student_attendance(
            collection_name, session_name, roll_no, status, manual=True
        )
        db_manager.close()
        
        if success:
            return jsonify({'success': True, 'message': f'Successfully marked as {status}'})
        else:
            return jsonify({'success': False, 'error': 'Failed to update'}), 500
            
    except Exception as e:
        print(f"Error in update_attendance_manual: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/attendance/clear', methods=['POST'])
def clear_attendance_data():
    """Clear all attendance data for current session"""
    try:
        data = request.get_json()
        collection_name = data.get('collection_name')
        session_name = data.get('session_name')
        
        if not all([collection_name, session_name]):
            return jsonify({'success': False, 'error': 'Missing required fields'}), 400
        
        if attendance_system:
            attendance_system.student_status = {}
            attendance_system.attendance_count = 0
            count = attendance_system.db_manager.clear_session_data(collection_name, session_name)
        else:
            db_manager = DatabaseManager(MONGODB_CONFIG)
            count = db_manager.clear_session_data(collection_name, session_name)
            db_manager.close()
        
        return jsonify({
            'success': True,
            'message': f'Cleared data for {count} students',
            'count': count
        })
        
    except Exception as e:
        print(f"Error in clear_attendance_data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/')
def index():
    """Render main HTML page"""
    return render_template_string(HTML_TEMPLATE)

# ============= HTML TEMPLATE =============
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Attendance System - Fixed Version</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 15px;
        }
        .container { max-width: 1800px; margin: 0 auto; }
        .header {
            background: white;
            padding: 20px 25px;
            border-radius: 12px;
            box-shadow: 0 5px 15px rgba(0,0,0,0.2);
            margin-bottom: 15px;
        }
        .header h1 { color: #667eea; font-size: 1.8em; margin-bottom: 5px; }
        .header p { color: #666; font-size: 0.9em; }
        
        .camera-section {
            background: white;
            padding: 15px;
            border-radius: 12px;
            box-shadow: 0 5px 15px rgba(0,0,0,0.1);
            margin-bottom: 15px;
        }
        
        .config-row {
            display: grid;
            grid-template-columns: 1fr 1fr 1fr 1fr auto;
            gap: 10px;
            margin-bottom: 10px;
            align-items: end;
        }
        
        .form-group {
            display: flex;
            flex-direction: column;
            gap: 5px;
        }
        
        .form-group label {
            font-weight: 600;
            color: #555;
            font-size: 0.85em;
        }
        
        .form-group input,
        .form-group select {
            padding: 8px 10px;
            border: 2px solid #ddd;
            border-radius: 6px;
            font-size: 0.9em;
        }
        
        .btn {
            padding: 8px 16px;
            border: none;
            border-radius: 6px;
            font-size: 0.9em;
            font-weight: bold;
            cursor: pointer;
            transition: all 0.3s ease;
        }
        .btn-confirm { background: #3b82f6; color: white; }
        .btn-confirm:hover { background: #2563eb; }
        .btn:disabled { background: #9ca3af; cursor: not-allowed; opacity: 0.6; }
        
        .camera-controls {
            display: flex;
            gap: 10px;
            align-items: center;
        }
        
        .btn-start { background: #10b981; color: white; flex: 1; }
        .btn-start:hover:not(:disabled) { background: #059669; }
        .btn-stop { background: #ef4444; color: white; flex: 1; }
        .btn-stop:hover:not(:disabled) { background: #dc2626; }
        
        .status-badge {
            display: inline-flex;
            align-items: center;
            gap: 6px;
            padding: 6px 12px;
            background: #f3f4f6;
            border-radius: 6px;
            font-weight: 600;
            font-size: 0.85em;
        }
        .status-dot {
            width: 10px;
            height: 10px;
            border-radius: 50%;
            background: #ef4444;
        }
        .status-dot.active {
            background: #10b981;
            animation: pulse 2s infinite;
        }
        @keyframes pulse {
            0%, 100% { opacity: 1; }
            50% { opacity: 0.5; }
        }
        
        .video-row {
            display: grid;
            grid-template-columns: 1fr auto;
            gap: 15px;
        }
        
        .video-container {
            background: #000;
            border-radius: 8px;
            overflow: hidden;
            height: 400px;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .video-container img { 
            max-width: 100%; 
            max-height: 100%; 
            object-fit: contain;
        }
        .video-placeholder {
            color: #9ca3af;
            font-size: 1.1em;
            text-align: center;
        }
        
        .info-stats {
            display: flex;
            flex-direction: column;
            gap: 8px;
            min-width: 200px;
        }
        
        .stat-box {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 12px;
            border-radius: 8px;
            color: white;
            text-align: center;
        }
        .stat-box h3 {
            font-size: 0.7em;
            opacity: 0.9;
            margin-bottom: 4px;
        }
        .stat-box .value {
            font-size: 1.6em;
            font-weight: bold;
        }
        
        .attendance-section {
            background: white;
            padding: 20px;
            border-radius: 12px;
            box-shadow: 0 5px 15px rgba(0,0,0,0.1);
        }
        
        .section-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 15px;
        }
        
        .section-header h2 {
            color: #333;
            font-size: 1.4em;
        }
        
        .action-buttons {
            display: flex;
            gap: 8px;
        }
        
        .btn-refresh {
            background: #667eea;
            color: white;
            border: none;
            padding: 8px 16px;
            border-radius: 6px;
            cursor: pointer;
            font-weight: 600;
            font-size: 0.9em;
        }
        .btn-refresh:hover { background: #5568d3; }
        
        .btn-clear {
            background: #f59e0b;
            color: white;
            border: none;
            padding: 8px 16px;
            border-radius: 6px;
            cursor: pointer;
            font-weight: 600;
            font-size: 0.9em;
        }
        .btn-clear:hover { background: #d97706; }
        
        .stats-grid {
            display: grid;
            grid-template-columns: repeat(5, 1fr);
            gap: 12px;
            margin-bottom: 15px;
        }
        .stat-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 15px;
            border-radius: 8px;
            color: white;
            text-align: center;
        }
        .stat-card h3 {
            font-size: 0.7em;
            text-transform: uppercase;
            margin-bottom: 6px;
            opacity: 0.9;
        }
        .stat-card .value {
            font-size: 1.8em;
            font-weight: bold;
        }
        
        .table-container {
            overflow-x: auto;
            max-height: 450px;
            overflow-y: auto;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            font-size: 0.85em;
        }
        table th {
            background: #f8f9fa;
            padding: 10px 8px;
            text-align: left;
            font-weight: bold;
            color: #333;
            position: sticky;
            top: 0;
            z-index: 10;
        }
        table td {
            padding: 10px 8px;
            border-bottom: 1px solid #eee;
        }
        table tr:hover { background: #f8f9fa; }
        .status-badge-table {
            padding: 3px 10px;
            border-radius: 20px;
            font-size: 0.8em;
            font-weight: bold;
            display: inline-block;
        }
        .status-badge-table.present { background: #d1fae5; color: #065f46; }
        .status-badge-table.absent { background: #fee2e2; color: #991b1b; }
        .status-badge-table.temporary-absent { background: #fef3c7; color: #92400e; }
        .status-badge-table.permanently-absent { background: #fee2e2; color: #991b1b; }
        .status-toggle {
            padding: 4px 8px;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 0.75em;
            font-weight: bold;
        }
        .status-toggle.to-present { background: #10b981; color: white; }
        .status-toggle.to-absent { background: #ef4444; color: white; }
        .manual-badge {
            background: #fbbf24;
            color: #78350f;
            padding: 2px 6px;
            border-radius: 4px;
            font-size: 0.7em;
            margin-left: 5px;
        }
        
        .config-confirmed {
            background: #d1fae5;
            border: 2px solid #10b981;
            padding: 10px;
            border-radius: 6px;
            margin-bottom: 10px;
            display: none;
        }
        .config-confirmed.show { display: block; }
        .config-confirmed strong { color: #065f46; }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Attendance System - Fixed Version</h1>
            <p>Excel Mapping: DEPT_YEARCODE | MongoDB: lecture_metadata + DEPT_YEARCODE_DATE | Full Day Creation</p>
        </div>

        <div class="camera-section">
            <div class="config-row">
                <div class="form-group">
                    <label>Year/Batch (2022-2025)</label>
                    <input type="text" id="year" placeholder="e.g., 2022">
                </div>
                <div class="form-group">
                    <label>Department</label>
                    <input type="text" id="department" placeholder="e.g., CSBS">
                </div>
                <div class="form-group">
                    <label>Classroom</label>
                    <input type="text" id="classroom" placeholder="e.g., 301">
                </div>
                <div class="form-group">
                    <label>Teacher Name</label>
                    <input type="text" id="teacherInput" placeholder="e.g., Prof. Smith">
                </div>
                <button class="btn btn-confirm" onclick="confirmConfig()">Confirm</button>
            </div>
            
            <div class="config-confirmed" id="confirmedBanner">
                <strong>Config:</strong> <span id="confirmedText"></span> | <strong>Sheet:</strong> <span id="sheetName"></span>
            </div>
            
            <div class="camera-controls" style="margin-bottom: 10px;">
                <div class="form-group" style="flex: 0 0 150px;">
                    <label>Recognition Mode</label>
                    <select id="modeSelect">
                        <option value="1">By Name</option>
                        <option value="2">By Roll Number</option>
                    </select>
                </div>
                <button class="btn btn-start" id="startBtn" onclick="startCamera()" disabled>Start Camera</button>
                <button class="btn btn-stop" id="stopBtn" onclick="stopCamera()" disabled>Stop Camera</button>
                <div class="status-badge">
                    <div class="status-dot" id="statusDot"></div>
                    <span id="statusText">Stopped</span>
                </div>
            </div>
            
            <div class="video-row">
                <div class="video-container" id="videoContainer">
                    <div class="video-placeholder">Configure settings and confirm to begin</div>
                </div>
                
                <div class="info-stats">
                    <div class="stat-box">
                        <h3>Faces</h3>
                        <div class="value" id="currentFaces">0</div>
                    </div>
                    <div class="stat-box">
                        <h3>Session</h3>
                        <div class="value" style="font-size: 1.2em;" id="currentSession">-</div>
                    </div>
                    <div class="stat-box">
                        <h3>Total</h3>
                        <div class="value" id="quickTotal">-</div>
                    </div>
                    <div class="stat-box">
                        <h3>Present</h3>
                        <div class="value" id="quickPresent">-</div>
                    </div>
                    <div class="stat-box">
                        <h3>Absent</h3>
                        <div class="value" id="quickAbsent">-</div>
                    </div>
                </div>
            </div>
        </div>

        <div class="attendance-section">
            <div class="section-header">
                <h2>Current Session Attendance</h2>
                <div class="action-buttons">
                    <button class="btn-refresh" onclick="refreshData()">Refresh</button>
                    <button class="btn-clear" onclick="clearSessionData()" id="clearBtn" disabled>Clear</button>
                </div>
            </div>

            <div class="stats-grid">
                <div class="stat-card">
                    <h3>Total</h3>
                    <div class="value" id="totalStudents">-</div>
                </div>
                <div class="stat-card">
                    <h3>Present</h3>
                    <div class="value" id="presentCount">-</div>
                </div>
                <div class="stat-card">
                    <h3>Absent</h3>
                    <div class="value" id="absentCount">-</div>
                </div>
                <div class="stat-card">
                    <h3>Temp Absent</h3>
                    <div class="value" id="tempAbsentCount">-</div>
                </div>
                <div class="stat-card">
                    <h3>Attendance %</h3>
                    <div class="value" id="attendancePercentage">-</div>
                </div>
            </div>

            <div class="table-container">
                <table>
                    <thead>
                        <tr>
                            <th>Roll No</th>
                            <th>Name</th>
                            <th>Status</th>
                            <th>First Seen</th>
                            <th>Last Seen</th>
                            <th>Present Time</th>
                            <th>Absent Time</th>
                            <th>Action</th>
                        </tr>
                    </thead>
                    <tbody id="attendanceBody">
                        <tr><td colspan="8" style="text-align: center;">Configure and confirm settings to load students</td></tr>
                    </tbody>
                </table>
            </div>
        </div>
    </div>

    <script>
        const API_BASE_URL = window.location.origin + '/api';
        const YEAR_MAP = { '2022': 'B.Tech', '2023': 'TY', '2024': 'SY', '2025': 'FY' };
        let cameraRunning = false;
        let refreshInterval = null;
        let currentCollection = null;
        let currentSession = null;
        let configConfirmed = false;
        let configData = {};

        function confirmConfig() {
            const year = document.getElementById('year').value.trim();
            const department = document.getElementById('department').value.trim();
            const classroom = document.getElementById('classroom').value.trim();
            const teacher = document.getElementById('teacherInput').value.trim();
            
            if (!year || !department || !classroom || !teacher) {
                alert('Please fill in all fields');
                return;
            }
            
            const yearCode = YEAR_MAP[year] || 'B.Tech';
            const sheetName = `${department}_${yearCode}`;
            
            configData = { year, department, classroom, teacher_name: teacher };
            configConfirmed = true;
            
            document.getElementById('confirmedText').textContent = 
                `${year} | ${department} | ${classroom} | ${teacher}`;
            document.getElementById('sheetName').textContent = sheetName;
            document.getElementById('confirmedBanner').classList.add('show');
            document.getElementById('startBtn').disabled = false;
            
            loadCurrentSession();
        }

        async function startCamera() {
            if (!configConfirmed) {
                alert('Please confirm configuration first');
                return;
            }
            
            const mode = document.getElementById('modeSelect').value;
            const startBtn = document.getElementById('startBtn');
            const stopBtn = document.getElementById('stopBtn');
            
            startBtn.disabled = true;
            startBtn.textContent = 'Starting...';
            try {
                const response = await fetch(`${API_BASE_URL}/camera/start`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ 
                        mode: parseInt(mode),
                        ...configData
                    })
                });
                
                const data = await response.json();
                if (data.success) {
                    cameraRunning = true;
                    updateCameraStatus(true);
                    startBtn.disabled = true;
                    stopBtn.disabled = false;
                    startBtn.textContent = 'Start Camera';
                    document.getElementById('videoContainer').innerHTML = 
                        '<img src="' + API_BASE_URL + '/video_feed?t=' + Date.now() + '">';
                    startAutoRefresh();
                    await loadCurrentSession();
                } else {
                    throw new Error(data.message || data.error);
                }
            } catch (error) {
                alert('Error: ' + error.message);
                startBtn.disabled = false;
                startBtn.textContent = 'Start Camera';
            }
        }

        async function stopCamera() {
            const startBtn = document.getElementById('startBtn');
            const stopBtn = document.getElementById('stopBtn');
            stopBtn.disabled = true;
            stopBtn.textContent = 'Stopping...';
            try {
                const response = await fetch(`${API_BASE_URL}/camera/stop`, { method: 'POST' });
                const data = await response.json();
                if (data.success) {
                    cameraRunning = false;
                    updateCameraStatus(false);
                    startBtn.disabled = false;
                    stopBtn.disabled = true;
                    stopBtn.textContent = 'Stop Camera';
                    document.getElementById('videoContainer').innerHTML = 
                        '<div class="video-placeholder">Camera stopped</div>';
                    stopAutoRefresh();
                }
            } catch (error) {
                alert('Error: ' + error.message);
                stopBtn.disabled = false;
                stopBtn.textContent = 'Stop Camera';
            }
        }

        function updateCameraStatus(isRunning) {
            const statusDot = document.getElementById('statusDot');
            const statusText = document.getElementById('statusText');
            if (isRunning) {
                statusDot.classList.add('active');
                statusText.textContent = 'Running';
            } else {
                statusDot.classList.remove('active');
                statusText.textContent = 'Stopped';
            }
        }

        async function updateCameraInfo() {
            try {
                const response = await fetch(`${API_BASE_URL}/camera/status`);
                const data = await response.json();
                document.getElementById('currentFaces').textContent = data.current_faces || 0;
                document.getElementById('currentSession').textContent = data.current_session || '-';
                
                if (data.current_session && data.current_session !== currentSession) {
                    currentSession = data.current_session;
                    await loadCurrentSession();
                }
            } catch (error) {
                console.error('Error updating camera info:', error);
            }
        }

        async function loadCurrentSession() {
            if (!configConfirmed) return;
            
            try {
                const response = await fetch(`${API_BASE_URL}/current-session`);
                const data = await response.json();
                if (data.success && data.active) {
                    currentCollection = data.collection_name;
                    currentSession = data.session_name;
                    updateStats(data.summary);
                    displayAttendanceData(data.attendance);
                    document.getElementById('clearBtn').disabled = false;
                }
            } catch (error) {
                console.error('Error loading session:', error);
            }
        }

        function updateStats(summary) {
            document.getElementById('totalStudents').textContent = summary.total || 0;
            document.getElementById('presentCount').textContent = summary.present || 0;
            document.getElementById('absentCount').textContent = summary.absent || 0;
            document.getElementById('tempAbsentCount').textContent = summary.temporary_absent || 0;
            document.getElementById('attendancePercentage').textContent = 
                (summary.attendance_percentage || 0).toFixed(2) + '%';
            
            document.getElementById('quickTotal').textContent = summary.total || 0;
            document.getElementById('quickPresent').textContent = summary.present || 0;
            document.getElementById('quickAbsent').textContent = summary.absent || 0;
        }

        function displayAttendanceData(records) {
            const tbody = document.getElementById('attendanceBody');
            if (!records || records.length === 0) {
                tbody.innerHTML = '<tr><td colspan="8" style="text-align:center;">No records found</td></tr>';
                return;
            }

            tbody.innerHTML = '';
            records.forEach(rec => {
                const row = document.createElement('tr');
                const status = rec.status || 'Absent';
                let badgeClass = 'absent';
                if (status === 'Present') badgeClass = 'present';
                else if (status === 'Temporary Absent') badgeClass = 'temporary-absent';
                else if (status === 'Permanently Absent') badgeClass = 'permanently-absent';

                const timestamps = rec.timestamps || {};
                const durations = rec.durations || {};
                const flags = rec.flags || {};

                row.innerHTML = `
                    <td>${rec.roll_no || '-'}</td>
                    <td>${rec.name || '-'}</td>
                    <td>
                        <span class="status-badge-table ${badgeClass}">${status}</span>
                        ${flags.manual_override ? '<span class="manual-badge">Manual</span>' : ''}
                    </td>
                    <td>${timestamps.first_seen || 'N/A'}</td>
                    <td>${timestamps.last_seen || 'N/A'}</td>
                    <td>${durations.total_present_human || '0 sec'}</td>
                    <td>${durations.total_absent_human || '0 sec'}</td>
                    <td>
                        <button class="status-toggle ${status === 'Present' ? 'to-absent' : 'to-present'}"
                            onclick="toggleAttendance('${rec.roll_no}', '${status}')">
                            ${status === 'Present' ? 'Mark Absent' : 'Mark Present'}
                        </button>
                    </td>`;
                tbody.appendChild(row);
            });
        }

        async function toggleAttendance(rollNo, currentStatus) {
            if (!currentCollection || !currentSession) {
                alert('No active session');
                return;
            }
            
            const newStatus = currentStatus === 'Present' ? 'Absent' : 'Present';
            const button = event.target;
            button.disabled = true;
            button.textContent = 'Updating...';
            
            try {
                const response = await fetch(`${API_BASE_URL}/attendance/update`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({
                        collection_name: currentCollection,
                        session_name: currentSession,
                        roll_no: rollNo,
                        status: newStatus
                    })
                });
                
                const data = await response.json();
                if (data.success) {
                    await refreshData();
                } else {
                    alert('Failed to update: ' + (data.error || 'Unknown error'));
                }
            } catch (error) {
                console.error('Error toggling attendance:', error);
                alert('Error: ' + error.message);
            } finally {
                button.disabled = false;
                button.textContent = currentStatus === 'Present' ? 'Mark Absent' : 'Mark Present';
            }
        }

        async function refreshData() {
            await loadCurrentSession();
        }

        async function clearSessionData() {
            if (!currentCollection || !currentSession) {
                alert('No active session to clear');
                return;
            }
            if (!confirm('Are you sure you want to clear all attendance data for this session?')) return;
            
            try {
                const response = await fetch(`${API_BASE_URL}/attendance/clear`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ 
                        collection_name: currentCollection,
                        session_name: currentSession 
                    })
                });
                const data = await response.json();
                if (data.success) {
                    alert(data.message);
                    await refreshData();
                } else {
                    alert('Failed to clear data: ' + (data.error || 'Unknown error'));
                }
            } catch (error) {
                console.error('Error clearing session data:', error);
                alert('Error: ' + error.message);
            }
        }

        function startAutoRefresh() {
            if (refreshInterval) clearInterval(refreshInterval);
            updateCameraInfo();
            refreshInterval = setInterval(() => {
                updateCameraInfo();
                refreshData();
            }, 5000);
        }

        function stopAutoRefresh() {
            if (refreshInterval) clearInterval(refreshInterval);
        }

        window.onload = function() {
            updateCameraStatus(false);
        };
    </script>
</body>
</html>
'''

if __name__ == '__main__':
    print("=" * 80)
    print("ATTENDANCE SYSTEM - FIXED & COMPLETE VERSION")
    print("=" * 80)
    print("\n✅ IMPLEMENTED FEATURES:")
    print("   1. MongoDB schema with proper nested structure")
    print("   2. Excel sheet configuration with YEAR_MAPPING")
    print("   3. Sheet naming: DEPT_YEARCODE (e.g., CSBS_B.Tech)")
    print("   4. Collection naming: DEPT_YEARCODE_DATE")
    print("   5. Full day creation with 8 sessions at once")
    print("   6. Robust duration tracking system")
    print("   7. Complete HTML template preserved")
    print("   8. All JavaScript functions completed")
    print("\n📊 MONGODB STRUCTURE:")
    print("   Database: Attendance_system")
    print("   ├── Collection: lecture_metadata")
    print("   └── Collection: DEPT_YEARCODE_DATE")
    print("       └── Document: {")
    print("           date, department, year, year_code,")
    print("           classroom, teacher_name, camera_ids,")
    print("           sessions: {")
    print("               'Session 1': { start_time, end_time,")
    print("                   students: { 'ROLL_NO': {...} } }")
    print("           }")
    print("       }")
    print("\n📋 YEAR MAPPING:")
    print("   2022 → B.Tech")
    print("   2023 → TY")
    print("   2024 → SY")
    print("   2025 → FY")
    print("\n📁 EXCEL STRUCTURE:")
    print("   File: Book2.xlsx")
    print("   Sheet Format: DEPT_YEARCODE")
    print("   Example Sheets: CSBS_B.Tech, IT_TY, ENTC_SY, AIDS_FY")
    print("\n🔄 WORKFLOW:")
    print("   1. User configures: year, department, classroom, teacher")
    print("   2. System maps year to code using YEAR_MAPPING")
    print("   3. Loads Excel sheet: DEPT_YEARCODE")
    print("   4. Creates/opens collection: DEPT_YEARCODE_DATE")
    print("   5. Initializes all 8 sessions with student data")
    print("   6. Updates current session as camera runs")
    print("   7. Tracks durations with proper timer finalization")
    print("\n⏱️ DURATION TRACKING:")
    print("   - Present timer starts when status becomes Present")
    print("   - Absent timer starts when status leaves Present")
    print("   - Finalizes previous period before status change")
    print("   - Accumulates total durations across changes")
    print("   - Works for both manual and automatic updates")
    print("\n🎯 KEY IMPROVEMENTS:")
    print("   ✓ Proper Excel header handling (multiple variations)")
    print("   ✓ Robust error handling and logging")
    print("   ✓ Complete JavaScript implementations")
    print("   ✓ No truncated or incomplete code")
    print("   ✓ Preserved original HTML template structure")
    print("   ✓ Thread-safe database operations")
    print("   ✓ Proper MongoDB connection management")
    print("\n🚀 Starting Server...")
    print(f"   MongoDB: {MONGODB_CONFIG['host']}:{MONGODB_CONFIG['port']}")
    print(f"   Database: {MONGODB_CONFIG['database']}")
    print(f"   Server URL: http://localhost:5000")
    print("\n" + "=" * 80 + "\n")
    
    try:
        app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)
    except KeyboardInterrupt:
        print("\n\n👋 Server stopped by user")
    except Exception as e:
        print(f"\n\n❌ Server error: {e}")
        import traceback
        traceback.print_exc()