from openpyxl.styles import Font, PatternFill, Border, Alignment
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook


# Load the workbook
file_path = 'Book1.xlsx'  # Path to your attendance workbook
workbook = load_workbook(file_path)

def copy_cell_format(source_cell, target_cell):
    target_cell.font = Font(name=source_cell.font.name,
                            size=source_cell.font.size,
                            bold=source_cell.font.bold,
                            italic=source_cell.font.italic,
                            vertAlign=source_cell.font.vertAlign,
                            underline=source_cell.font.underline,
                            strike=source_cell.font.strike,
                            color=source_cell.font.color)
    
    target_cell.fill = PatternFill(fill_type=source_cell.fill.fill_type,
                                   fgColor=source_cell.fill.fgColor,
                                   bgColor=source_cell.fill.bgColor)
    
    target_cell.border = Border(left=source_cell.border.left,
                                right=source_cell.border.right,
                                top=source_cell.border.top,
                                bottom=source_cell.border.bottom)
    
    target_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                      vertical=source_cell.alignment.vertical,
                                      text_rotation=source_cell.alignment.text_rotation,
                                      wrap_text=source_cell.alignment.wrap_text,
                                      shrink_to_fit=source_cell.alignment.shrink_to_fit,
                                      indent=source_cell.alignment.indent)

def copy_merged_cells(source_sheet, target_sheet):
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))
        first_cell = merged_range.start_cell
        target_cell = target_sheet.cell(row=first_cell.row, column=first_cell.column)
        copy_cell_format(source_sheet[first_cell.coordinate], target_cell)


def get_or_create_today_sheet(workbook, source_sheet_name='Sheet 1'):
    today_date = datetime.now().strftime('%Y-%m-%d')
    if today_date in workbook.sheetnames:
        target_sheet = workbook[today_date]
        print(f"Using existing sheet for today's date: {today_date}")
    else:
        source_sheet = workbook[source_sheet_name]
        target_sheet = workbook.create_sheet(title=today_date)
        print(f"Creating new sheet for today's date: {today_date}")

        # Copy headers and formats from source sheet
        for column in source_sheet.columns:
            col_letter = get_column_letter(column[0].column)
            target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width
            for source_cell in column:
                target_cell = target_sheet.cell(row=source_cell.row, column=source_cell.column)
                target_cell.value = source_cell.value
                copy_cell_format(source_cell, target_cell)

        copy_merged_cells(source_sheet, target_sheet)
    
    return target_sheet



# Get or create today's sheet
target_sheet = get_or_create_today_sheet(workbook)

def get_current_session():
    now = datetime.now().time()
    sessions = {
        'Session 1': (datetime.strptime('09:00:00', '%H:%M:%S').time(), datetime.strptime('10:00:00', '%H:%M:%S').time()),
        'Session 2': (datetime.strptime('10:00:01', '%H:%M:%S').time(), datetime.strptime('11:00:00', '%H:%M:%S').time()),
        'Session 3': (datetime.strptime('11:00:01', '%H:%M:%S').time(), datetime.strptime('11:45:00', '%H:%M:%S').time()),
        'Session 4': (datetime.strptime('11:45:01', '%H:%M:%S').time(), datetime.strptime('12:40:00', '%H:%M:%S').time()),
        'Session 5': (datetime.strptime('12:40:01', '%H:%M:%S').time(), datetime.strptime('12:50:00', '%H:%M:%S').time()),
        'Session 6': (datetime.strptime('12:50:01', '%H:%M:%S').time(), datetime.strptime('14:45:00', '%H:%M:%S').time()),
        'Session 7': (datetime.strptime('14:45:01', '%H:%M:%S').time(), datetime.strptime('23:45:00', '%H:%M:%S').time()),
    }
    for session, (start, end) in sessions.items():
        if start <= now <= end:
            return session
    return None

def search_keyword_ranges(sheet, keyword):
    keyword = keyword.upper()
    keyword_rows = []
    
    for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for cell_value in row:
            if isinstance(cell_value, str) and keyword in cell_value.upper():
                keyword_rows.append(row_num)
    return keyword_rows
# def search_keyword_ranges(sheet, keyword):
#     if not keyword:
#         print("Keyword is None or empty. Please provide a valid keyword.")
#         return []

#     keyword = keyword.upper()
#     keyword_rows = []

#     for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
#         for cell_value in row:
#             if isinstance(cell_value, str) and keyword in cell_value.upper():
#                 keyword_rows.append(row_num)
#     return keyword_rows


def find_end_limit(sheet, keyword_start_row, keyword_column=1):
    """
    Finds the end row of the current session, stopping when it finds the next keyword or an empty row.
    
    Parameters:
    - sheet: The Excel sheet to search.
    - keyword_start_row: The starting row where the keyword (session) was found.
    - keyword_column: The column in which the session keywords are located. Default is 1 (column A).
    
    Returns:
    - The row number where the session ends.
    """
    for row_num in range(keyword_start_row + 1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_num, column=keyword_column).value
        # Stop if the next keyword is found or if the row is empty
        if cell_value is None or (isinstance(cell_value, str) and "session" in cell_value.upper()):
            return row_num - 1  # Return the row before the next keyword or empty row
    
    # If no keyword or empty row is found, return the last row of the sheet
    return sheet.max_row

def get_top_left_cell(sheet, cell):
    """Get the top-left cell of the merged cell range for the given cell."""
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(merged_range.min_row, merged_range.min_col)
    return cell

# def update_student_info(sheet, student_name, last_seen_time, current_time, start_row, end_row,
#                         status_column_offset, last_seen_column_offset, absence_timer_start_column_offset,
#                         status=None, update_absence_timer=True, time_in_seconds=None):
#     """
#     Updates the student's information in the Excel sheet.
#     """
#     print(f"Updating student info for {student_name}...")

#     # Locate the row for the student
#     for row in range(start_row, end_row + 1):
#         if sheet.cell(row=row, column=4).value == student_name:  # Assuming names are in column 3
#             # Update status
#             status_cell = sheet.cell(row=row, column=4 + status_column_offset)
#             if status:
#                 status_cell.value = status
#                 if status == "Temporary Absent":
#                     status_cell.font = Font(color="FFA500")  # Orange
#                 elif status == "Permanently Absent":
#                     status_cell.font = Font(color="FF0000")  # Red
#                 elif status == "Present":
#                     status_cell.font = Font(color="00FF00")  # Green

#             # Update last seen time
#             last_seen_cell = sheet.cell(row=row, column=3 + last_seen_column_offset)
#             last_seen_cell.value = last_seen_time.strftime("%H:%M:%S") if last_seen_time else "N/A"

#             # Update absence timer start time
#             absence_timer_start_cell = sheet.cell(row=row, column=3 + absence_timer_start_column_offset)
#             absence_timer_start_cell.value = current_time.strftime("%H:%M:%S") if update_absence_timer else absence_timer_start_cell.value

#             # Update absence timer duration if provided
#             if time_in_seconds is not None:
#                 absence_timer_start_cell.value = f"{time_in_seconds} seconds"

#             break
from openpyxl.styles import Font

def update_student_info(sheet, student_name, last_seen_time, current_time, start_row, end_row,
                        status_column_offset, last_seen_column_offset, absence_timer_start_column_offset,
                        status=None, update_absence_timer=True, time_in_seconds=None):
    """
    Updates the student's information in the Excel sheet, including font color changes for the status, 
    last seen time, and absence timer.
    """
    print(f"Updating student info for {student_name}...")

    # Locate the row for the student
    student_found = False
    for row in range(start_row, end_row + 1):
        student_cell_value = sheet.cell(row=row, column=4).value
        print(f"Checking row {row}, student: {student_cell_value}")  # Debugging line

        # Strip any leading/trailing spaces and compare
        if student_cell_value and student_cell_value.strip() == student_name.strip():
            student_found = True
            # Update status
            status_cell = sheet.cell(row=row, column=4 + status_column_offset)
            if status:
                # status_cell.value = status
                print(status_cell.value)
                if status == "Temporary Absent":
                    status_cell.font = Font(color="FFA500")  # Orange
                    status_cell.value = status
                elif status == "Permanently Absent":
                    status_cell.font = Font(color="FF0000")  # Red
                    status_cell.value = status
                elif status == "Present":
                    status_cell.font = Font(color="00FF00")  # Green
                    status_cell.value = status
                else:
                    status_cell.font = Font(color="000000")  # Default color for other statuses

            # Update last seen time
            last_seen_cell = sheet.cell(row=row, column=3 + last_seen_column_offset)
            last_seen_cell.value = last_seen_time.strftime("%H:%M:%S") if last_seen_time else "N/A"
            if status == "Present":
                last_seen_cell.font = Font(color="00FF00")  # Green
            elif status == "Temporary Absent":
                last_seen_cell.font = Font(color="FFA500")  # Orange
            elif status == "Permanently Absent":
                last_seen_cell.font = Font(color="FF0000")  # Red
            else:
                last_seen_cell.font = Font(color="000000")  # Default color

            # Update absence timer start time
            absence_timer_start_cell = sheet.cell(row=row, column=3 + absence_timer_start_column_offset)
            if update_absence_timer:
                absence_timer_start_cell.value = current_time.strftime("%H:%M:%S")
            else:
                # Only update the timer if it's not being reset
                absence_timer_start_cell.value = absence_timer_start_cell.value

            # Update absence timer duration if provided
            if time_in_seconds is not None:
                timer_duration_cell = sheet.cell(row=row, column=3 + absence_timer_start_column_offset + 1)  # Assuming next column
                timer_duration_cell.value = f"{time_in_seconds} seconds"
                # Change font color based on status
                if status == "Present":
                    timer_duration_cell.font = Font(color="00FF00")  # Green
                elif status == "Temporary Absent":
                    timer_duration_cell.font = Font(color="FFA500")  # Orange
                elif status == "Permanently Absent":
                    timer_duration_cell.font = Font(color="FF0000")  # Red
                else:
                    timer_duration_cell.font = Font(color="000000")  # Default color

            break

    if not student_found:
        print(f"Student {student_name} not found in the specified range.")

    # Save the workbook
    try:
        sheet.parent.save("Attendance_updated.xlsx")  # Replace with the actual file name
        print("Workbook saved successfully.")
    except Exception as e:
        print(f"Error saving workbook: {e}")
