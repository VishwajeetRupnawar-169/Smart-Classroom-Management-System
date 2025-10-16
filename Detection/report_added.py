"""
Enhanced Attendance System - COMPLETE PRODUCTION READY VERSION
All features implemented with no truncation
"""

from flask import Flask, jsonify, request, render_template_string, Response, send_file
from flask_cors import CORS
from pymongo import MongoClient
from datetime import datetime, timedelta
import os
import threading
import cv2
import numpy as np
import face_recognition
import sys
from time import sleep
from threading import Lock, Event
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
from bson.objectid import ObjectId
import traceback

sys.path.append(os.path.abspath('../'))
try:
    from Excel_Format import get_current_session
except ImportError:
    def get_current_session():
        current_hour = datetime.now().hour
        if 9 <= current_hour < 10:
            return "Session 1"
        elif 10 <= current_hour < 11:
            return "Session 2"
        elif 11 <= current_hour < 12:
            return "Session 3"
        elif 12 <= current_hour < 13:
            return "Session 4"
        elif 13 <= current_hour < 14:
            return "Session 5"
        elif 14 <= current_hour < 15:
            return "Session 6"
        elif 15 <= current_hour < 16:
            return "Session 7"
        else:
            return "Session 8"

app = Flask(__name__)
CORS(app)

MONGODB_CONFIG = {
    'host': 'localhost',
    'port': 27017,
    'database': 'Attendance_system'
}

TEMPLATE_FILE = 'Book2.xlsx'
ALL_SESSIONS = ["Session 1", "Session 2", "Session 3", "Session 4", 
                "Session 5", "Session 6", "Session 7", "Session 8"]

YEAR_MAPPING = {
    '2022': 'B.Tech',
    '2023': 'TY',
    '2024': 'SY',
    '2025': 'FY'
}

attendance_system = None
camera_running = False
camera_lock = Lock()


class FilePathResolver:
    """Robust file path resolution"""
    
    @staticmethod
    def find_file(filename, search_dirs=None):
        if search_dirs is None:
            search_dirs = [
                '.',
                '..',
                '../..',
                os.path.dirname(__file__),
                os.path.join(os.path.dirname(__file__), '..'),
                os.path.join(os.path.dirname(__file__), '../..'),
            ]
        
        for base_dir in search_dirs:
            file_path = os.path.join(base_dir, filename)
            if os.path.exists(file_path):
                abs_path = os.path.abspath(file_path)
                print(f"✓ Found file: {abs_path}")
                return abs_path
        
        raise FileNotFoundError(
            f"File '{filename}' not found in any location"
        )
    
    @staticmethod
    def find_training_folder(dept_year_code, mode_name):
        base_dirs = [
            'Training_images',
            '../Training_images',
            '../../Training_images',
            os.path.join(os.path.dirname(__file__), 'Training_images'),
            os.path.join(os.path.dirname(__file__), '../Training_images'),
        ]
        
        for base in base_dirs:
            structured_path = os.path.join(base, dept_year_code, mode_name)
            if os.path.exists(structured_path):
                abs_path = os.path.abspath(structured_path)
                print(f"✓ Found training folder: {abs_path}")
                return abs_path
        
        for base in base_dirs:
            flat_path = os.path.join(base, mode_name)
            if os.path.exists(flat_path):
                abs_path = os.path.abspath(flat_path)
                print(f"✓ Found training folder (flat): {abs_path}")
                return abs_path
        
        raise FileNotFoundError(
            f"Training images folder not found for {dept_year_code}/{mode_name}"
        )


class AttendanceConfig:
    TEMPLATE_FILE = 'Book2.xlsx'
    ABSENCE_DETECTION_DELAY = 5
    TEMPORARY_ABSENT_THRESHOLD = 10
    PERMANENT_ABSENT_THRESHOLD = 15
    ABSENCE_CHECK_INTERVAL = 2
    MODE_NAME = 1
    MODE_ROLL_NO = 2


class DatabaseManager:
    def __init__(self, mongodb_config):
        self.mongodb_config = mongodb_config
        self.client = None
        self.db = None
        self.lock = Lock()
        self._initialize_db()
    
    def _get_connection(self):
        try:
            if self.client is None:
                self.client = MongoClient(
                    self.mongodb_config['host'], 
                    self.mongodb_config['port'],
                    serverSelectionTimeoutMS=5000
                )
                self.db = self.client[self.mongodb_config['database']]
            return self.db
        except Exception as e:
            print(f"Error connecting to MongoDB: {e}")
            raise
    
    def _initialize_db(self):
        try:
            self.db = self._get_connection()
            try:
                self.db.lecture_metadata.create_index([('collection_name', 1)], unique=True)
            except:
                pass
            print("✓ MongoDB initialized")
        except Exception as e:
            print(f"Error initializing database: {e}")
            raise
    
    def _get_year_code(self, year_input):
        return YEAR_MAPPING.get(str(year_input), 'B.Tech')
    
    def _get_sheet_name(self, department, year_input):
        year_code = self._get_year_code(year_input)
        return f"{department}_{year_code}"
    
    def load_students_from_excel(self, excel_file, sheet_name):
        try:
            excel_path = FilePathResolver.find_file(excel_file)
            wb = load_workbook(excel_path, data_only=True)
            
            if sheet_name not in wb.sheetnames:
                available = ", ".join(wb.sheetnames)
                wb.close()
                raise ValueError(f"Sheet '{sheet_name}' not found. Available: {available}")
            
            sheet = wb[sheet_name]
            data = []
            headers = None
            
            for row in sheet.iter_rows(values_only=True):
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                
                row_data = [cell if cell is not None else '' for cell in row]
                
                if headers is None:
                    headers = [str(h).strip() for h in row_data]
                else:
                    data.append(row_data)
            
            wb.close()
            print(f"✓ Loaded {len(data)} students from '{sheet_name}'")
            return headers, data
        except Exception as e:
            print(f"Error loading Excel: {e}")
            raise
    
    def create_or_get_daily_collection(self, department, year_input, date_str, 
                                      classroom, teacher_name, template_headers, 
                                      template_data, camera_ids=None):
        year_code = self._get_year_code(year_input)
        collection_name = f"{department}_{year_code}_{date_str}"
        
        with self.lock:
            try:
                db = self._get_connection()
                
                if collection_name in db.list_collection_names():
                    print(f"✓ Using existing: {collection_name}")
                    return collection_name
                
                collection = db[collection_name]
                current_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                sessions = {}
                for session_name in ALL_SESSIONS:
                    students_dict = {}
                    
                    for row in template_data:
                        doc_data = {}
                        for i, header in enumerate(template_headers):
                            if i < len(row):
                                value = row[i]
                                doc_data[header] = str(value).strip() if value not in (None, '') else ''
                            else:
                                doc_data[header] = ''
                        
                        roll_no = (doc_data.get('Roll No') or 
                                  doc_data.get('Roll NO') or 
                                  doc_data.get('Roll_No') or '')
                        
                        if not roll_no:
                            continue
                        
                        students_dict[roll_no] = {
                            'sr_no': doc_data.get('Sr. No', ''),
                            'roll_no': roll_no,
                            'prn_no': doc_data.get('PRN No.', ''),
                            'name': doc_data.get('Name', ''),
                            'status': 'Absent',
                            'timestamps': {
                                'first_seen': None,
                                'last_seen': None,
                                'present_timer_start': None,
                                'absence_timer_start': None,
                                'temp_absent_time': None,
                                'perm_absent_time': None,
                                'last_updated': current_time_str
                            },
                            'durations': {
                                'total_present_seconds': 0,
                                'total_absent_seconds': 0,
                                'total_present_human': '0 sec',
                                'total_absent_human': '0 sec'
                            },
                            'flags': {
                                'manual_override': False,
                                'is_temp_absent': False,
                                'is_perm_absent': False
                            }
                        }
                    
                    sessions[session_name] = {
                        'start_time': None,
                        'end_time': None,
                        'students': students_dict
                    }
                
                daily_doc = {
                    'date': date_str,
                    'department': department,
                    'year': year_input,
                    'year_code': year_code,
                    'classroom': classroom,
                    'teacher_name': teacher_name,
                    'camera_ids': camera_ids or [],
                    'created_at': datetime.now(),
                    'sessions': sessions
                }
                
                collection.insert_one(daily_doc)
                
                metadata = {
                    'collection_name': collection_name,
                    'date': date_str,
                    'department': department,
                    'year': year_input,
                    'year_code': year_code,
                    'classroom': classroom,
                    'teacher_name': teacher_name,
                    'created_at': datetime.now()
                }
                db.lecture_metadata.insert_one(metadata)
                
                print(f"✓ Created: {collection_name}")
                return collection_name
            except Exception as e:
                print(f"Error creating collection: {e}")
                traceback.print_exc()
                raise
    
    def update_student_attendance(self, collection_name, session_name, roll_no, status, manual=False):
        with self.lock:
            try:
                db = self._get_connection()
                collection = db[collection_name]
                doc = collection.find_one({})
                
                if not doc or session_name not in doc.get('sessions', {}):
                    return False
                
                student_path = f'sessions.{session_name}.students.{roll_no}'
                student = doc['sessions'][session_name]['students'].get(roll_no)
                
                if not student:
                    return False
                
                current_time = datetime.now()
                current_time_str = current_time.strftime('%Y-%m-%d %H:%M:%S')
                
                prev_status = student.get('status', 'Absent')
                present_start = student['timestamps'].get('present_timer_start')
                absent_start = student['timestamps'].get('absence_timer_start')
                total_present = student['durations'].get('total_present_seconds', 0)
                total_absent = student['durations'].get('total_absent_seconds', 0)
                
                update_fields = {}
                
                if prev_status == 'Present' and status != 'Present':
                    if present_start:
                        try:
                            start_dt = datetime.strptime(present_start, '%Y-%m-%d %H:%M:%S')
                            duration = (current_time - start_dt).total_seconds()
                            total_present += duration
                        except:
                            pass
                    update_fields[f'{student_path}.timestamps.present_timer_start'] = None
                    update_fields[f'{student_path}.timestamps.absence_timer_start'] = current_time_str
                elif prev_status != 'Present' and status == 'Present':
                    if absent_start:
                        try:
                            start_dt = datetime.strptime(absent_start, '%Y-%m-%d %H:%M:%S')
                            duration = (current_time - start_dt).total_seconds()
                            total_absent += duration
                        except:
                            pass
                    update_fields[f'{student_path}.timestamps.absence_timer_start'] = None
                    update_fields[f'{student_path}.timestamps.present_timer_start'] = current_time_str
                elif status == 'Present' and not present_start:
                    update_fields[f'{student_path}.timestamps.present_timer_start'] = current_time_str
                elif status != 'Present' and not absent_start:
                    update_fields[f'{student_path}.timestamps.absence_timer_start'] = current_time_str
                
                update_fields[f'{student_path}.status'] = status
                update_fields[f'{student_path}.timestamps.last_updated'] = current_time_str
                update_fields[f'{student_path}.flags.manual_override'] = manual
                
                if student['timestamps']['first_seen'] is None and status == 'Present':
                    update_fields[f'{student_path}.timestamps.first_seen'] = current_time_str
                
                if status == 'Present':
                    update_fields[f'{student_path}.timestamps.last_seen'] = current_time.strftime('%H:%M:%S')
                
                update_fields[f'{student_path}.durations.total_present_seconds'] = int(total_present)
                update_fields[f'{student_path}.durations.total_absent_seconds'] = int(total_absent)
                update_fields[f'{student_path}.durations.total_present_human'] = self._format_duration(total_present)
                update_fields[f'{student_path}.durations.total_absent_human'] = self._format_duration(total_absent)
                
                if status == 'Temporary Absent':
                    update_fields[f'{student_path}.timestamps.temp_absent_time'] = current_time_str
                    update_fields[f'{student_path}.flags.is_temp_absent'] = True
                elif status == 'Permanently Absent':
                    update_fields[f'{student_path}.timestamps.perm_absent_time'] = current_time_str
                    update_fields[f'{student_path}.flags.is_perm_absent'] = True
                elif status == 'Present':
                    update_fields[f'{student_path}.flags.is_temp_absent'] = False
                    update_fields[f'{student_path}.flags.is_perm_absent'] = False
                
                if doc['sessions'][session_name]['start_time'] is None:
                    update_fields[f'sessions.{session_name}.start_time'] = current_time_str
                
                result = collection.update_one({}, {'$set': update_fields})
                return result.modified_count > 0
            except Exception as e:
                print(f"Error updating: {e}")
                return False
    
    def _format_duration(self, seconds):
        try:
            seconds = int(seconds)
        except:
            seconds = 0
        
        if seconds < 60:
            return f"{seconds} sec"
        elif seconds < 3600:
            minutes = seconds // 60
            secs = seconds % 60
            return f"{minutes} min {secs} sec"
        else:
            hours = seconds // 3600
            minutes = (seconds % 3600) // 60
            return f"{hours} hr {minutes} min"
    
    def get_session_attendance(self, collection_name, session_name):
        with self.lock:
            try:
                db = self._get_connection()
                collection = db[collection_name]
                doc = collection.find_one({})
                
                if not doc or session_name not in doc.get('sessions', {}):
                    return []
                
                students = doc['sessions'][session_name]['students']
                return list(students.values())
            except Exception as e:
                return []
    
    def get_session_summary(self, collection_name, session_name):
        with self.lock:
            try:
                db = self._get_connection()
                collection = db[collection_name]
                doc = collection.find_one({})
                
                if not doc or session_name not in doc.get('sessions', {}):
                    return {}
                
                students = doc['sessions'][session_name]['students']
                total = len(students)
                present = sum(1 for s in students.values() if s['status'] == 'Present')
                temp_absent = sum(1 for s in students.values() if s['status'] == 'Temporary Absent')
                perm_absent = sum(1 for s in students.values() if s['status'] == 'Permanently Absent')
                absent = sum(1 for s in students.values() if s['status'] == 'Absent')
                
                return {
                    'total': total,
                    'present': present,
                    'temporary_absent': temp_absent,
                    'permanently_absent': perm_absent,
                    'absent': absent,
                    'attendance_percentage': round((present / total * 100), 2) if total > 0 else 0
                }
            except Exception as e:
                return {}
    
    def get_all_daily_collections(self):
        with self.lock:
            try:
                db = self._get_connection()
                collections = list(db.lecture_metadata.find().sort('created_at', -1))
                for col in collections:
                    col['_id'] = str(col['_id'])
                    if 'created_at' in col:
                        col['created_at'] = col['created_at'].strftime('%Y-%m-%d %H:%M:%S')
                return collections
            except Exception as e:
                return []
    
    def get_student_history(self, identifier, search_field='roll_no'):
        with self.lock:
            try:
                db = self._get_connection()
                all_metadata = list(db.lecture_metadata.find().sort('created_at', -1))
                history = []
                
                for meta in all_metadata:
                    collection = db[meta['collection_name']]
                    doc = collection.find_one({})
                    
                    if not doc:
                        continue
                    
                    for session_name, session_data in doc.get('sessions', {}).items():
                        students = session_data.get('students', {})
                        
                        student = None
                        if search_field == 'roll_no':
                            student = students.get(identifier)
                        else:
                            for s in students.values():
                                if s.get('name', '').upper() == identifier.upper():
                                    student = s
                                    break
                        
                        if student:
                            history.append({
                                'date': meta['date'],
                                'session': session_name,
                                'status': student.get('status', 'N/A'),
                                'first_seen': student.get('timestamps', {}).get('first_seen', 'N/A'),
                                'last_seen': student.get('timestamps', {}).get('last_seen', 'N/A'),
                                'present_duration': student.get('durations', {}).get('total_present_human', '0 sec'),
                                'department': meta.get('department', ''),
                                'classroom': meta.get('classroom', '')
                            })
                
                return history
            except Exception as e:
                return []
    
    def clear_session_data(self, collection_name, session_name):
        with self.lock:
            try:
                db = self._get_connection()
                collection = db[collection_name]
                doc = collection.find_one({})
                
                if not doc or session_name not in doc.get('sessions', {}):
                    return 0
                
                students = doc['sessions'][session_name]['students']
                current_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                update_fields = {}
                for roll_no in students.keys():
                    prefix = f'sessions.{session_name}.students.{roll_no}'
                    update_fields[f'{prefix}.status'] = 'Absent'
                    update_fields[f'{prefix}.timestamps.first_seen'] = None
                    update_fields[f'{prefix}.timestamps.last_seen'] = None
                    update_fields[f'{prefix}.timestamps.present_timer_start'] = None
                    update_fields[f'{prefix}.timestamps.absence_timer_start'] = None
                    update_fields[f'{prefix}.timestamps.temp_absent_time'] = None
                    update_fields[f'{prefix}.timestamps.perm_absent_time'] = None
                    update_fields[f'{prefix}.timestamps.last_updated'] = current_time_str
                    update_fields[f'{prefix}.durations.total_present_seconds'] = 0
                    update_fields[f'{prefix}.durations.total_absent_seconds'] = 0
                    update_fields[f'{prefix}.durations.total_present_human'] = '0 sec'
                    update_fields[f'{prefix}.durations.total_absent_human'] = '0 sec'
                    update_fields[f'{prefix}.flags.manual_override'] = False
                    update_fields[f'{prefix}.flags.is_temp_absent'] = False
                    update_fields[f'{prefix}.flags.is_perm_absent'] = False
                
                collection.update_one({}, {'$set': update_fields})
                return len(students)
            except Exception as e:
                return 0
    
    def generate_excel_report(self, collection_name, session_name=None):
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            db = self._get_connection()
            collection = db[collection_name]
            doc = collection.find_one({})
            
            if not doc:
                return None
            
            summary = wb.create_sheet("Summary", 0)
            summary['A1'] = f"Attendance Report - {doc['date']}"
            summary['A1'].font = Font(bold=True, size=14)
            summary['A2'] = f"Department: {doc['department']}"
            summary['A3'] = f"Classroom: {doc['classroom']}"
            summary['A4'] = f"Teacher: {doc['teacher_name']}"
            
            sessions_to_export = [session_name] if session_name else ALL_SESSIONS
            
            for sess in sessions_to_export:
                if sess not in doc['sessions']:
                    continue
                
                sheet = wb.create_sheet(sess[:31])
                students = list(doc['sessions'][sess]['students'].values())
                
                headers = ['Roll No', 'Name', 'Status', 'First Seen', 'Last Seen', 
                          'Present Duration', 'Absent Duration']
                
                for col, header in enumerate(headers, 1):
                    cell = sheet.cell(1, col, header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
                
                for row, student in enumerate(students, 2):
                    sheet.cell(row, 1, student['roll_no'])
                    sheet.cell(row, 2, student['name'])
                    sheet.cell(row, 3, student['status'])
                    sheet.cell(row, 4, student['timestamps'].get('first_seen', 'N/A'))
                    sheet.cell(row, 5, student['timestamps'].get('last_seen', 'N/A'))
                    sheet.cell(row, 6, student['durations'].get('total_present_human', '0 sec'))
                    sheet.cell(row, 7, student['durations'].get('total_absent_human', '0 sec'))
            
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            return excel_file
        except Exception as e:
            print(f"Error generating Excel: {e}")
            return None
    
    def close(self):
        if self.client:
            self.client.close()


class AttendanceSystem:
    def __init__(self, mode, year_input, department, classroom, teacher_name, camera_ids=None):
        self.mode = mode
        self.config = AttendanceConfig()
        self.student_status = {}
        self.stop_event = Event()
        self.attendance_count = 0
        self.total_students = 0
        self.search_mode = 'roll' if mode == AttendanceConfig.MODE_ROLL_NO else 'name'
        self.current_faces_count = 0
        
        self.year_input = year_input
        self.department = department
        self.classroom = classroom
        self.teacher_name = teacher_name
        self.camera_ids = camera_ids or ['CAM-01']
        
        self.db_manager = DatabaseManager(MONGODB_CONFIG)
        
        sheet_name = self.db_manager._get_sheet_name(department, year_input)
        
        self.template_headers, self.template_data = self.db_manager.load_students_from_excel(
            self.config.TEMPLATE_FILE, sheet_name
        )
        self.total_students = len(self.template_data)
        
        self.current_session = None
        self.current_collection = None
        self.current_date = None
        
        year_code = self.db_manager._get_year_code(year_input)
        dept_year_code = f"{department}_{year_code}"
        self.class_names, self.known_encodings = self._load_training_data(dept_year_code)
    
    def _load_training_data(self, dept_year_code):
        try:
            mode_name = 'Name' if self.mode == self.config.MODE_NAME else 'Roll No.'
            path = FilePathResolver.find_training_folder(dept_year_code, mode_name)
            
            images = []
            class_names = []
            image_files = [f for f in os.listdir(path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
            
            for filename in image_files:
                img_path = os.path.join(path, filename)
                img = cv2.imread(img_path)
                if img is None:
                    continue
                images.append(img)
                class_names.append(os.path.splitext(filename)[0])
            
            print(f"✓ Loaded {len(class_names)} training images")
            encodings = self._find_encodings(images)
            return class_names, encodings
        except Exception as e:
            print(f"Error loading training data: {e}")
            raise
    
    def _find_encodings(self, images):
        encode_list = []
        for img in images:
            try:
                img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                encodings = face_recognition.face_encodings(img_rgb)
                if encodings:
                    encode_list.append(encodings[0])
            except:
                pass
        return encode_list
    
    def mark_attendance(self, identifier):
        identifier = identifier.upper()
        current_time = datetime.now()
        
        if not self.current_session or not self.current_collection:
            return False
        
        if identifier not in self.student_status:
            self.student_status[identifier] = {
                'last_seen': current_time,
                'status': 'Present',
                'timer_start': None
            }
            self.attendance_count += 1
        else:
            self.student_status[identifier]['last_seen'] = current_time
            self.student_status[identifier]['status'] = 'Present'
            self.student_status[identifier]['timer_start'] = None
        
        success = self.db_manager.update_student_attendance(
            self.current_collection,
            self.current_session,
            identifier,
            'Present',
            manual=False
        )
        
        return success
    
    def check_absence_continuously(self):
        while not self.stop_event.is_set():
            if not self.current_session:
                sleep(self.config.ABSENCE_CHECK_INTERVAL)
                continue
            
            current_time = datetime.now()
            for identifier, info in list(self.student_status.items()):
                last_seen = info['last_seen']
                current_status = info['status']
                timer_start = info['timer_start']
                
                time_since_seen = current_time - last_seen
                
                if current_status == 'Present':
                    if time_since_seen >= timedelta(seconds=self.config.ABSENCE_DETECTION_DELAY):
                        if timer_start is None:
                            timer_start = current_time
                            self.student_status[identifier]['timer_start'] = timer_start
                        
                        time_in_absence = current_time - timer_start
                        
                        if time_in_absence >= timedelta(seconds=self.config.PERMANENT_ABSENT_THRESHOLD):
                            self.student_status[identifier]['status'] = 'Permanently Absent'
                            self.db_manager.update_student_attendance(
                                self.current_collection,
                                self.current_session,
                                identifier,
                                'Permanently Absent'
                            )
                        elif time_in_absence >= timedelta(seconds=self.config.TEMPORARY_ABSENT_THRESHOLD):
                            if current_status != 'Temporary Absent':
                                self.student_status[identifier]['status'] = 'Temporary Absent'
                                self.db_manager.update_student_attendance(
                                    self.current_collection,
                                    self.current_session,
                                    identifier,
                                    'Temporary Absent'
                                )
            
            sleep(self.config.ABSENCE_CHECK_INTERVAL)
    
    def process_frame(self, frame):
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
        
        self.current_faces_count = len(face_locations)
        session = get_current_session()
        date_str = datetime.now().strftime('%Y-%m-%d')
        
        if session != self.current_session or date_str != self.current_date:
            print(f"\nSession Changed: {self.current_session} -> {session}")
            
            self.current_session = session
            self.current_date = date_str
            self.student_status = {}
            self.attendance_count = 0
            
            self.current_collection = self.db_manager.create_or_get_daily_collection(
                self.department,
                self.year_input,
                date_str,
                self.classroom,
                self.teacher_name,
                self.template_headers,
                self.template_data,
                camera_ids=self.camera_ids
            )
        
        for face_encoding, face_loc in zip(face_encodings, face_locations):
            matches = face_recognition.compare_faces(self.known_encodings, face_encoding, tolerance=0.6)
            face_distances = face_recognition.face_distance(self.known_encodings, face_encoding)
            
            if len(face_distances) > 0:
                best_match_idx = np.argmin(face_distances)
                if matches[best_match_idx]:
                    name = self.class_names[best_match_idx].upper()
                    color = (0, 255, 0)
                    self.mark_attendance(name)
                else:
                    name = "UNKNOWN"
                    color = (0, 0, 255)
            else:
                name = "UNKNOWN"
                color = (0, 0, 255)
            
            y1, x2, y2, x1 = [coord * 4 for coord in face_loc]
            cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
            cv2.rectangle(frame, (x1, y2 - 35), (x2, y2), color, cv2.FILLED)
            cv2.putText(frame, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        
        overlay = frame.copy()
        cv2.rectangle(overlay, (0, 0), (frame.shape[1], 80), (0, 0, 0), -1)
        cv2.addWeighted(overlay, 0.7, frame, 0.3, 0, frame)
        
        cv2.putText(frame, f"Session: {session}", (20, 25), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
        cv2.putText(frame, f"Attendance: {self.attendance_count}/{self.total_students}", (20, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0) if self.attendance_count > 0 else (255, 255, 255), 2)
        cv2.putText(frame, f"Faces: {self.current_faces_count}", (20, 70), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)
        
        return frame
    
    def stop(self):
        self.stop_event.set()


@app.route('/api/health', methods=['GET'])
def health_check():
    global attendance_system, camera_running
    try:
        client = MongoClient(MONGODB_CONFIG['host'], MONGODB_CONFIG['port'], serverSelectionTimeoutMS=2000)
        client.server_info()
        db_connected = True
        client.close()
    except:
        db_connected = False
    return jsonify({
        'status': 'healthy' if db_connected else 'degraded',
        'database': 'connected' if db_connected else 'disconnected',
        'camera_status': 'running' if camera_running else 'stopped',
        'system_initialized': attendance_system is not None,
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/camera/start', methods=['POST'])
def start_camera():
    global attendance_system, camera_running
    try:
        data = request.get_json()
        mode = data.get('mode', 1)
        year_input = data.get('year', '')
        department = data.get('department', '')
        classroom = data.get('classroom', '')
        teacher_name = data.get('teacher_name', '')
        camera_ids = data.get('camera_ids', ['CAM-01'])
        
        if not all([year_input, department, classroom, teacher_name]):
            return jsonify({'success': False, 'message': 'All fields required'}), 400
        
        if camera_running:
            return jsonify({'success': False, 'message': 'Camera already running'}), 400
        
        attendance_system = AttendanceSystem(
            mode=mode,
            year_input=year_input,
            department=department,
            classroom=classroom,
            teacher_name=teacher_name,
            camera_ids=camera_ids
        )
        
        absence_thread = threading.Thread(target=attendance_system.check_absence_continuously)
        absence_thread.daemon = True
        absence_thread.start()
        
        camera_running = True
        year_code = YEAR_MAPPING.get(year_input, 'B.Tech')
        
        return jsonify({
            'success': True,
            'message': 'Camera started',
            'year_code': year_code,
            'sheet_loaded': f"{department}_{year_code}"
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/camera/stop', methods=['POST'])
def stop_camera():
    global attendance_system, camera_running
    try:
        if not camera_running:
            return jsonify({'success': False, 'message': 'Camera not running'}), 400
        
        camera_running = False
        if attendance_system:
            attendance_system.stop()
        
        return jsonify({'success': True, 'message': 'Camera stopped'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/camera/status', methods=['GET'])
def camera_status():
    global camera_running, attendance_system
    status = {
        'running': camera_running,
        'attendance_count': 0,
        'current_faces': 0,
        'current_session': None,
        'current_collection': None,
        'camera_ids': []
    }
    if attendance_system:
        status['attendance_count'] = attendance_system.attendance_count
        status['current_faces'] = attendance_system.current_faces_count
        status['current_session'] = attendance_system.current_session
        status['current_collection'] = attendance_system.current_collection
        status['camera_ids'] = attendance_system.camera_ids
    return jsonify(status)

def generate_frames():
    global attendance_system, camera_running
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        return
    try:
        while camera_running:
            ret, frame = cap.read()
            if not ret:
                break
            if attendance_system:
                frame = attendance_system.process_frame(frame)
            ret, buffer = cv2.imencode('.jpg', frame)
            if not ret:
                continue
            frame = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
    except Exception as e:
        print(f"Error in video: {e}")
    finally:
        cap.release()

@app.route('/api/video_feed')
def video_feed():
    if not camera_running:
        return jsonify({'error': 'Camera not running'}), 400
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/api/current-session', methods=['GET'])
def get_current_session_data():
    try:
        if attendance_system:
            if not attendance_system.current_collection or not attendance_system.current_session:
                return jsonify({'success': True, 'active': False})
            
            attendance = attendance_system.db_manager.get_session_attendance(
                attendance_system.current_collection,
                attendance_system.current_session
            )
            summary = attendance_system.db_manager.get_session_summary(
                attendance_system.current_collection,
                attendance_system.current_session
            )
            
            return jsonify({
                'success': True,
                'active': True,
                'collection_name': attendance_system.current_collection,
                'session_name': attendance_system.current_session,
                'date': attendance_system.current_date,
                'summary': summary,
                'attendance': attendance
            })
        else:
            return jsonify({'success': True, 'active': False})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/attendance/update', methods=['POST'])
def update_attendance_manual():
    try:
        data = request.get_json()
        collection_name = data.get('collection_name')
        session_name = data.get('session_name')
        roll_no = data.get('roll_no')
        status = data.get('status')
        
        if not all([collection_name, session_name, roll_no, status]):
            return jsonify({'success': False, 'error': 'Missing fields'}), 400
        
        db_manager = DatabaseManager(MONGODB_CONFIG)
        success = db_manager.update_student_attendance(
            collection_name, session_name, roll_no, status, manual=True
        )
        db_manager.close()
        
        if success:
            return jsonify({'success': True, 'message': f'Marked as {status}'})
        else:
            return jsonify({'success': False, 'error': 'Failed to update'}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/attendance/clear', methods=['POST'])
def clear_attendance_data():
    try:
        data = request.get_json()
        collection_name = data.get('collection_name')
        session_name = data.get('session_name')
        
        if not all([collection_name, session_name]):
            return jsonify({'success': False, 'error': 'Missing fields'}), 400
        
        if attendance_system:
            attendance_system.student_status = {}
            attendance_system.attendance_count = 0
            count = attendance_system.db_manager.clear_session_data(collection_name, session_name)
        else:
            db_manager = DatabaseManager(MONGODB_CONFIG)
            count = db_manager.clear_session_data(collection_name, session_name)
            db_manager.close()
        
        return jsonify({'success': True, 'message': f'Cleared {count} students', 'count': count})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/collections', methods=['GET'])
def get_collections():
    try:
        db_manager = DatabaseManager(MONGODB_CONFIG)
        collections = db_manager.get_all_daily_collections()
        db_manager.close()
        return jsonify({'success': True, 'data': collections})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/student/history/<identifier>', methods=['GET'])
def get_student_history(identifier):
    try:
        search_field = request.args.get('search_by', 'roll_no')
        db_manager = DatabaseManager(MONGODB_CONFIG)
        history = db_manager.get_student_history(identifier, search_field)
        db_manager.close()
        
        total = len(history)
        present = sum(1 for h in history if h['status'] == 'Present')
        percentage = (present / total * 100) if total > 0 else 0
        
        return jsonify({
            'success': True,
            'identifier': identifier,
            'statistics': {
                'total_sessions': total,
                'present': present,
                'absent': total - present,
                'attendance_percentage': round(percentage, 2)
            },
            'history': history
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reports/export/<collection_name>', methods=['GET'])
def export_report(collection_name):
    try:
        session_name = request.args.get('session')
        db_manager = DatabaseManager(MONGODB_CONFIG)
        excel_file = db_manager.generate_excel_report(collection_name, session_name)
        db_manager.close()
        
        if excel_file:
            filename = f"{collection_name}_{session_name if session_name else 'all'}.xlsx"
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        else:
            return jsonify({'success': False, 'error': 'Failed to generate'}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/reports')
def reports_page():
    return render_template_string(REPORTS_HTML_TEMPLATE)

@app.route('/student')
def student_page():
    return render_template_string(STUDENT_HTML_TEMPLATE)


HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Attendance System</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',sans-serif;background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);min-height:100vh;padding:15px}
.container{max-width:1800px;margin:0 auto}
.header{background:white;padding:20px 25px;border-radius:12px;box-shadow:0 5px 15px rgba(0,0,0,0.2);margin-bottom:15px;display:flex;justify-content:space-between;align-items:center}
.header h1{color:#667eea;font-size:1.8em}
.nav-links{display:flex;gap:10px}
.nav-links a{padding:8px 16px;background:#667eea;color:white;text-decoration:none;border-radius:6px;font-weight:600;font-size:0.9em;transition:background 0.3s}
.nav-links a:hover{background:#5568d3}
.camera-section{background:white;padding:15px;border-radius:12px;box-shadow:0 5px 15px rgba(0,0,0,0.1);margin-bottom:15px}
.config-row{display:grid;grid-template-columns:repeat(5,1fr) auto;gap:10px;margin-bottom:10px;align-items:end}
.form-group{display:flex;flex-direction:column;gap:5px}
.form-group label{font-weight:600;color:#555;font-size:0.85em}
.form-group input,.form-group select{padding:8px 10px;border:2px solid #ddd;border-radius:6px;font-size:0.9em}
.btn{padding:8px 16px;border:none;border-radius:6px;font-size:0.9em;font-weight:bold;cursor:pointer;transition:all 0.3s}
.btn-confirm{background:#3b82f6;color:white}
.btn-confirm:hover{background:#2563eb}
.btn:disabled{background:#9ca3af;cursor:not-allowed;opacity:0.6}
.camera-controls{display:flex;gap:10px;align-items:center}
.btn-start{background:#10b981;color:white;flex:1}
.btn-start:hover:not(:disabled){background:#059669}
.btn-stop{background:#ef4444;color:white;flex:1}
.btn-stop:hover:not(:disabled){background:#dc2626}
.status-badge{display:inline-flex;align-items:center;gap:6px;padding:6px 12px;background:#f3f4f6;border-radius:6px;font-weight:600;font-size:0.85em}
.status-dot{width:10px;height:10px;border-radius:50%;background:#ef4444}
.status-dot.active{background:#10b981;animation:pulse 2s infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:0.5}}
.video-row{display:grid;grid-template-columns:1fr auto;gap:15px}
.video-container{background:#000;border-radius:8px;overflow:hidden;height:400px;display:flex;align-items:center;justify-content:center}
.video-container img{max-width:100%;max-height:100%;object-fit:contain}
.video-placeholder{color:#9ca3af;font-size:1.1em;text-align:center}
.info-stats{display:flex;flex-direction:column;gap:8px;min-width:200px}
.stat-box{background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);padding:12px;border-radius:8px;color:white;text-align:center}
.stat-box h3{font-size:0.7em;opacity:0.9;margin-bottom:4px}
.stat-box .value{font-size:1.6em;font-weight:bold}
.attendance-section{background:white;padding:20px;border-radius:12px;box-shadow:0 5px 15px rgba(0,0,0,0.1)}
.section-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:15px}
.section-header h2{color:#333;font-size:1.4em}
.action-buttons{display:flex;gap:8px}
.btn-refresh{background:#667eea;color:white;border:none;padding:8px 16px;border-radius:6px;cursor:pointer;font-weight:600;font-size:0.9em}
.btn-refresh:hover{background:#5568d3}
.btn-clear{background:#f59e0b;color:white;border:none;padding:8px 16px;border-radius:6px;cursor:pointer;font-weight:600;font-size:0.9em}
.btn-clear:hover{background:#d97706}
.stats-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin-bottom:15px}
.stat-card{background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);padding:15px;border-radius:8px;color:white;text-align:center}
.stat-card h3{font-size:0.7em;text-transform:uppercase;margin-bottom:6px;opacity:0.9}
.stat-card .value{font-size:1.8em;font-weight:bold}
.table-container{overflow-x:auto;max-height:450px;overflow-y:auto}
table{width:100%;border-collapse:collapse;font-size:0.85em}
table th{background:#f8f9fa;padding:10px 8px;text-align:left;font-weight:bold;color:#333;position:sticky;top:0;z-index:10}
table td{padding:10px 8px;border-bottom:1px solid #eee}
table tr:hover{background:#f8f9fa}
.status-badge-table{padding:3px 10px;border-radius:20px;font-size:0.8em;font-weight:bold;display:inline-block}
.status-badge-table.present{background:#d1fae5;color:#065f46}
.status-badge-table.absent{background:#fee2e2;color:#991b1b}
.status-badge-table.temporary-absent{background:#fef3c7;color:#92400e}
.status-badge-table.permanently-absent{background:#fee2e2;color:#991b1b}
.status-toggle{padding:4px 8px;border:none;border-radius:4px;cursor:pointer;font-size:0.75em;font-weight:bold}
.status-toggle.to-present{background:#10b981;color:white}
.status-toggle.to-absent{background:#ef4444;color:white}
.manual-badge{background:#fbbf24;color:#78350f;padding:2px 6px;border-radius:4px;font-size:0.7em;margin-left:5px}
.config-confirmed{background:#d1fae5;border:2px solid #10b981;padding:10px;border-radius:6px;margin-bottom:10px;display:none}
.config-confirmed.show{display:block}
.config-confirmed strong{color:#065f46}
</style>
</head>
<body>
<div class="container">
<div class="header">
<h1>📸 Attendance System</h1>
<div class="nav-links">
<a href="/">Dashboard</a>
<a href="/reports">Reports</a>
<a href="/student">Student View</a>
</div>
</div>
<div class="camera-section">
<div class="config-row">
<div class="form-group">
<label>Year (2022-2025)</label>
<input type="text" id="year" placeholder="e.g., 2022">
</div>
<div class="form-group">
<label>Department</label>
<input type="text" id="department" placeholder="e.g., CSBS">
</div>
<div class="form-group">
<label>Classroom</label>
<input type="text" id="classroom" placeholder="e.g., 301">
</div>
<div class="form-group">
<label>Teacher</label>
<input type="text" id="teacherInput" placeholder="Prof. Name">
</div>
<div class="form-group">
<label>Camera ID</label>
<input type="text" id="cameraId" placeholder="CAM-01" value="CAM-01">
</div>
<button class="btn btn-confirm" onclick="confirmConfig()">Confirm</button>
</div>
<div class="config-confirmed" id="confirmedBanner">
<strong>Config:</strong> <span id="confirmedText"></span> | <strong>Sheet:</strong> <span id="sheetName"></span> | <strong>Cam:</strong> <span id="confirmedCamera"></span>
</div>
<div class="camera-controls" style="margin-bottom:10px">
<div class="form-group" style="flex:0 0 150px">
<label>Mode</label>
<select id="modeSelect">
<option value="1">By Name</option>
<option value="2">By Roll Number</option>
</select>
</div>
<button class="btn btn-start" id="startBtn" onclick="startCamera()" disabled>Start</button>
<button class="btn btn-stop" id="stopBtn" onclick="stopCamera()" disabled>Stop</button>
<div class="status-badge">
<div class="status-dot" id="statusDot"></div>
<span id="statusText">Stopped</span>
</div>
</div>
<div class="video-row">
<div class="video-container" id="videoContainer">
<div class="video-placeholder">Configure and confirm</div>
</div>
<div class="info-stats">
<div class="stat-box"><h3>Faces</h3><div class="value" id="currentFaces">0</div></div>
<div class="stat-box"><h3>Session</h3><div class="value" style="font-size:1.2em" id="currentSession">-</div></div>
<div class="stat-box"><h3>Total</h3><div class="value" id="quickTotal">-</div></div>
<div class="stat-box"><h3>Present</h3><div class="value" id="quickPresent">-</div></div>
<div class="stat-box"><h3>Absent</h3><div class="value" id="quickAbsent">-</div></div>
</div>
</div>
</div>
<div class="attendance-section">
<div class="section-header">
<h2>Current Session</h2>
<div class="action-buttons">
<button class="btn-refresh" onclick="refreshData()">Refresh</button>
<button class="btn-clear" onclick="clearSessionData()" id="clearBtn" disabled>Clear</button>
</div>
</div>
<div class="stats-grid">
<div class="stat-card"><h3>Total</h3><div class="value" id="totalStudents">-</div></div>
<div class="stat-card"><h3>Present</h3><div class="value" id="presentCount">-</div></div>
<div class="stat-card"><h3>Absent</h3><div class="value" id="absentCount">-</div></div>
<div class="stat-card"><h3>Temp Absent</h3><div class="value" id="tempAbsentCount">-</div></div>
<div class="stat-card"><h3>Attendance %</h3><div class="value" id="attendancePercentage">-</div></div>
</div>
<div class="table-container">
<table>
<thead>
<tr><th>Roll</th><th>Name</th><th>Status</th><th>First</th><th>Last</th><th>Present</th><th>Absent</th><th>Action</th></tr>
</thead>
<tbody id="attendanceBody">
<tr><td colspan="8" style="text-align:center">Configure settings</td></tr>
</tbody>
</table>
</div>
</div>
</div>
<script>
const API=''+window.location.origin+'/api';
const YM={'2022':'B.Tech','2023':'TY','2024':'SY','2025':'FY'};
let running=false,interval=null,col=null,sess=null,conf=false,cfg={};
function confirmConfig(){
const y=document.getElementById('year').value.trim();
const d=document.getElementById('department').value.trim();
const c=document.getElementById('classroom').value.trim();
const t=document.getElementById('teacherInput').value.trim();
const cam=document.getElementById('cameraId').value.trim();
if(!y||!d||!c||!t){alert('Fill all fields');return}
const yc=YM[y]||'B.Tech';
const sn=`${d}_${yc}`;
cfg={year:y,department:d,classroom:c,teacher_name:t,camera_ids:[cam]};
conf=true;
document.getElementById('confirmedText').textContent=`${y}|${d}|${c}|${t}`;
document.getElementById('sheetName').textContent=sn;
document.getElementById('confirmedCamera').textContent=cam;
document.getElementById('confirmedBanner').classList.add('show');
document.getElementById('startBtn').disabled=false;
loadSession();
}
async function startCamera(){
if(!conf){alert('Confirm first');return}
const m=document.getElementById('modeSelect').value;
const sb=document.getElementById('startBtn');
const stb=document.getElementById('stopBtn');
sb.disabled=true;sb.textContent='Starting...';
try{
const r=await fetch(`${API}/camera/start`,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({mode:parseInt(m),...cfg})});
const d=await r.json();
if(d.success){
running=true;updateStatus(true);sb.disabled=true;stb.disabled=false;sb.textContent='Start';
document.getElementById('videoContainer').innerHTML='<img src="'+API+'/video_feed?t='+Date.now()+'">';
startRefresh();await loadSession();
}else throw new Error(d.message||d.error);
}catch(e){alert('Error: '+e.message);sb.disabled=false;sb.textContent='Start'}
}
async function stopCamera(){
const sb=document.getElementById('startBtn');
const stb=document.getElementById('stopBtn');
stb.disabled=true;stb.textContent='Stopping...';
try{
const r=await fetch(`${API}/camera/stop`,{method:'POST'});
const d=await r.json();
if(d.success){
running=false;updateStatus(false);sb.disabled=false;stb.disabled=true;stb.textContent='Stop';
document.getElementById('videoContainer').innerHTML='<div class="video-placeholder">Stopped</div>';
stopRefresh();
}
}catch(e){alert('Error: '+e.message);stb.disabled=false;stb.textContent='Stop'}
}
function updateStatus(r){
const dot=document.getElementById('statusDot');
const txt=document.getElementById('statusText');
if(r){dot.classList.add('active');txt.textContent='Running'}
else{dot.classList.remove('active');txt.textContent='Stopped'}
}
async function updateInfo(){
try{
const r=await fetch(`${API}/camera/status`);
const d=await r.json();
document.getElementById('currentFaces').textContent=d.current_faces||0;
document.getElementById('currentSession').textContent=d.current_session||'-';
if(d.current_session&&d.current_session!==sess){sess=d.current_session;await loadSession()}
}catch(e){console.error(e)}
}
async function loadSession(){
if(!conf)return;
try{
const r=await fetch(`${API}/current-session`);
const d=await r.json();
if(d.success&&d.active){
col=d.collection_name;sess=d.session_name;
updateStats(d.summary);displayData(d.attendance);
document.getElementById('clearBtn').disabled=false;
}
}catch(e){console.error(e)}
}
function updateStats(s){
document.getElementById('totalStudents').textContent=s.total||0;
document.getElementById('presentCount').textContent=s.present||0;
document.getElementById('absentCount').textContent=s.absent||0;
document.getElementById('tempAbsentCount').textContent=s.temporary_absent||0;
document.getElementById('attendancePercentage').textContent=(s.attendance_percentage||0).toFixed(2)+'%';
document.getElementById('quickTotal').textContent=s.total||0;
document.getElementById('quickPresent').textContent=s.present||0;
document.getElementById('quickAbsent').textContent=s.absent||0;
}
function displayData(recs){
const tb=document.getElementById('attendanceBody');
if(!recs||recs.length===0){tb.innerHTML='<tr><td colspan="8" style="text-align:center">No records</td></tr>';return}
tb.innerHTML='';
recs.forEach(r=>{
const row=document.createElement('tr');
const st=r.status||'Absent';
let bc='absent';
if(st==='Present')bc='present';
else if(st==='Temporary Absent')bc='temporary-absent';
else if(st==='Permanently Absent')bc='permanently-absent';
const ts=r.timestamps||{};
const du=r.durations||{};
const fl=r.flags||{};
row.innerHTML=`<td>${r.roll_no||'-'}</td><td>${r.name||'-'}</td><td><span class="status-badge-table ${bc}">${st}</span>${fl.manual_override?'<span class="manual-badge">Manual</span>':''}</td><td>${ts.first_seen||'N/A'}</td><td>${ts.last_seen||'N/A'}</td><td>${du.total_present_human||'0 sec'}</td><td>${du.total_absent_human||'0 sec'}</td><td><button class="status-toggle ${st==='Present'?'to-absent':'to-present'}" onclick="toggleAtt('${r.roll_no}','${st}')">${st==='Present'?'Mark Absent':'Mark Present'}</button></td>`;
tb.appendChild(row);
});
}
async function toggleAtt(roll,cur){
if(!col||!sess){alert('No session');return}
const ns=cur==='Present'?'Absent':'Present';
const btn=event.target;
btn.disabled=true;btn.textContent='Updating...';
try{
const r=await fetch(`${API}/attendance/update`,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({collection_name:col,session_name:sess,roll_no:roll,status:ns})});
const d=await r.json();
if(d.success)await refreshData();
else alert('Failed: '+(d.error||'Unknown'));
}catch(e){alert('Error: '+e.message)}finally{btn.disabled=false;btn.textContent=cur==='Present'?'Mark Absent':'Mark Present'}
}
async function refreshData(){await loadSession()}
async function clearSessionData(){
if(!col||!sess){alert('No session');return}
if(!confirm('Clear all data?'))return;
try{
const r=await fetch(`${API}/attendance/clear`,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({collection_name:col,session_name:sess})});
const d=await r.json();
if(d.success){alert(d.message);await refreshData()}
else alert('Failed: '+(d.error||'Unknown'));
}catch(e){console.error(e)}
}
function startRefresh(){
if(interval)clearInterval(interval);
updateInfo();
interval=setInterval(()=>{updateInfo();refreshData()},5000);
}
function stopRefresh(){if(interval)clearInterval(interval)}
window.onload=()=>updateStatus(false);
</script>
</body>
</html>'''

REPORTS_HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Reports</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',sans-serif;background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);min-height:100vh;padding:20px}
.container{max-width:1400px;margin:0 auto}
.header{background:white;padding:25px 30px;border-radius:15px;box-shadow:0 10px 30px rgba(0,0,0,0.2);margin-bottom:20px;display:flex;justify-content:space-between;align-items:center}
.header h1{color:#667eea;font-size:2em}
.nav-links{display:flex;gap:15px}
.nav-links a{padding:10px 20px;background:#667eea;color:white;text-decoration:none;border-radius:8px;font-weight:bold;transition:background 0.3s}
.nav-links a:hover{background:#5568d3}
.reports-section{background:white;padding:25px;border-radius:15px;box-shadow:0 5px 15px rgba(0,0,0,0.1);margin-bottom:20px}
.reports-section h2{color:#333;margin-bottom:20px;font-size:1.5em}
.collections-grid{display:grid;gap:15px}
.collection-card{background:linear-gradient(135deg,#f8f9fa 0%,#e9ecef 100%);padding:20px;border-radius:10px;border-left:5px solid #667eea;transition:transform 0.3s,box-shadow 0.3s}
.collection-card:hover{transform:translateY(-5px);box-shadow:0 8px 20px rgba(0,0,0,0.1)}
.collection-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:15px}
.collection-title{font-size:1.2em;font-weight:bold;color:#333}
.collection-meta{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:15px}
.meta-item{background:white;padding:10px;border-radius:6px;text-align:center}
.meta-item label{display:block;font-size:0.75em;color:#666;margin-bottom:5px}
.meta-item value{font-weight:bold;color:#667eea;font-size:0.95em}
.collection-actions{display:flex;gap:10px}
.btn{padding:8px 16px;border:none;border-radius:6px;font-weight:bold;cursor:pointer;transition:all 0.3s;font-size:0.9em}
.btn-export{background:#10b981;color:white}
.btn-export:hover{background:#059669}
.btn-view{background:#3b82f6;color:white}
.btn-view:hover{background:#2563eb}
.session-details{display:none;margin-top:15px;padding-top:15px;border-top:2px solid #e5e7eb}
.session-details.show{display:block}
.sessions-list{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-top:10px}
.session-btn{padding:10px;background:white;border:2px solid #667eea;border-radius:6px;cursor:pointer;font-weight:600;color:#667eea;transition:all 0.3s}
.session-btn:hover{background:#667eea;color:white}
.loading{text-align:center;padding:40px;color:#666;font-size:1.1em}
.no-data{text-align:center;padding:40px;color:#999;font-size:1.1em}
</style>
</head>
<body>
<div class="container">
<div class="header">
<h1>📊 Reports</h1>
<div class="nav-links">
<a href="/">Dashboard</a>
<a href="/reports">Reports</a>
<a href="/student">Student View</a>
</div>
</div>
<div class="reports-section">
<h2>Daily Collections</h2>
<div id="collectionsContainer" class="collections-grid">
<div class="loading">Loading...</div>
</div>
</div>
</div>
<script>
const API=window.location.origin+'/api';
async function loadCollections(){
try{
const r=await fetch(`${API}/collections`);
const d=await r.json();
if(d.success&&d.data.length>0)displayCollections(d.data);
else document.getElementById('collectionsContainer').innerHTML='<div class="no-data">No collections found</div>';
}catch(e){document.getElementById('collectionsContainer').innerHTML='<div class="no-data">Error loading</div>'}
}
function displayCollections(cols){
const c=document.getElementById('collectionsContainer');
c.innerHTML='';
cols.forEach(col=>{
const card=document.createElement('div');
card.className='collection-card';
card.innerHTML=`<div class="collection-header"><div class="collection-title">${col.collection_name}</div></div><div class="collection-meta"><div class="meta-item"><label>Date</label><value>${col.date}</value></div><div class="meta-item"><label>Department</label><value>${col.department}</value></div><div class="meta-item"><label>Classroom</label><value>${col.classroom}</value></div></div><div class="collection-meta"><div class="meta-item"><label>Teacher</label><value>${col.teacher_name}</value></div><div class="meta-item"><label>Year</label><value>${col.year} (${col.year_code})</value></div><div class="meta-item"><label>Created</label><value>${col.created_at}</value></div></div><div class="collection-actions"><button class="btn btn-view" onclick="toggleSessions('${col.collection_name}')">View Sessions</button><button class="btn btn-export" onclick="exportReport('${col.collection_name}')">Export All</button></div><div class="session-details" id="sessions-${col.collection_name}"><h3 style="margin-bottom:10px;color:#555">Export by Session:</h3><div class="sessions-list"><button class="session-btn" onclick="exportSession('${col.collection_name}','Session 1')">Session 1</button><button class="session-btn" onclick="exportSession('${col.collection_name}','Session 2')">Session 2</button><button class="session-btn" onclick="exportSession('${col.collection_name}','Session 3')">Session 3</button><button class="session-btn" onclick="exportSession('${col.collection_name}','Session 4')">Session 4</button><button class="session-btn" onclick="exportSession('${col.collection_name}','Session 5')">Session 5</button><button class="session-btn" onclick="exportSession('${col.collection_name}','Session 6')">Session 6</button><button class="session-btn" onclick="exportSession('${col.collection_name}','Session 7')">Session 7</button><button class="session-btn" onclick="exportSession('${col.collection_name}','Session 8')">Session 8</button></div></div>`;
c.appendChild(card);
});
}
function toggleSessions(cn){
const sd=document.getElementById(`sessions-${cn}`);
sd.classList.toggle('show');
}
async function exportReport(cn){
try{window.location.href=`${API}/reports/export/${cn}`}catch(e){alert('Failed to export')}
}
async function exportSession(cn,sn){
try{window.location.href=`${API}/reports/export/${cn}?session=${encodeURIComponent(sn)}`}catch(e){alert('Failed to export')}
}
window.onload=()=>loadCollections();
</script>
</body>
</html>'''

STUDENT_HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Student View</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',sans-serif;background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);min-height:100vh;padding:20px}
.container{max-width:1200px;margin:0 auto}
.header{background:white;padding:25px 30px;border-radius:15px;box-shadow:0 10px 30px rgba(0,0,0,0.2);margin-bottom:20px;display:flex;justify-content:space-between;align-items:center}
.header h1{color:#667eea;font-size:2em}
.nav-links{display:flex;gap:15px}
.nav-links a{padding:10px 20px;background:#667eea;color:white;text-decoration:none;border-radius:8px;font-weight:bold;transition:background 0.3s}
.nav-links a:hover{background:#5568d3}
.search-section{background:white;padding:25px;border-radius:15px;box-shadow:0 5px 15px rgba(0,0,0,0.1);margin-bottom:20px}
.search-form{display:grid;grid-template-columns:150px 1fr auto;gap:15px;align-items:end}
.form-group{display:flex;flex-direction:column;gap:8px}
.form-group label{font-weight:bold;color:#555;font-size:0.9em}
.form-group select,.form-group input{padding:10px;border:2px solid #ddd;border-radius:8px;font-size:1em}
.btn-search{padding:10px 30px;background:#667eea;color:white;border:none;border-radius:8px;font-weight:bold;cursor:pointer;font-size:1em;transition:background 0.3s}
.btn-search:hover{background:#5568d3}
.results-section{background:white;padding:25px;border-radius:15px;box-shadow:0 5px 15px rgba(0,0,0,0.1);display:none}
.results-section.show{display:block}
.student-info{background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);padding:20px;border-radius:10px;color:white;margin-bottom:20px}
.student-info h2{font-size:1.5em;margin-bottom:10px}
.stats-row{display:grid;grid-template-columns:repeat(4,1fr);gap:15px;margin-bottom:20px}
.stat-box{background:white;padding:20px;border-radius:10px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,0.1)}
.stat-box h3{font-size:0.8em;color:#666;margin-bottom:8px;text-transform:uppercase}
.stat-box .value{font-size:2em;font-weight:bold;color:#667eea}
.history-table{overflow-x:auto;max-height:500px;overflow-y:auto}
table{width:100%;border-collapse:collapse}
table th{background:#f8f9fa;padding:12px;text-align:left;font-weight:bold;color:#333;position:sticky;top:0;z-index:10}
table td{padding:12px;border-bottom:1px solid #eee}
table tr:hover{background:#f8f9fa}
.status-badge{padding:4px 12px;border-radius:20px;font-size:0.85em;font-weight:bold;display:inline-block}
.status-badge.present{background:#d1fae5;color:#065f46}
.status-badge.absent{background:#fee2e2;color:#991b1b}
.no-results{text-align:center;padding:40px;color:#999;font-size:1.1em}
.loading{text-align:center;padding:40px;color:#666}
</style>
</head>
<body>
<div class="container">
<div class="header">
<h1>👤 Student View</h1>
<div class="nav-links">
<a href="/">Dashboard</a>
<a href="/reports">Reports</a>
<a href="/student">Student View</a>
</div>
</div>
<div class="search-section">
<h2 style="margin-bottom:20px;color:#333">Search Student</h2>
<div class="search-form">
<div class="form-group">
<label>Search By</label>
<select id="searchBy">
<option value="roll_no">Roll Number</option>
<option value="name">Name</option>
</select>
</div>
<div class="form-group">
<label>Enter Roll or Name</label>
<input type="text" id="searchInput" placeholder="e.g., 101 or John Doe">
</div>
<button class="btn-search" onclick="searchStudent()">Search</button>
</div>
</div>
<div class="results-section" id="resultsSection">
<div class="student-info" id="studentInfo">
<h2 id="studentName">-</h2>
<p id="studentId">-</p>
</div>
<div class="stats-row">
<div class="stat-box"><h3>Total Sessions</h3><div class="value" id="totalSessions">0</div></div>
<div class="stat-box"><h3>Present</h3><div class="value" id="presentSessions">0</div></div>
<div class="stat-box"><h3>Absent</h3><div class="value" id="absentSessions">0</div></div>
<div class="stat-box"><h3>Attendance %</h3><div class="value" id="attendancePercent">0%</div></div>
</div>
<h3 style="margin-bottom:15px;color:#333">Attendance History</h3>
<div class="history-table">
<table>
<thead>
<tr><th>Date</th><th>Session</th><th>Department</th><th>Classroom</th><th>Status</th><th>First</th><th>Last</th><th>Duration</th></tr>
</thead>
<tbody id="historyBody">
<tr><td colspan="8" class="no-results">Search for a student</td></tr>
</tbody>
</table>
</div>
</div>
</div>
<script>
const API=window.location.origin+'/api';
async function searchStudent(){
const sb=document.getElementById('searchBy').value;
const id=document.getElementById('searchInput').value.trim();
if(!id){alert('Enter roll or name');return}
const hb=document.getElementById('historyBody');
hb.innerHTML='<tr><td colspan="8" class="loading">Loading...</td></tr>';
document.getElementById('resultsSection').classList.add('show');
try{
const r=await fetch(`${API}/student/history/${encodeURIComponent(id)}?search_by=${sb}`);
const d=await r.json();
if(d.success)displayData(d);
else hb.innerHTML='<tr><td colspan="8" class="no-results">Error: '+d.error+'</td></tr>';
}catch(e){hb.innerHTML='<tr><td colspan="8" class="no-results">Error loading</td></tr>'}
}
function displayData(d){
document.getElementById('studentName').textContent='Student: '+d.identifier;
document.getElementById('studentId').textContent='Search by: '+document.getElementById('searchBy').options[document.getElementById('searchBy').selectedIndex].text;
const s=d.statistics;
document.getElementById('totalSessions').textContent=s.total_sessions;
document.getElementById('presentSessions').textContent=s.present;
document.getElementById('absentSessions').textContent=s.absent;
document.getElementById('attendancePercent').textContent=s.attendance_percentage.toFixed(2)+'%';
const hb=document.getElementById('historyBody');
if(d.history.length===0){hb.innerHTML='<tr><td colspan="8" class="no-results">No records</td></tr>';return}
hb.innerHTML='';
d.history.forEach(r=>{
const row=document.createElement('tr');
const sc=r.status==='Present'?'present':'absent';
row.innerHTML=`<td>${r.date}</td><td>${r.session}</td><td>${r.department||'-'}</td><td>${r.classroom||'-'}</td><td><span class="status-badge ${sc}">${r.status}</span></td><td>${r.first_seen||'N/A'}</td><td>${r.last_seen||'N/A'}</td><td>${r.present_duration||'0 sec'}</td>`;
hb.appendChild(row);
});
}
document.getElementById('searchInput').addEventListener('keypress',e=>{if(e.key==='Enter')searchStudent()});
</script>
</body>
</html>'''

if __name__=='__main__':
    print("="*80)
    print("ATTENDANCE SYSTEM - COMPLETE PRODUCTION READY")
    print("="*80)
    print("\n✅ FEATURES:")
    print("   1. Robust file handling")
    print("   2. Training images: Training_images/DEPT_YEAR/Name|Roll No.")
    print("   3. Camera ID tracking")
    print("   4. Reports page with Excel export")
    print("   5. Student view page")
    print("   6. MongoDB nested schema")
    print("   7. Excel: DEPT_YEARCODE")
    print("   8. Full error handling")
    print("\n📁 STRUCTURE:")
    print("   Training_images/")
    print("   ├── CSBS_B.Tech/Name/")
    print("   ├── CSBS_B.Tech/Roll No./")
    print("   └── ...")
    print("\n🚀 Starting...")
    print(f"   MongoDB: {MONGODB_CONFIG['host']}:{MONGODB_CONFIG['port']}")
    print(f"   URL: http://localhost:5000")
    print("\n"+"="*80+"\n")
    try:
        app.run(host='0.0.0.0',port=5000,debug=False,threaded=True)
    except KeyboardInterrupt:
        print("\n\n👋 Stopped")
    except Exception as e:
        print(f"\n\n❌ Error: {e}")
        traceback.print_exc()