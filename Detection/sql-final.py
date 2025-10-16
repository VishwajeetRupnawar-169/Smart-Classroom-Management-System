"""
Optimized Face Recognition Attendance System with SQL Storage
- Improved absence timer logic
- SQL database storage instead of Excel
- Excel Sheet 1 as template
"""

import cv2
import numpy as np
import face_recognition
import os
import sys
from datetime import datetime, timedelta
from time import sleep
import threading
from threading import Lock, Event
from pathlib import Path
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Font

# Add utils folder to path
sys.path.append(os.path.abspath('../'))
from Excel_Format import get_current_session

class AttendanceConfig:
    """Configuration constants"""
    TEMPLATE_FILE = 'Book2.xlsx'
    DB_FILE = 'attendance.db'
    
    # Absence timing configuration (optimized)
    ABSENCE_DETECTION_DELAY = timedelta(seconds=5)  # Time before starting absence check
    TEMPORARY_ABSENT_THRESHOLD = timedelta(seconds=10)  # Temporary absent after this
    PERMANENT_ABSENT_THRESHOLD = timedelta(seconds=15)  # Permanent absent after this
    
    ABSENCE_CHECK_INTERVAL = 2  # Check every 2 seconds
    
    # Recognition mode
    MODE_NAME = 1
    MODE_ROLL_NO = 2

class DatabaseManager:
    """Handles all database operations"""
    
    def __init__(self, db_file):
        self.db_file = db_file
        self.conn = None
        self.lock = Lock()
        self._initialize_db()
    
    def _initialize_db(self):
        """Initialize database connection and create metadata table"""
        self.conn = sqlite3.connect(self.db_file, check_same_thread=False)
        self.conn.execute('''
            CREATE TABLE IF NOT EXISTS lecture_metadata (
                table_name TEXT PRIMARY KEY,
                session_name TEXT,
                date TEXT,
                start_time TEXT,
                created_at TEXT
            )
        ''')
        self.conn.commit()
    
    def load_template_from_excel(self, excel_file, sheet_name='Sheet1'):
        """Load template structure from Excel Sheet 1"""
        try:
            wb = load_workbook(excel_file, data_only=True)
            
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in {excel_file}")
            
            sheet = wb[sheet_name]
            
            # Read all data from sheet
            data = []
            headers = None
            
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                # Skip completely empty rows
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                
                # Convert row to list and handle None values
                row_data = [cell if cell is not None else '' for cell in row]
                
                if headers is None:
                    # First non-empty row is headers
                    headers = [str(h).strip() for h in row_data]
                else:
                    data.append(row_data)
            
            wb.close()
            
            if not headers:
                raise ValueError("No headers found in template")
            
            print(f"Loaded template with headers: {headers}")
            print(f"Loaded {len(data)} rows of data")
            
            return headers, data
            
        except Exception as e:
            print(f"Error loading template from Excel: {e}")
            raise
    
    def create_lecture_table(self, session_name, template_headers, template_data):
        """Create a new table for the current lecture session"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        table_name = f"lecture_{timestamp}"
        
        with self.lock:
            try:
                # Build CREATE TABLE statement with all template columns
                columns = []
                clean_headers = []
                
                for header in template_headers:
                    col_name = str(header).replace(' ', '_').replace('.', '').replace('(', '').replace(')', '').replace('-', '_').replace('/', '_')
                    if not col_name or col_name.isspace():
                        col_name = f"Column_{len(columns)}"
                    clean_headers.append(col_name)
                    columns.append(f"{col_name} TEXT")
                
                # Add attendance tracking columns if not present
                required_cols = ['Status', 'Last_Seen', 'Absence_Timer_Start']
                
                for req_col in required_cols:
                    if req_col not in clean_headers:
                        clean_headers.append(req_col)
                        columns.append(f"{req_col} TEXT")
                
                columns_str = ', '.join(columns)
                create_query = f"CREATE TABLE {table_name} (id INTEGER PRIMARY KEY AUTOINCREMENT, {columns_str})"
                
                self.conn.execute(create_query)
                
                # Insert template data with all columns filled
                for row in template_data:
                    # Create a complete row with all data
                    row_data = []
                    
                    for i, value in enumerate(row):
                        if i < len(template_headers):
                            # Convert None or empty to empty string, preserve all other values
                            if value is None:
                                row_data.append('')
                            elif isinstance(value, (int, float)):
                                row_data.append(str(value))
                            else:
                                row_data.append(str(value).strip() if value else '')
                    
                    # Pad with empty strings if row is shorter than headers
                    while len(row_data) < len(template_headers):
                        row_data.append('')
                    
                    # Add default values for attendance tracking columns
                    for req_col in required_cols:
                        if req_col not in template_headers:
                            row_data.append('')
                    
                    # Build dynamic insert query
                    placeholders = ', '.join(['?'] * len(clean_headers))
                    insert_query = f"INSERT INTO {table_name} ({', '.join(clean_headers)}) VALUES ({placeholders})"
                    
                    self.conn.execute(insert_query, row_data[:len(clean_headers)])
                
                # Record metadata
                self.conn.execute('''
                    INSERT INTO lecture_metadata (table_name, session_name, date, start_time, created_at)
                    VALUES (?, ?, ?, ?, ?)
                ''', (table_name, session_name, datetime.now().strftime('%Y-%m-%d'),
                      datetime.now().strftime('%H:%M:%S'), datetime.now().isoformat()))
                
                self.conn.commit()
                
                print(f"Created table '{table_name}' with {len(template_data)} students")
                return table_name
                
            except Exception as e:
                print(f"Error creating lecture table: {e}")
                import traceback
                traceback.print_exc()
                self.conn.rollback()
                raise
    
    def get_current_lecture_table(self, session_name):
        """Get the table for current session or create new one"""
        with self.lock:
            cursor = self.conn.execute('''
                SELECT table_name FROM lecture_metadata 
                WHERE session_name = ? AND date = ?
                ORDER BY created_at DESC LIMIT 1
            ''', (session_name, datetime.now().strftime('%Y-%m-%d')))
            
            result = cursor.fetchone()
            return result[0] if result else None
    
    def update_attendance(self, table_name, identifier, status, last_seen, absence_timer_start=None, search_mode='name'):
        """Update student attendance in database"""
        with self.lock:
            try:
                # Find the student by name or roll number
                cursor = self.conn.execute(f"PRAGMA table_info({table_name})")
                columns = [col[1] for col in cursor.fetchall()]
                
                # Determine which column to search based on mode
                if search_mode == 'roll':
                    search_columns = [col for col in columns if 'roll' in col.lower()]
                else:
                    search_columns = [col for col in columns if 'name' in col.lower() and 'roll' not in col.lower()]
                
                if not search_columns:
                    print(f"Warning: No {'roll' if search_mode == 'roll' else 'name'} column found in {table_name}")
                    return False
                
                search_col = search_columns[0]
                
                # Update query
                updates = [f"Status = ?", f"Last_Seen = ?"]
                values = [status, last_seen]
                
                if absence_timer_start:
                    updates.append(f"Absence_Timer_Start = ?")
                    values.append(absence_timer_start)
                
                values.append(identifier.upper())
                
                update_query = f"UPDATE {table_name} SET {', '.join(updates)} WHERE UPPER({search_col}) = ?"
                
                cursor = self.conn.execute(update_query, values)
                self.conn.commit()
                
                if cursor.rowcount > 0:
                    return True
                else:
                    print(f"Student '{identifier}' not found in table")
                    return False
                    
            except Exception as e:
                print(f"Error updating attendance: {e}")
                self.conn.rollback()
                return False
    
    def get_student_info(self, table_name, identifier, search_mode='name'):
        """Get student information from database"""
        with self.lock:
            try:
                cursor = self.conn.execute(f"PRAGMA table_info({table_name})")
                columns = [col[1] for col in cursor.fetchall()]
                
                # Determine which column to search based on mode
                if search_mode == 'roll':
                    search_columns = [col for col in columns if 'roll' in col.lower()]
                else:
                    search_columns = [col for col in columns if 'name' in col.lower() and 'roll' not in col.lower()]
                
                if not search_columns:
                    return None
                
                search_col = search_columns[0]
                
                cursor = self.conn.execute(
                    f"SELECT Status, Last_Seen, Absence_Timer_Start FROM {table_name} WHERE UPPER({search_col}) = ?",
                    (identifier.upper(),)
                )
                
                return cursor.fetchone()
                
            except Exception as e:
                print(f"Error getting student info: {e}")
                return None
    
    def close(self):
        """Close database connection"""
        if self.conn:
            self.conn.close()

class AttendanceSystem:
    def __init__(self, mode=AttendanceConfig.MODE_NAME):
        """Initialize the attendance system"""
        self.mode = mode
        self.config = AttendanceConfig()
        self.student_status = {}  # {identifier: {'last_seen': datetime, 'status': str, 'timer_start': datetime}}
        self.stop_event = Event()
        self.last_recognized_faces = {}
        self.attendance_count = 0  # Track attendance count
        self.total_students = 73  # Total number of students
        self.search_mode = 'roll' if mode == AttendanceConfig.MODE_ROLL_NO else 'name'
        
        # Initialize database
        self.db_manager = DatabaseManager(self.config.DB_FILE)
        
        # Load template
        self.template_headers, self.template_data = self.db_manager.load_template_from_excel(
            self.config.TEMPLATE_FILE, 
            sheet_name='Sheet1'
        )
        
        # Current lecture table
        self.current_table = None
        self.current_session = None
        
        # Load training images
        self.class_names, self.known_encodings = self._load_training_data()
    
    def _load_training_data(self):
        """Load and encode training images"""
        path = '../Training_images/Name' if self.mode == self.config.MODE_NAME else '../Training_images/Roll No.'
        
        if not os.path.exists(path):
            raise FileNotFoundError(f"Training images directory not found: {path}")
        
        images = []
        class_names = []
        
        image_files = [f for f in os.listdir(path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
        print(f"Loading {len(image_files)} training images from {path}")
        
        for filename in image_files:
            img_path = os.path.join(path, filename)
            img = cv2.imread(img_path)
            
            if img is None:
                print(f"Warning: Could not load image {filename}")
                continue
            
            images.append(img)
            class_names.append(os.path.splitext(filename)[0])
        
        if not images:
            raise ValueError("No valid training images found")
        
        print(f"Loaded class names: {class_names}")
        encodings = self._find_encodings(images)
        print(f"Generated {len(encodings)} face encodings")
        
        return class_names, encodings
    
    def _find_encodings(self, images):
        """Generate face encodings from images"""
        encode_list = []
        for idx, img in enumerate(images):
            try:
                img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                encodings = face_recognition.face_encodings(img_rgb)
                
                if encodings:
                    encode_list.append(encodings[0])
                else:
                    print(f"Warning: No face detected in image {idx}")
            except Exception as e:
                print(f"Error encoding image {idx}: {e}")
        
        return encode_list
    
    def mark_attendance(self, identifier):
        """Mark attendance for a student - Optimized logic"""
        identifier = identifier.upper()
        current_time = datetime.now()
        current_time_str = current_time.strftime('%H:%M:%S')
        
        if not self.current_table:
            return False
        
        # Update or initialize student status
        if identifier not in self.student_status:
            self.student_status[identifier] = {
                'last_seen': current_time,
                'status': 'Present',
                'timer_start': None
            }
            self.attendance_count += 1  # Increment count for new attendance
        else:
            # Student seen again - reset everything
            self.student_status[identifier] = {
                'last_seen': current_time,
                'status': 'Present',
                'timer_start': None
            }
        
        # Update database
        success = self.db_manager.update_attendance(
            self.current_table,
            identifier,
            'Present',
            current_time_str,
            None,  # Clear absence timer
            self.search_mode
        )
        
        if success:
            print(f"✓ Marked {identifier} as Present at {current_time_str}")
        
        return success
    
    def check_absence_continuously(self):
        """Background thread to check student absence status - Optimized logic"""
        while not self.stop_event.is_set():
            if not self.current_table:
                sleep(self.config.ABSENCE_CHECK_INTERVAL)
                continue
            
            current_time = datetime.now()
            
            for identifier, info in list(self.student_status.items()):
                last_seen = info['last_seen']
                current_status = info['status']
                timer_start = info['timer_start']
                
                time_since_seen = current_time - last_seen
                
                # Only process if student was Present
                if current_status == 'Present':
                    # Start absence detection after delay
                    if time_since_seen >= self.config.ABSENCE_DETECTION_DELAY:
                        # Start timer if not started
                        if timer_start is None:
                            timer_start = current_time
                            self.student_status[identifier]['timer_start'] = timer_start
                            print(f"⏱ Started absence timer for {identifier}")
                        
                        time_in_absence = current_time - timer_start
                        
                        # Check for permanent absence
                        if time_in_absence >= self.config.PERMANENT_ABSENT_THRESHOLD:
                            self.student_status[identifier]['status'] = 'Permanently Absent'
                            self.db_manager.update_attendance(
                                self.current_table,
                                identifier,
                                'Permanently Absent',
                                last_seen.strftime('%H:%M:%S'),
                                timer_start.strftime('%H:%M:%S'),
                                self.search_mode
                            )
                            print(f"❌ {identifier} marked as Permanently Absent")
                        
                        # Check for temporary absence
                        elif time_in_absence >= self.config.TEMPORARY_ABSENT_THRESHOLD:
                            if current_status != 'Temporary Absent':
                                self.student_status[identifier]['status'] = 'Temporary Absent'
                                self.db_manager.update_attendance(
                                    self.current_table,
                                    identifier,
                                    'Temporary Absent',
                                    last_seen.strftime('%H:%M:%S'),
                                    timer_start.strftime('%H:%M:%S'),
                                    self.search_mode
                                )
                                print(f"⚠ {identifier} marked as Temporary Absent")
            
            sleep(self.config.ABSENCE_CHECK_INTERVAL)
    
    def draw_header(self, img, session_name, attendance_count, total_students, current_faces):
        """Draw header with session name, attendance count, and current faces detected"""
        # Create a dark overlay for header
        overlay = img.copy()
        cv2.rectangle(overlay, (0, 0), (img.shape[1], 100), (0, 0, 0), -1)
        cv2.addWeighted(overlay, 0.7, img, 0.3, 0, img)
        
        # Session name
        session_text = f"Session: {session_name}" if session_name else "Session: Not Active"
        cv2.putText(img, session_text, (20, 30), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        
        # Total Attendance count (unique students marked)
        attendance_text = f"Total Attendance: {attendance_count}/{total_students}"
        attendance_color = (0, 255, 0) if attendance_count > 0 else (255, 255, 255)
        cv2.putText(img, attendance_text, (20, 60), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, attendance_color, 2)
        
        # Current faces detected (real-time)
        faces_text = f"Current Faces: {current_faces}"
        faces_color = (0, 255, 255) if current_faces > 0 else (150, 150, 150)
        cv2.putText(img, faces_text, (20, 90), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, faces_color, 2)
    
    def draw_face_box(self, img, face_loc, name, color):
        """Draw bounding box and name on face"""
        y1, x2, y2, x1 = [coord * 4 for coord in face_loc]
        
        cv2.rectangle(img, (x1, y1), (x2, y2), color, 2)
        
        font_scale = 0.7
        thickness = 2
        (text_width, text_height), _ = cv2.getTextSize(name, cv2.FONT_HERSHEY_SIMPLEX, font_scale, thickness)
        
        cv2.rectangle(img, (x1, y2 - text_height - 10), (x2, y2), color, cv2.FILLED)
        cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_SIMPLEX, font_scale, (255, 255, 255), thickness)
    
    def run(self):
        """Main execution loop"""
        # Start absence checking thread
        absence_thread = threading.Thread(target=self.check_absence_continuously)
        absence_thread.daemon = True
        absence_thread.start()
        
        # Initialize webcam
        cap = cv2.VideoCapture(0)
        
        if not cap.isOpened():
            print("Error: Could not open camera")
            return
        
        print("Camera opened successfully. Press 'q' to quit.")
        
        try:
            while True:
                ret, frame = cap.read()
                if not ret:
                    print("Failed to capture frame")
                    break
                
                # Resize for faster processing
                small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
                rgb_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
                
                # Detect faces
                face_locations = face_recognition.face_locations(rgb_frame)
                face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
                
                # Count current faces detected (for dynamic display)
                current_faces_count = len(face_locations)
                
                # Get current session
                session = get_current_session()
                
                if session != self.current_session:
                    print(f"\n📚 Current Session: {session}")
                    self.current_session = session
                    self.last_recognized_faces = {}
                    self.student_status = {}
                    self.attendance_count = 0  # Reset attendance count for new session
                    
                    if session:
                        # Get or create table for this session
                        table_name = self.db_manager.get_current_lecture_table(session)
                        
                        if not table_name:
                            table_name = self.db_manager.create_lecture_table(
                                session,
                                self.template_headers,
                                self.template_data
                            )
                        
                        self.current_table = table_name
                        print(f"Using table: {table_name}")
                
                # Draw header with session, attendance count, and current faces
                self.draw_header(frame, session, self.attendance_count, self.total_students, current_faces_count)
                
                if not session:
                    cv2.imshow("Attendance System", frame)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break
                    continue
                
                # Process each face
                for face_encoding, face_loc in zip(face_encodings, face_locations):
                    matches = face_recognition.compare_faces(self.known_encodings, face_encoding, tolerance=0.6)
                    face_distances = face_recognition.face_distance(self.known_encodings, face_encoding)
                    
                    if len(face_distances) > 0:
                        best_match_idx = np.argmin(face_distances)
                        
                        if matches[best_match_idx]:
                            name = self.class_names[best_match_idx].upper()
                            color = (0, 255, 0)
                            
                            # Mark attendance
                            self.mark_attendance(name)
                        else:
                            name = "UNKNOWN"
                            color = (0, 0, 255)
                    else:
                        name = "UNKNOWN"
                        color = (0, 0, 255)
                    
                    self.draw_face_box(frame, face_loc, name, color)
                
                cv2.imshow("Attendance System", frame)
                
                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break
        
        except KeyboardInterrupt:
            print("\n⚠ Program interrupted by user")
        except Exception as e:
            print(f"❌ Error: {e}")
            import traceback
            traceback.print_exc()
        finally:
            self.cleanup(cap)
    
    def cleanup(self, cap):
        """Clean up resources"""
        self.stop_event.set()
        
        cap.release()
        cv2.destroyAllWindows()
        
        self.db_manager.close()
        print("✓ System shutdown complete")

def main():
    """Main entry point"""
    print("=" * 60)
    print("Face Recognition Attendance System with SQL Storage")
    print("=" * 60)
    print("\nSelect mode:")
    print("1. Mark attendance by Name")
    print("2. Mark attendance by Roll No")
    
    try:
        mode = int(input("\nEnter choice (1 or 2): "))
        if mode not in [1, 2]:
            print("Invalid choice. Defaulting to Name mode.")
            mode = 1
    except ValueError:
        print("Invalid input. Defaulting to Name mode.")
        mode = 1
    
    system = AttendanceSystem(mode=mode)
    system.run()

if __name__ == "__main__":
    main()