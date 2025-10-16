"""
Enhanced Attendance System - Final Version
- Robust duration tracking (manual + face recognition)
- Compact UI with confirmation flow
- New schema: Daily document with embedded sessions
"""

from flask import Flask, jsonify, request, render_template_string, Response, send_file
from flask_cors import CORS
from pymongo import MongoClient
from datetime import datetime, timedelta
import os
import threading
import cv2
import numpy as np
import face_recognition
import sys
from time import sleep
from threading import Lock, Event
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import io
from bson.objectid import ObjectId

sys.path.append(os.path.abspath('../'))
try:
    from Excel_Format import get_current_session
except ImportError:
    def get_current_session():
        current_hour = datetime.now().hour
        if 9 <= current_hour < 10:
            return "Session 1"
        elif 10 <= current_hour < 11:
            return "Session 2"
        elif 11 <= current_hour < 12:
            return "Session 3"
        elif 12 <= current_hour < 13:
            return "Session 4"
        elif 13 <= current_hour < 14:
            return "Session 5"
        elif 14 <= current_hour < 15:
            return "Session 6"
        elif 15 <= current_hour < 16:
            return "Session 7"
        else:
            return "Session 8"

app = Flask(__name__)
CORS(app)

MONGODB_CONFIG = {
    'host': 'localhost',
    'port': 27017,
    'database': 'Attendance_system'
}

TEMPLATE_FILE = 'Book2.xlsx'

attendance_system = None
camera_running = False
camera_lock = Lock()


class AttendanceConfig:
    TEMPLATE_FILE = 'Book2.xlsx'
    ABSENCE_DETECTION_DELAY = 5
    TEMPORARY_ABSENT_THRESHOLD = 10
    PERMANENT_ABSENT_THRESHOLD = 15
    ABSENCE_CHECK_INTERVAL = 2
    MODE_NAME = 1
    MODE_ROLL_NO = 2


class DatabaseManager:
    """Handles MongoDB with NEW SCHEMA: Daily docs with embedded sessions"""
    
    def __init__(self, mongodb_config):
        self.mongodb_config = mongodb_config
        self.client = None
        self.db = None
        self.lock = Lock()
        self._initialize_db()
    
    def _get_connection(self):
        try:
            if self.client is None:
                self.client = MongoClient(
                    self.mongodb_config['host'], 
                    self.mongodb_config['port']
                )
                self.db = self.client[self.mongodb_config['database']]
            return self.db
        except Exception as e:
            print(f"Error connecting to MongoDB: {e}")
            raise
    
    def _initialize_db(self):
        try:
            self.db = self._get_connection()
            # Create index on date + class_id for fast lookups
            self.db.daily_attendance.create_index([('date', 1), ('class_id', 1)], unique=True)
            print("✓ MongoDB initialized successfully")
        except Exception as e:
            print(f"Error initializing database: {e}")
            raise
    
    def load_template_from_excel(self, excel_file, sheet_name='Sheet1'):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_file, data_only=True)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found")
            sheet = wb[sheet_name]
            data = []
            headers = None
            for row in sheet.iter_rows(values_only=True):
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                row_data = [cell if cell is not None else '' for cell in row]
                if headers is None:
                    headers = [str(h).strip() for h in row_data]
                else:
                    data.append(row_data)
            wb.close()
            print(f"✓ Loaded template: {len(headers)} columns, {len(data)} rows")
            return headers, data
        except Exception as e:
            print(f"Error loading template: {e}")
            raise
    
    def get_or_create_daily_doc(self, date_str, class_id, year, department, classroom, teacher_name):
        """Get or create daily attendance document"""
        with self.lock:
            try:
                db = self._get_connection()
                collection = db.daily_attendance
                
                # Try to find existing document
                doc = collection.find_one({
                    'date': date_str,
                    'class_id': class_id
                })
                
                if doc:
                    return doc
                
                # Create new document with structure
                new_doc = {
                    'date': date_str,
                    'class_id': class_id,
                    'year': year,
                    'department': department,
                    'classroom': classroom,
                    'teacher_name': teacher_name,
                    'created_at': datetime.now(),
                    'sessions': {}
                }
                
                result = collection.insert_one(new_doc)
                new_doc['_id'] = result.inserted_id
                print(f"✓ Created daily document for {class_id} on {date_str}")
                return new_doc
                
            except Exception as e:
                print(f"Error getting/creating daily doc: {e}")
                raise
    
    def initialize_session_students(self, date_str, class_id, session_name, template_data, template_headers):
        """Initialize students for a session"""
        with self.lock:
            try:
                db = self._get_connection()
                collection = db.daily_attendance
                
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # Build students dictionary
                students_dict = {}
                for row in template_data:
                    doc_data = {}
                    for i, header in enumerate(template_headers):
                        if i < len(row):
                            value = row[i]
                            doc_data[header] = str(value) if value not in (None, '') else ''
                        else:
                            doc_data[header] = ''
                    
                    roll_no = doc_data.get('Roll No', '')
                    if not roll_no:
                        continue
                    
                    students_dict[roll_no] = {
                        'sr_no': doc_data.get('Sr. No', ''),
                        'roll_no': roll_no,
                        'prn_no': doc_data.get('PRN No.', ''),
                        'name': doc_data.get('Name', ''),
                        'status': 'Absent',
                        'timestamps': {
                            'first_seen': None,
                            'last_seen': None,
                            'present_timer_start': None,
                            'absence_timer_start': None,
                            'temp_absent_time': None,
                            'perm_absent_time': None,
                            'last_updated': current_time
                        },
                        'durations': {
                            'total_present_seconds': 0,
                            'total_absent_seconds': 0,
                            'total_present_human': '0 sec',
                            'total_absent_human': '0 sec'
                        },
                        'flags': {
                            'manual_override': False,
                            'is_temp_absent': False,
                            'is_perm_absent': False
                        }
                    }
                
                # Update document with session
                session_obj = {
                    'start_time': datetime.now().strftime('%H:%M:%S'),
                    'end_time': None,
                    'students': students_dict
                }
                
                collection.update_one(
                    {'date': date_str, 'class_id': class_id},
                    {'$set': {f'sessions.{session_name}': session_obj}}
                )
                
                print(f"✓ Initialized {len(students_dict)} students for {session_name}")
                return True
                
            except Exception as e:
                print(f"Error initializing session students: {e}")
                import traceback
                traceback.print_exc()
                return False
    
    def update_student_attendance(self, date_str, class_id, session_name, roll_no, status, manual=False):
        """Update student attendance with ROBUST duration tracking"""
        with self.lock:
            try:
                db = self._get_connection()
                collection = db.daily_attendance
                
                # Get current document
                doc = collection.find_one({'date': date_str, 'class_id': class_id})
                if not doc or session_name not in doc.get('sessions', {}):
                    return False
                
                student_path = f'sessions.{session_name}.students.{roll_no}'
                student = doc['sessions'][session_name]['students'].get(roll_no)
                
                if not student:
                    return False
                
                current_time = datetime.now()
                current_time_str = current_time.strftime('%Y-%m-%d %H:%M:%S')
                
                # Get previous state
                prev_status = student.get('status', 'Absent')
                present_start = student['timestamps'].get('present_timer_start')
                absent_start = student['timestamps'].get('absence_timer_start')
                total_present = student['durations'].get('total_present_seconds', 0)
                total_absent = student['durations'].get('total_absent_seconds', 0)
                
                update_fields = {}
                
                # ROBUST DURATION LOGIC
                # Finalize previous duration before changing status
                if prev_status == 'Present' and status != 'Present':
                    # Was present, now becoming absent
                    if present_start:
                        try:
                            start_dt = datetime.strptime(present_start, '%Y-%m-%d %H:%M:%S')
                            duration = (current_time - start_dt).total_seconds()
                            total_present += duration
                            print(f"  Added {duration:.0f}s to present time (total: {total_present:.0f}s)")
                        except:
                            pass
                    update_fields[f'{student_path}.timestamps.present_timer_start'] = None
                    update_fields[f'{student_path}.timestamps.absence_timer_start'] = current_time_str
                    
                elif prev_status != 'Present' and status == 'Present':
                    # Was absent, now becoming present
                    if absent_start:
                        try:
                            start_dt = datetime.strptime(absent_start, '%Y-%m-%d %H:%M:%S')
                            duration = (current_time - start_dt).total_seconds()
                            total_absent += duration
                            print(f"  Added {duration:.0f}s to absent time (total: {total_absent:.0f}s)")
                        except:
                            pass
                    update_fields[f'{student_path}.timestamps.absence_timer_start'] = None
                    update_fields[f'{student_path}.timestamps.present_timer_start'] = current_time_str
                    
                elif status == 'Present' and not present_start:
                    # Starting present tracking
                    update_fields[f'{student_path}.timestamps.present_timer_start'] = current_time_str
                    
                elif status != 'Present' and not absent_start:
                    # Starting absent tracking
                    update_fields[f'{student_path}.timestamps.absence_timer_start'] = current_time_str
                
                # Update basic fields
                update_fields[f'{student_path}.status'] = status
                update_fields[f'{student_path}.timestamps.last_updated'] = current_time_str
                update_fields[f'{student_path}.flags.manual_override'] = manual
                
                # Handle first_seen
                if student['timestamps']['first_seen'] is None and status == 'Present':
                    update_fields[f'{student_path}.timestamps.first_seen'] = current_time_str
                
                # Handle last_seen
                if status == 'Present':
                    update_fields[f'{student_path}.timestamps.last_seen'] = current_time_str
                
                # Update durations
                update_fields[f'{student_path}.durations.total_present_seconds'] = total_present
                update_fields[f'{student_path}.durations.total_absent_seconds'] = total_absent
                update_fields[f'{student_path}.durations.total_present_human'] = self._format_duration(total_present)
                update_fields[f'{student_path}.durations.total_absent_human'] = self._format_duration(total_absent)
                
                # Handle status flags
                if status == 'Temporary Absent':
                    update_fields[f'{student_path}.timestamps.temp_absent_time'] = current_time_str
                    update_fields[f'{student_path}.flags.is_temp_absent'] = True
                elif status == 'Permanently Absent':
                    update_fields[f'{student_path}.timestamps.perm_absent_time'] = current_time_str
                    update_fields[f'{student_path}.flags.is_perm_absent'] = True
                elif status == 'Present':
                    update_fields[f'{student_path}.flags.is_temp_absent'] = False
                    update_fields[f'{student_path}.flags.is_perm_absent'] = False
                
                # Apply update
                result = collection.update_one(
                    {'date': date_str, 'class_id': class_id},
                    {'$set': update_fields}
                )
                
                return result.modified_count > 0
                
            except Exception as e:
                print(f"Error updating student attendance: {e}")
                import traceback
                traceback.print_exc()
                return False
    
    def _format_duration(self, seconds):
        if seconds < 60:
            return f"{int(seconds)} sec"
        elif seconds < 3600:
            minutes = int(seconds / 60)
            secs = int(seconds % 60)
            return f"{minutes} min {secs} sec"
        else:
            hours = int(seconds / 3600)
            minutes = int((seconds % 3600) / 60)
            return f"{hours} hr {minutes} min"
    
    def get_session_attendance(self, date_str, class_id, session_name):
        """Get attendance for a specific session"""
        with self.lock:
            try:
                db = self._get_connection()
                doc = db.daily_attendance.find_one({'date': date_str, 'class_id': class_id})
                
                if not doc or session_name not in doc.get('sessions', {}):
                    return []
                
                students = doc['sessions'][session_name]['students']
                return list(students.values())
                
            except Exception as e:
                print(f"Error getting session attendance: {e}")
                return []
    
    def get_session_summary(self, date_str, class_id, session_name):
        """Get summary statistics for a session"""
        with self.lock:
            try:
                db = self._get_connection()
                doc = db.daily_attendance.find_one({'date': date_str, 'class_id': class_id})
                
                if not doc or session_name not in doc.get('sessions', {}):
                    return {}
                
                students = doc['sessions'][session_name]['students']
                total = len(students)
                present = sum(1 for s in students.values() if s['status'] == 'Present')
                temp_absent = sum(1 for s in students.values() if s['status'] == 'Temporary Absent')
                perm_absent = sum(1 for s in students.values() if s['status'] == 'Permanently Absent')
                absent = sum(1 for s in students.values() if s['status'] == 'Absent')
                
                return {
                    'total': total,
                    'present': present,
                    'temporary_absent': temp_absent,
                    'permanently_absent': perm_absent,
                    'absent': absent,
                    'attendance_percentage': round((present / total * 100), 2) if total > 0 else 0
                }
                
            except Exception as e:
                print(f"Error getting summary: {e}")
                return {}
    
    def clear_session_data(self, date_str, class_id, session_name):
        """Clear all attendance data for a session"""
        with self.lock:
            try:
                db = self._get_connection()
                doc = db.daily_attendance.find_one({'date': date_str, 'class_id': class_id})
                
                if not doc or session_name not in doc.get('sessions', {}):
                    return 0
                
                students = doc['sessions'][session_name]['students']
                current_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                update_fields = {}
                for roll_no in students.keys():
                    prefix = f'sessions.{session_name}.students.{roll_no}'
                    update_fields[f'{prefix}.status'] = 'Absent'
                    update_fields[f'{prefix}.timestamps.first_seen'] = None
                    update_fields[f'{prefix}.timestamps.last_seen'] = None
                    update_fields[f'{prefix}.timestamps.present_timer_start'] = None
                    update_fields[f'{prefix}.timestamps.absence_timer_start'] = None
                    update_fields[f'{prefix}.timestamps.temp_absent_time'] = None
                    update_fields[f'{prefix}.timestamps.perm_absent_time'] = None
                    update_fields[f'{prefix}.timestamps.last_updated'] = current_time_str
                    update_fields[f'{prefix}.durations.total_present_seconds'] = 0
                    update_fields[f'{prefix}.durations.total_absent_seconds'] = 0
                    update_fields[f'{prefix}.durations.total_present_human'] = '0 sec'
                    update_fields[f'{prefix}.durations.total_absent_human'] = '0 sec'
                    update_fields[f'{prefix}.flags.manual_override'] = False
                    update_fields[f'{prefix}.flags.is_temp_absent'] = False
                    update_fields[f'{prefix}.flags.is_perm_absent'] = False
                
                result = db.daily_attendance.update_one(
                    {'date': date_str, 'class_id': class_id},
                    {'$set': update_fields}
                )
                
                print(f"✓ Cleared {len(students)} students in {session_name}")
                return len(students)
                
            except Exception as e:
                print(f"Error clearing session: {e}")
                return 0
    
    def close(self):
        if self.client:
            self.client.close()


class AttendanceSystem:
    def __init__(self, mode, year, department, classroom, teacher_name, class_id):
        self.mode = mode
        self.config = AttendanceConfig()
        self.student_status = {}
        self.stop_event = Event()
        self.attendance_count = 0
        self.total_students = 0
        self.search_mode = 'roll' if mode == AttendanceConfig.MODE_ROLL_NO else 'name'
        self.current_faces_count = 0
        self.year = year
        self.department = department
        self.classroom = classroom
        self.teacher_name = teacher_name
        self.class_id = class_id
        
        self.db_manager = DatabaseManager(MONGODB_CONFIG)
        self.template_headers, self.template_data = self.db_manager.load_template_from_excel(
            self.config.TEMPLATE_FILE, 'Sheet1'
        )
        self.total_students = len(self.template_data)
        
        self.current_session = None
        self.current_date = None
        self.class_names, self.known_encodings = self._load_training_data()
    
    def _load_training_data(self):
        path = '../Training_images/Name' if self.mode == self.config.MODE_NAME else '../Training_images/Roll No.'
        if not os.path.exists(path):
            raise FileNotFoundError(f"Training images directory not found: {path}")
        images = []
        class_names = []
        image_files = [f for f in os.listdir(path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
        print(f"Loading {len(image_files)} training images from {path}")
        for filename in image_files:
            img_path = os.path.join(path, filename)
            img = cv2.imread(img_path)
            if img is None:
                continue
            images.append(img)
            class_names.append(os.path.splitext(filename)[0])
        print(f"✓ Loaded {len(class_names)} training images")
        encodings = self._find_encodings(images)
        return class_names, encodings
    
    def _find_encodings(self, images):
        encode_list = []
        for img in images:
            try:
                img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                encodings = face_recognition.face_encodings(img_rgb)
                if encodings:
                    encode_list.append(encodings[0])
            except Exception as e:
                pass
        return encode_list
    
    def mark_attendance(self, identifier):
        identifier = identifier.upper()
        current_time = datetime.now()
        
        if not self.current_session or not self.current_date:
            return False
        
        if identifier not in self.student_status:
            self.student_status[identifier] = {
                'last_seen': current_time,
                'status': 'Present',
                'timer_start': None
            }
            self.attendance_count += 1
        else:
            self.student_status[identifier]['last_seen'] = current_time
            self.student_status[identifier]['status'] = 'Present'
            self.student_status[identifier]['timer_start'] = None
        
        success = self.db_manager.update_student_attendance(
            self.current_date,
            self.class_id,
            self.current_session,
            identifier,
            'Present',
            manual=False
        )
        
        if success:
            print(f"✓ Marked {identifier} as Present")
        return success
    
    def check_absence_continuously(self):
        while not self.stop_event.is_set():
            if not self.current_session:
                sleep(self.config.ABSENCE_CHECK_INTERVAL)
                continue
            
            current_time = datetime.now()
            for identifier, info in list(self.student_status.items()):
                last_seen = info['last_seen']
                current_status = info['status']
                timer_start = info['timer_start']
                
                time_since_seen = current_time - last_seen
                
                if current_status == 'Present':
                    if time_since_seen >= timedelta(seconds=self.config.ABSENCE_DETECTION_DELAY):
                        if timer_start is None:
                            timer_start = current_time
                            self.student_status[identifier]['timer_start'] = timer_start
                            print(f"⏱ Started absence timer for {identifier}")
                        
                        time_in_absence = current_time - timer_start
                        
                        if time_in_absence >= timedelta(seconds=self.config.PERMANENT_ABSENT_THRESHOLD):
                            self.student_status[identifier]['status'] = 'Permanently Absent'
                            self.db_manager.update_student_attendance(
                                self.current_date,
                                self.class_id,
                                self.current_session,
                                identifier,
                                'Permanently Absent'
                            )
                            print(f"❌ {identifier} marked as Permanently Absent")
                        elif time_in_absence >= timedelta(seconds=self.config.TEMPORARY_ABSENT_THRESHOLD):
                            if current_status != 'Temporary Absent':
                                self.student_status[identifier]['status'] = 'Temporary Absent'
                                self.db_manager.update_student_attendance(
                                    self.current_date,
                                    self.class_id,
                                    self.current_session,
                                    identifier,
                                    'Temporary Absent'
                                )
                                print(f"⚠ {identifier} marked as Temporary Absent")
            
            sleep(self.config.ABSENCE_CHECK_INTERVAL)
    
    def process_frame(self, frame):
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
        
        self.current_faces_count = len(face_locations)
        session = get_current_session()
        date_str = datetime.now().strftime('%Y-%m-%d')
        
        if session != self.current_session or date_str != self.current_date:
            print(f"\n📚 Session Changed: {self.current_session} -> {session}")
            
            self.current_session = session
            self.current_date = date_str
            self.student_status = {}
            self.attendance_count = 0
            
            # Get or create daily doc
            self.db_manager.get_or_create_daily_doc(
                date_str, self.class_id, self.year, self.department, 
                self.classroom, self.teacher_name
            )
            
            # Initialize session students
            self.db_manager.initialize_session_students(
                date_str, self.class_id, session, 
                self.template_data, self.template_headers
            )
        
        for face_encoding, face_loc in zip(face_encodings, face_locations):
            matches = face_recognition.compare_faces(self.known_encodings, face_encoding, tolerance=0.6)
            face_distances = face_recognition.face_distance(self.known_encodings, face_encoding)
            
            if len(face_distances) > 0:
                best_match_idx = np.argmin(face_distances)
                if matches[best_match_idx]:
                    name = self.class_names[best_match_idx].upper()
                    color = (0, 255, 0)
                    self.mark_attendance(name)
                else:
                    name = "UNKNOWN"
                    color = (0, 0, 255)
            else:
                name = "UNKNOWN"
                color = (0, 0, 255)
            
            y1, x2, y2, x1 = [coord * 4 for coord in face_loc]
            cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
            cv2.rectangle(frame, (x1, y2 - 35), (x2, y2), color, cv2.FILLED)
            cv2.putText(frame, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        
        # Overlay
        overlay = frame.copy()
        cv2.rectangle(overlay, (0, 0), (frame.shape[1], 80), (0, 0, 0), -1)
        cv2.addWeighted(overlay, 0.7, frame, 0.3, 0, frame)
        
        cv2.putText(frame, f"Session: {session}", (20, 25), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
        cv2.putText(frame, f"Attendance: {self.attendance_count}/{self.total_students}", (20, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0) if self.attendance_count > 0 else (255, 255, 255), 2)
        cv2.putText(frame, f"Faces: {self.current_faces_count}", (20, 70), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)
        
        return frame
    
    def stop(self):
        self.stop_event.set()


# ============= FLASK ROUTES =============

@app.route('/api/health', methods=['GET'])
def health_check():
    global attendance_system, camera_running
    try:
        client = MongoClient(MONGODB_CONFIG['host'], MONGODB_CONFIG['port'], serverSelectionTimeoutMS=2000)
        client.server_info()
        db_connected = True
        client.close()
    except:
        db_connected = False
    return jsonify({
        'status': 'healthy' if db_connected else 'degraded',
        'database': 'connected' if db_connected else 'disconnected',
        'camera_status': 'running' if camera_running else 'stopped',
        'system_initialized': attendance_system is not None,
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/camera/start', methods=['POST'])
def start_camera():
    global attendance_system, camera_running
    try:
        data = request.get_json()
        mode = data.get('mode', 1)
        year = data.get('year', '')
        department = data.get('department', '')
        classroom = data.get('classroom', '')
        teacher_name = data.get('teacher_name', '')
        
        if not all([year, department, classroom, teacher_name]):
            return jsonify({'success': False, 'message': 'All fields are required'}), 400
        
        if camera_running:
            return jsonify({'success': False, 'message': 'Camera is already running'}), 400
        
        class_id = f"{year}_{department}_{classroom}"
        
        attendance_system = AttendanceSystem(
            mode=mode,
            year=year,
            department=department,
            classroom=classroom,
            teacher_name=teacher_name,
            class_id=class_id
        )
        
        absence_thread = threading.Thread(target=attendance_system.check_absence_continuously)
        absence_thread.daemon = True
        absence_thread.start()
        
        camera_running = True
        
        return jsonify({
            'success': True,
            'message': 'Camera started successfully',
            'class_id': class_id
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/camera/stop', methods=['POST'])
def stop_camera():
    global attendance_system, camera_running
    try:
        if not camera_running:
            return jsonify({'success': False, 'message': 'Camera is not running'}), 400
        
        camera_running = False
        if attendance_system:
            attendance_system.stop()
        
        return jsonify({'success': True, 'message': 'Camera stopped successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/camera/status', methods=['GET'])
def camera_status():
    global camera_running, attendance_system
    status = {
        'running': camera_running,
        'attendance_count': 0,
        'current_faces': 0,
        'current_session': None
    }
    if attendance_system:
        status['attendance_count'] = attendance_system.attendance_count
        status['current_faces'] = attendance_system.current_faces_count
        status['current_session'] = attendance_system.current_session
    return jsonify(status)

def generate_frames():
    global attendance_system, camera_running
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("Error: Could not open camera")
        return
    try:
        while camera_running:
            ret, frame = cap.read()
            if not ret:
                break
            if attendance_system:
                frame = attendance_system.process_frame(frame)
            ret, buffer = cv2.imencode('.jpg', frame)
            if not ret:
                continue
            frame = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
    except Exception as e:
        print(f"Error in video stream: {e}")
    finally:
        cap.release()

@app.route('/api/video_feed')
def video_feed():
    if not camera_running:
        return jsonify({'error': 'Camera not running'}), 400
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/api/current-session', methods=['GET'])
def get_current_session_data():
    try:
        if attendance_system:
            date_str = attendance_system.current_date or datetime.now().strftime('%Y-%m-%d')
            session_name = attendance_system.current_session or get_current_session()
            
            attendance = attendance_system.db_manager.get_session_attendance(
                date_str, attendance_system.class_id, session_name
            )
            summary = attendance_system.db_manager.get_session_summary(
                date_str, attendance_system.class_id, session_name
            )
            
            return jsonify({
                'success': True,
                'active': True,
                'date': date_str,
                'class_id': attendance_system.class_id,
                'session_name': session_name,
                'summary': summary,
                'attendance': attendance
            })
        else:
            return jsonify({'success': True, 'active': False, 'message': 'No active session'})
    except Exception as e:
        print(f"Error in get_current_session_data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/attendance/update', methods=['POST'])
def update_attendance_manual():
    """Manual attendance update endpoint"""
    try:
        data = request.get_json()
        
        date_str = data.get('date')
        class_id = data.get('class_id')
        session_name = data.get('session_name')
        roll_no = data.get('roll_no')
        status = data.get('status')
        
        if not all([date_str, class_id, session_name, roll_no, status]):
            return jsonify({'success': False, 'error': 'Missing required fields'}), 400
        
        db_manager = DatabaseManager(MONGODB_CONFIG)
        success = db_manager.update_student_attendance(
            date_str, class_id, session_name, roll_no, status, manual=True
        )
        db_manager.close()
        
        if success:
            return jsonify({'success': True, 'message': f'Successfully marked as {status}'})
        else:
            return jsonify({'success': False, 'error': 'Failed to update'}), 500
            
    except Exception as e:
        print(f"Error in update_attendance_manual: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/attendance/clear', methods=['POST'])
def clear_attendance_data():
    """Clear all attendance data for current session"""
    try:
        data = request.get_json()
        date_str = data.get('date')
        class_id = data.get('class_id')
        session_name = data.get('session_name')
        
        if not all([date_str, class_id, session_name]):
            return jsonify({'success': False, 'error': 'Missing required fields'}), 400
        
        if attendance_system:
            attendance_system.student_status = {}
            attendance_system.attendance_count = 0
            count = attendance_system.db_manager.clear_session_data(date_str, class_id, session_name)
        else:
            db_manager = DatabaseManager(MONGODB_CONFIG)
            count = db_manager.clear_session_data(date_str, class_id, session_name)
            db_manager.close()
        
        return jsonify({
            'success': True,
            'message': f'Cleared data for {count} students',
            'count': count
        })
        
    except Exception as e:
        print(f"Error in clear_attendance_data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

from template_copy import HTML_TEMPLATE, REPORTS_HTML_TEMPLATE, STUDENT_HTML_TEMPLATE
if __name__ == '__main__':
    print("=" * 70)
    print("Enhanced Attendance System - FINAL VERSION")
    print("=" * 70)
    print("\nKey Features:")
    print("   ✓ ROBUST duration tracking (handles manual + face recognition)")
    print("   ✓ COMPACT UI with confirmation flow")
    print("   ✓ NEW SCHEMA: Daily documents with embedded sessions")
    print("   ✓ Space-efficient button layout (single line)")
    print("   ✓ No wasted space in video section")
    print("\nSchema Structure:")
    print("   {")
    print("     date: '2025-10-06',")
    print("     class_id: '2025_CSE_301',")
    print("     sessions: {")
    print("       'Session 1': {")
    print("         students: { 'ROLL001': {...}, ... }")
    print("       }")
    print("     }")
    print("   }")
    print("\nStarting Server...")
    print(f"Database: MongoDB ({MONGODB_CONFIG['host']}:{MONGODB_CONFIG['port']})")
    print(f"Server URL: http://localhost:5000")
    print("\n" + "=" * 70 + "\n")
    
    app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)